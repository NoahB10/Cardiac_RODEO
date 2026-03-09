"""
Create PNG plots for Cumulative Feature Importance with threshold lines.
Drugs that FAIL (negative for condition) are plotted with X markers.
Drugs that PASS (positive for condition) are plotted with circle markers.
Also creates Excel files with embedded charts.
"""
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import json
from pathlib import Path

# Register Helvetica fonts from local fonts folder
_font_dir = Path(__file__).parent / 'fonts'
if _font_dir.exists():
    for _font_file in _font_dir.glob('*.ttf'):
        fm.fontManager.addfont(str(_font_file))
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Helvetica', 'Arial', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils.dataframe import dataframe_to_rows

# Paths
PROJECT_ROOT = Path(__file__).parent
DATA_PATH = PROJECT_ROOT / 'Output' / 'Cumulative_Plot_Data' / 'cumulative_feature_importance.xlsx'
THRESHOLD_PATH = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data' / 'prediction_thresholds.json'
CLASSIFICATION_PATH = PROJECT_ROOT / 'Cleaned_Data' / 'drug_classification.csv'
OUTPUT_DIR = PROJECT_ROOT / 'Output' / 'Excel_Figures'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Load thresholds
with open(THRESHOLD_PATH, 'r') as f:
    thresholds = json.load(f)

# Load drug classification (ground truth)
drug_class = pd.read_csv(CLASSIFICATION_PATH)
drug_class.set_index('Drug', inplace=True)

# Threshold mapping for each sheet
THRESHOLD_MAP = {
    'Arrhythmia': thresholds['Arrhythmia_threshold_pct'],
    'Heart Damage': thresholds['Heart_Damage_threshold_pct'],
    'Concern Binary': thresholds['Concern_Binary_threshold_pct'],
    'Concern No': thresholds['Concern_thresholds_pct']['No Concern'],
    'Concern Less': thresholds['Concern_thresholds_pct']['Less Concern'],
    'Concern Most': thresholds['Concern_thresholds_pct']['Most Concern'],
}

# Colors for plots
PLOT_COLORS = {
    'Arrhythmia': 'tab:blue',
    'Heart Damage': 'tab:red',
    'Concern Binary': 'tab:purple',
    'Concern No': 'tab:green',
    'Concern Less': 'tab:orange',
    'Concern Most': 'tab:red',
}


def get_passing_drugs(sheet_name):
    """
    Return set of drugs that are POSITIVE for the condition (pass).
    Drugs not in this set are considered to FAIL (negative).
    """
    if sheet_name == 'Arrhythmia':
        # Arrhythmia = True
        return set(drug_class[drug_class['Arrhythmia'] == True].index)
    elif sheet_name == 'Heart Damage':
        # heart_damage = True
        return set(drug_class[drug_class['heart_damage'] == True].index)
    elif sheet_name == 'Concern Binary':
        # Binary: most concern vs others (most = positive)
        return set(drug_class[drug_class['Concern'] == 'most'].index)
    elif sheet_name == 'Concern No':
        # No concern = positive
        return set(drug_class[drug_class['Concern'] == 'no'].index)
    elif sheet_name == 'Concern Less':
        # Less concern = positive
        return set(drug_class[drug_class['Concern'] == 'less'].index)
    elif sheet_name == 'Concern Most':
        # Most concern = positive
        return set(drug_class[drug_class['Concern'] == 'most'].index)
    else:
        return set()


def create_cumulative_plot(df, sheet_name, threshold, output_dir):
    """Create cumulative feature importance plot with threshold line.
    Passing drugs = circles, Failing drugs = X markers.
    """
    fig, ax = plt.subplots(figsize=(14, 8))

    # Get feature labels (first column)
    features = df['Cumulative_Coefficients'].values
    n_features = len(features)

    # Get drug columns (all except first)
    drug_cols = [col for col in df.columns if col != 'Cumulative_Coefficients']

    # Get passing drugs for this target
    passing_drugs = get_passing_drugs(sheet_name)

    # Create rainbow colormap for drugs
    drug_list = sorted(drug_cols)
    cmap = plt.get_cmap('rainbow')
    color_map = {
        drug: cmap(0.0 if len(drug_list) == 1 else i / (len(drug_list) - 1))
        for i, drug in enumerate(drug_list)
    }

    # Plot each drug
    x = np.arange(1, n_features + 1)
    for drug in drug_list:
        y = df[drug].values
        is_passing = drug in passing_drugs

        # Marker style: 'o' for passing, 'x' for failing
        marker = 'o' if is_passing else 'x'
        markersize = 4 if is_passing else 6

        ax.plot(x, y, marker=marker, label=drug, alpha=0.7, linewidth=1.5,
                markersize=markersize, color=color_map[drug])

    # Add threshold line
    ax.axhline(y=threshold, color='red', linestyle='--', linewidth=2.5,
               label=f'Threshold = {threshold:.1f}%')

    # Create x-tick labels (stacked feature names)
    def stacked_label(rank):
        parts = features[rank-1].split(' + ')
        if len(parts) <= 5:
            return '\n'.join(parts)
        else:
            return '\n'.join(parts[:5]) + f'\n... (+{len(parts)-5})'

    xtick_labels = [stacked_label(i) for i in range(1, n_features + 1)]

    # Formatting
    ax.set_xticks(x)
    ax.set_xticklabels(xtick_labels, rotation=45, ha='right', fontsize=7)
    ax.set_yticks(np.linspace(0, 100, 11))
    ax.set_ylim(0, 100)
    ax.set_xlabel('Cumulative Features (ranked by importance)', fontsize=12, fontweight='bold')
    ax.set_ylabel('Predicted Probability (%)', fontsize=12, fontweight='bold')
    ax.set_title(f'Cumulative Feature Importance - {sheet_name}\n(circles = positive class, X = negative class)',
                 fontsize=14, fontweight='bold')
    ax.grid(True, axis='y', alpha=0.3)

    # Legend outside plot
    ax.legend(bbox_to_anchor=(1.02, 1), loc='upper left', ncol=1, fontsize=8)

    plt.tight_layout()

    # Save PNG
    safe_name = sheet_name.replace(' ', '_')
    png_path = output_dir / f'Cumulative_Feature_Importance_{safe_name}.png'
    fig.savefig(png_path, dpi=300, bbox_inches='tight')
    plt.close(fig)

    print(f"Saved PNG: {png_path.name}")
    return png_path


def create_excel_with_chart(df, sheet_name, threshold, output_dir, passing_drugs):
    """Create Excel file with data and embedded line chart."""
    safe_name = sheet_name.replace(' ', '_')
    excel_path = output_dir / f'Cumulative_Feature_Importance_{safe_name}.xlsx'

    # Transpose data for Excel (features as columns, drugs as rows)
    df_transposed = df.set_index('Cumulative_Coefficients').T
    df_transposed.index.name = 'Drug'
    df_transposed.reset_index(inplace=True)

    # Add a "Pass/Fail" column
    df_transposed['Status'] = df_transposed['Drug'].apply(
        lambda d: 'Pass' if d in passing_drugs else 'Fail'
    )

    # Write to Excel
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_transposed.to_excel(writer, sheet_name='Data', index=False)

        # Also write original format for reference
        df.to_excel(writer, sheet_name='Original', index=False)

    # Add chart
    wb = load_workbook(excel_path)
    ws = wb['Original']

    # Create line chart
    chart = LineChart()
    chart.title = f"Cumulative Feature Importance - {sheet_name}"
    chart.style = 10
    chart.y_axis.title = "Predicted Probability (%)"
    chart.x_axis.title = "Feature Rank"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.width = 20
    chart.height = 12

    # Data: columns 2 onwards (B onwards) are drugs
    n_drugs = len(df.columns) - 1
    n_features = len(df)

    # Add each drug as a series
    for i in range(2, n_drugs + 2):  # Excel columns B, C, D, ...
        values = Reference(ws, min_col=i, min_row=1, max_row=n_features + 1)
        chart.add_data(values, titles_from_data=True)

    # Categories (feature ranks)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_features + 1)
    chart.set_categories(cats)

    # Add chart to new sheet
    chart_sheet = wb.create_sheet('Chart')
    chart_sheet.add_chart(chart, 'A1')

    wb.save(excel_path)
    print(f"Saved Excel: {excel_path.name}")
    return excel_path


# Main
print("="*60)
print("CREATING CUMULATIVE FEATURE IMPORTANCE PLOTS")
print("(Circles = Pass/Positive, X = Fail/Negative)")
print("="*60)

# Load Excel
xlsx = pd.ExcelFile(DATA_PATH)

for sheet_name in xlsx.sheet_names:
    print(f"\nProcessing: {sheet_name}")
    df = pd.read_excel(xlsx, sheet_name=sheet_name)
    threshold = THRESHOLD_MAP.get(sheet_name, 50.0)
    passing_drugs = get_passing_drugs(sheet_name)

    print(f"  Threshold: {threshold}%")
    print(f"  Passing drugs ({len(passing_drugs)}): {sorted(passing_drugs)[:5]}...")

    # Create PNG plot
    create_cumulative_plot(df, sheet_name, threshold, OUTPUT_DIR)

    # Create Excel with chart
    create_excel_with_chart(df, sheet_name, threshold, OUTPUT_DIR, passing_drugs)

print("\n" + "="*60)
print("COMPLETE!")
print("="*60)
print(f"\nOutput directory: {OUTPUT_DIR}")
