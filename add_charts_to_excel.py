"""
Add 2D PK-PD charts to the Excel file using openpyxl.
Creates combined charts with model lines + raw data scatter on same plot.

Reference: https://openpyxl.readthedocs.io/en/stable/charts/scatter.html
"""

from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.text import RichText
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont
from pathlib import Path

EXCEL_PATH = Path("Output/2D_Plots/2D_PKPD_Plot_Data.xlsx")

# Chart size in cm (3.5 x 2.5 inches)
CHART_WIDTH = 8.89
CHART_HEIGHT = 6.35

# Scaled-down sizes for compact chart
FONT_SIZE = 700       # Font size in 1/100 pt (700 = 7pt)
TITLE_SIZE = 800      # Title font size (800 = 8pt)
LINE_WIDTH = 12000    # Line width in EMUs (about 1pt)
RANGE_LINE_WIDTH = 8000
MARKER_SIZE = 3       # Marker size (small)


def make_font(size=FONT_SIZE, bold=False):
    """Create a font specification for chart text."""
    return DrawingFont(typeface='Arial')


def add_combined_chart_to_sheet(wb, sheet_name):
    """Add a combined scatter+line chart showing model and raw data together."""
    ws = wb[sheet_name]

    is_o2 = 'O2' in sheet_name
    y_title = 'O₂ (%)' if is_o2 else 'Contractility'
    drug_name = sheet_name.replace('_O2', '').replace('_Contractility', '').replace('_', ' ')

    # Colors
    if is_o2:
        model_color = "000080"
        range_color = "6495ED"
        raw_color = "DC143C"
    else:
        model_color = "006400"
        range_color = "32CD32"
        raw_color = "FF8C00"

    max_row = ws.max_row

    # Find raw data columns
    raw_time_col = None
    max_col = ws.max_column
    for col in range(1, max_col + 1):
        header = ws.cell(row=1, column=col).value
        if header and 'Raw_Time' in str(header):
            raw_time_col = col
            break

    raw_data_rows = 0
    if raw_time_col:
        for row in range(2, max_row + 1):
            val = ws.cell(row=row, column=raw_time_col).value
            if val is not None:
                raw_data_rows = row
            else:
                break

    # Create chart
    chart = ScatterChart()
    chart.title = drug_name
    chart.style = 10
    chart.width = CHART_WIDTH
    chart.height = CHART_HEIGHT
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 96

    # Axis titles (short)
    chart.x_axis.title = "Time (h)"
    chart.y_axis.title = y_title

    # Reduce axis label font sizes
    chart.x_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=FONT_SIZE)), endParaRPr=CharacterProperties(sz=FONT_SIZE))]
    )
    chart.y_axis.txPr = RichText(
        p=[Paragraph(pPr=ParagraphProperties(defRPr=CharacterProperties(sz=FONT_SIZE)), endParaRPr=CharacterProperties(sz=FONT_SIZE))]
    )

    # Hide legend to save space (or make it smaller)
    chart.legend = None

    # === MODEL DATA AS SMOOTH LINES ===
    x_model = Reference(ws, min_col=1, min_row=2, max_row=max_row)

    # Model Mean
    y_mean = Reference(ws, min_col=2, min_row=2, max_row=max_row)
    series_mean = Series(y_mean, x_model, title="Mean")
    series_mean.marker = Marker(symbol='none')
    series_mean.graphicalProperties.line.solidFill = model_color
    series_mean.graphicalProperties.line.width = LINE_WIDTH
    series_mean.smooth = True
    chart.series.append(series_mean)

    # Model Min
    y_min = Reference(ws, min_col=3, min_row=2, max_row=max_row)
    series_min = Series(y_min, x_model, title="Min")
    series_min.marker = Marker(symbol='none')
    series_min.graphicalProperties.line.solidFill = range_color
    series_min.graphicalProperties.line.width = RANGE_LINE_WIDTH
    series_min.graphicalProperties.line.dashStyle = "dash"
    series_min.smooth = True
    chart.series.append(series_min)

    # Model Max
    y_max = Reference(ws, min_col=4, min_row=2, max_row=max_row)
    series_max = Series(y_max, x_model, title="Max")
    series_max.marker = Marker(symbol='none')
    series_max.graphicalProperties.line.solidFill = range_color
    series_max.graphicalProperties.line.width = RANGE_LINE_WIDTH
    series_max.graphicalProperties.line.dashStyle = "dash"
    series_max.smooth = True
    chart.series.append(series_max)

    # === RAW DATA AS SCATTER POINTS ===
    if raw_time_col and raw_data_rows > 1:
        x_raw = Reference(ws, min_col=raw_time_col, min_row=2, max_row=raw_data_rows)

        raw_cols = []
        for col in range(raw_time_col + 1, max_col + 1):
            header = ws.cell(row=1, column=col).value
            if header and 'Raw_' in str(header):
                raw_cols.append((col, header))

        for i, (col, header) in enumerate(raw_cols):
            y_raw = Reference(ws, min_col=col, min_row=2, max_row=raw_data_rows)
            series_raw = Series(y_raw, x_raw, title=None)

            # Small markers, no line
            series_raw.marker = Marker(symbol='circle', size=MARKER_SIZE)
            series_raw.marker.graphicalProperties.solidFill = raw_color
            series_raw.marker.graphicalProperties.line.noFill = True
            series_raw.graphicalProperties.line.noFill = True

            chart.series.append(series_raw)

    ws.add_chart(chart, "P2")
    return True


def main():
    print("Adding combined charts to Excel file...")
    print(f"File: {EXCEL_PATH}")
    print(f"Chart size: {CHART_WIDTH}cm x {CHART_HEIGHT}cm")

    wb = load_workbook(EXCEL_PATH)

    for sheet_name in wb.sheetnames:
        if sheet_name == 'Parameters':
            continue
        print(f"  Adding chart to: {sheet_name}")
        add_combined_chart_to_sheet(wb, sheet_name)

    output_path = EXCEL_PATH.parent / "2D_PKPD_Plot_Data_with_Charts.xlsx"
    wb.save(output_path)
    print(f"\nSaved to: {output_path}")


if __name__ == '__main__':
    main()
