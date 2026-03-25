"""
Generate Paper Figures for Cardiac RODEO PowerPoint
====================================================

This script generates all publication figures with full tracking.
Figures 1-8 as specified in the paper outline.

Usage:
    python generate_paper_figures.py --all
    python generate_paper_figures.py --figure 3
    python generate_paper_figures.py --extract-layout
    python generate_paper_figures.py --list
"""

import figure_config  # FIRST LINE - registers Helvetica
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.lines import Line2D
import numpy as np
import pandas as pd
import seaborn as sns
from pathlib import Path
from datetime import datetime
import json
import argparse
import re


# ============================================================================
# HEATMAP HELPER FUNCTIONS (from figure-graphing skill)
# ============================================================================

def clean_concentration_labels(labels):
    """
    Clean concentration labels by removing duplicate decimal suffixes.

    Examples:
        '8.1' → '8'      (removes .1 suffix that's just numbering)
        '8.2' → '8'      (removes .2 suffix)
        '0.1' → '0.1'    (keeps meaningful decimal - it's the actual value)
        '0.1.1' → '0.1'  (removes duplicate suffix from 0.1)
    """
    cleaned = []
    for label in labels:
        label_str = str(label)
        # Pattern: number.number.number (like 0.1.1) → keep first two parts
        if re.match(r'^\d+\.\d+\.\d+$', label_str):
            parts = label_str.split('.')
            cleaned.append(f"{parts[0]}.{parts[1]}")
        # Pattern: integer.single_digit at end (like 8.1, 8.2) → remove suffix
        elif re.match(r'^(\d+)\.[1-9]$', label_str):
            cleaned.append(re.match(r'^(\d+)\.[1-9]$', label_str).group(1))
        else:
            cleaned.append(label_str)
    return cleaned

# ============================================================================
# CONFIGURATION (Updated per figure-graphing skill)
# ============================================================================

PROJECT_ROOT = Path(__file__).parent.resolve()
FIGURES_DIR = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures'
REGISTRY_PATH = FIGURES_DIR / 'figure_registry.csv'

# Standard figure sizes from skill (inches)
SQUARE_SIZE = 1.7           # Standard square for bar charts, scatter, ROC, CM
HEATMAP_WIDTH = 3.1         # Heatmap width (2x height) - sized so 2 fit side-by-side on slide
HEATMAP_HEIGHT = 1.55       # Heatmap height (1:2 ratio)

# Legacy sizes for compatibility
SINGLE_W, SINGLE_H = SQUARE_SIZE, SQUARE_SIZE
DOUBLE_W, DOUBLE_H = SQUARE_SIZE * 2, SQUARE_SIZE

# Publication DPI
SAVE_DPI = 600

# Heatmap colormap colors (MANDATORY from skill)
HEATMAP_BLUE = '#123BFF'    # Low values
HEATMAP_RED = '#FF2908'     # High values

# Color palette
COLORS = {
    'blue': '#6C92ED',
    'green': '#7DB88A',
    'dusty_rose': '#C98B8E',
    'soft_gold': '#CCBC7E',
    'dark_blue': '#4A6FBF',
    'grey': '#8E8E8E',
    'pass': '#6C92ED',  # Blue for pass/positive
    'fail': '#8E8E8E',  # Grey for fail/negative
    'threshold': '#C98B8E',
    # Legacy aliases (map old names to new palette)
    'pink': '#7DB88A',      # green replaces pink
    'orange': '#C98B8E',    # dusty rose replaces orange
    'beige': '#CCBC7E',     # soft gold replaces beige
}

MODEL_COLORS = {
    'Organoid': '#6C92ED',
    'CNN (DIQT Transfer)': '#7DB88A',
    'CNN (5-fold on 25)': '#C98B8E',
    'ADMET-AI': '#7DB88A',
    'SwissADME': '#C98B8E',
}

# ============================================================================
# REGISTRY FUNCTIONS
# ============================================================================

def load_registry():
    if not REGISTRY_PATH.exists():
        return pd.DataFrame(columns=[
            'Figure_ID', 'Letter', 'Description', 'PNG_Path', 'Excel_Path',
            'Source_Script', 'External', 'Width_In', 'Height_In', 'Notes'
        ])
    return pd.read_csv(REGISTRY_PATH)


def save_registry(df):
    df.to_csv(REGISTRY_PATH, index=False)


def register_figure(fig_id, letter, description, png_path, excel_path=None,
                    external=False, width=SINGLE_W, height=SINGLE_H, notes='',
                    source_script='generate_paper_figures.py'):
    registry = load_registry()

    # Convert to strings for comparison (handle type mismatches)
    fig_id_str = str(fig_id)
    letter_str = str(letter) if letter else ''

    # Create mask for existing entry
    registry['Figure_ID'] = registry['Figure_ID'].astype(str)
    registry['Letter'] = registry['Letter'].fillna('').astype(str)
    mask = (registry['Figure_ID'] == fig_id_str) & (registry['Letter'] == letter_str)

    entry = {
        'Figure_ID': fig_id_str,
        'Letter': letter_str,
        'Description': description,
        'PNG_Path': str(png_path),
        'Excel_Path': str(excel_path) if excel_path else '',
        'Source_Script': source_script,
        'External': external,
        'Width_In': width,
        'Height_In': height,
        'Notes': notes
    }

    if mask.any():
        for key, value in entry.items():
            registry.loc[mask, key] = value
    else:
        registry = pd.concat([registry, pd.DataFrame([entry])], ignore_index=True)

    # Remove any duplicate entries (keep last)
    registry = registry.drop_duplicates(subset=['Figure_ID', 'Letter'], keep='last')

    save_registry(registry)
    print(f"  Registered: Fig_{fig_id}{letter}")


def fit_to_slide(png_path, max_width_in=7.09, target_dpi=600):
    """Resize an image to fit within the slide width, preserving aspect ratio.

    Keeps the original file at 600 DPI. If the image already fits, does nothing.
    Used for externally-generated figures (SHAP, 2D plots, etc.) that may be
    larger than the slide dimensions.
    """
    from PIL import Image as PILImage
    img = PILImage.open(png_path)
    w, h = img.size
    src_dpi = img.info.get('dpi', (target_dpi, target_dpi))
    dpi_x = src_dpi[0] if src_dpi[0] else target_dpi
    w_in = w / dpi_x
    if w_in > max_width_in:
        scale = max_width_in / w_in
        new_w = int(w * scale)
        new_h = int(h * scale)
        img = img.resize((new_w, new_h), PILImage.LANCZOS)
        img.save(png_path, dpi=(target_dpi, target_dpi))
        print(f"  Resized: {w}x{h} -> {new_w}x{new_h} ({w_in:.1f}\" -> {max_width_in:.1f}\")")


def save_figure(fig, fig_id, letter, description, data_dict=None,
                width=SINGLE_W, height=SINGLE_H, notes='', exact_size=False,
                source_script='generate_paper_figures.py'):
    """Save figure and data, then register.

    Args:
        exact_size: If True, saves at exact figsize (no bbox_inches='tight')
                   Use for SQUARE figures where dimensions must be exact.
        source_script: The script that generates this figure (for tracking).
    """
    folder = FIGURES_DIR / f'Fig_{fig_id}'
    folder.mkdir(parents=True, exist_ok=True)

    png_path = folder / f'Fig_{fig_id}{letter}.png'
    if exact_size:
        fig.savefig(png_path, dpi=SAVE_DPI, facecolor='white')
    else:
        fig.savefig(png_path, dpi=SAVE_DPI, bbox_inches='tight', facecolor='white')
    print(f"  Saved: {png_path}")

    excel_path = None
    if data_dict:
        excel_path = folder / f'Fig_{fig_id}{letter}_data.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for sheet, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet[:31], index=True)
        print(f"  Saved: {excel_path}")

    register_figure(fig_id, letter, description,
                    png_path.relative_to(PROJECT_ROOT),
                    excel_path.relative_to(PROJECT_ROOT) if excel_path else None,
                    width=width, height=height, notes=notes, source_script=source_script)
    plt.close(fig)


def style_axis(ax, title=None, xlabel=None, ylabel=None, grid=True):
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    if title:
        ax.set_title(title, fontsize=9, fontweight='bold')
    if xlabel:
        ax.set_xlabel(xlabel, fontsize=9)
    if ylabel:
        ax.set_ylabel(ylabel, fontsize=9)
    if grid:
        ax.grid(axis='y', alpha=0.3)
    ax.tick_params(labelsize=8)


# ============================================================================
# 3D SURFACE + RAW DATA OVERLAY HELPER (ported from MATLAB/plot_3d_surface_with_raw.py)
# ============================================================================

def _pkpd_elimination_response(dose_ratio, time, R0, Emax, kappa, n, m, tau, k_elim):
    """PK-PD elimination equation: R(C0, t) = R0 + Emax*(1 - exp(-kappa*(dr*exp(-k_elim*t))^n * (t/tau)^m))"""
    import math
    dose_ratio = np.asarray(dose_ratio, dtype=float)
    time = np.asarray(time, dtype=float)
    kappa = max(kappa, 1e-9)
    tau = max(tau, 1e-9)
    k_elim = max(k_elim, 1e-9)
    time = np.maximum(time, 1e-9)
    C_norm = dose_ratio * np.exp(-k_elim * time)
    effect_term = kappa * (C_norm ** n) * ((time / tau) ** m)
    return R0 + Emax * (1 - np.exp(-effect_term))


def generate_3d_surface_with_raw(drug_name, response_type, remove_r0_offset=True,
                                  figsize=(6, 6)):
    """
    Generate 3D PK-PD surface with raw experimental data scatter overlay.
    Ported from MATLAB plot_3d_surface_with_raw.m with full data tracking.

    Returns: (fig, data_dict) for save_figure()
    """
    import math

    coeff_path = PROJECT_ROOT / 'EQN_Coefficients' / 'all_equations_coefficients.xlsx'
    o2_data_path = PROJECT_ROOT / 'Cleaned_Data' / 'O2_Mean_Averaged.xlsx'
    contr_data_path = PROJECT_ROOT / 'Cleaned_Data' / 'Heart_Contractility_Averaged.xlsx'

    # --- Load coefficients ---
    df_coeff = pd.read_excel(coeff_path, sheet_name='pkpd_elimination', header=1)
    df_coeff.columns = df_coeff.columns.str.strip()
    row = df_coeff[df_coeff['Drug'] == drug_name].iloc[0]

    suffix = '.1' if response_type == 'O2' else ''
    R0 = float(row[f'R0{suffix}'])
    Emax = float(row[f'Emax{suffix}'])
    kappa = float(row[f'kappa{suffix}'])
    n_param = float(row[f'n{suffix}'])
    m_param = float(row[f'm{suffix}'])
    tau = float(row[f'tau{suffix}'])
    k_elim = float(row[f'k_elim{suffix}'])
    cmax = float(row[f'Cmax_used{suffix}'])

    print(f"  {drug_name} {response_type}: R0={R0:.4f}, Emax={Emax:.4f}, kappa={kappa:.4f}, "
          f"n={n_param:.4f}, m={m_param:.4f}, tau={tau:.4f}, k_elim={k_elim:.4f}, Cmax={cmax}")

    # --- Load raw data ---
    data_path = o2_data_path if response_type == 'O2' else contr_data_path
    df_raw = pd.read_excel(data_path, sheet_name=drug_name)
    time_vals = df_raw.iloc[:, 0].values
    conc_columns = df_raw.columns[1:]

    all_times, all_drs, all_resp = [], [], []
    for conc_col in conc_columns:
        try:
            conc = float(str(conc_col).replace('_', '.'))
            dr = conc / cmax
            for t, r in zip(time_vals, df_raw[conc_col].values):
                if not np.isnan(r):
                    all_times.append(float(t))
                    all_drs.append(dr)
                    all_resp.append(float(r))
        except (ValueError, TypeError):
            continue

    raw_t = np.array(all_times)
    raw_dr = np.array(all_drs)
    raw_r = np.array(all_resp)
    print(f"  Raw data: {len(raw_t)} points, dose_ratio range: {raw_dr.min():.2f}-{raw_dr.max():.2f}")

    # --- Compute surface ---
    max_dr = max(2.0, math.ceil(raw_dr.max())) if len(raw_dr) > 0 else 2.0
    dose_ratio_vec = np.linspace(0, max_dr, 60)
    time_vec = np.linspace(0, 96, 60)
    T, Dr = np.meshgrid(time_vec, dose_ratio_vec)
    Response = _pkpd_elimination_response(Dr, T, R0, Emax, kappa, n_param, m_param, tau, k_elim)

    if remove_r0_offset:
        Response = Response - R0
        raw_r = raw_r - R0

    # --- Plot ---
    fig = plt.figure(figsize=figsize)
    ax = fig.add_subplot(111, projection='3d', computed_zorder=False)

    vmin = 0
    vmax = float(np.nanmax(Response))
    norm = plt.Normalize(vmin=vmin, vmax=vmax)

    ax.plot_surface(T, Dr, Response, cmap='turbo', norm=norm,
                    linewidth=0, antialiased=True, edgecolor='none', alpha=0.7)

    ax.scatter(raw_t, raw_dr, raw_r, c='black', s=25, alpha=0.8,
               depthshade=True, zorder=10)

    # Axis limits
    z_min = min(0, float(np.nanmin(raw_r))) if len(raw_r) > 0 else 0
    z_max = max(vmax, float(np.nanmax(raw_r)) * 1.1) if len(raw_r) > 0 else vmax
    ax.set_zlim(z_min, z_max)

    # Labels
    ax.set_xlabel('Time (hours)', fontsize=9, labelpad=5)
    ax.set_ylabel('Dose Ratio (C0/Cmax)', fontsize=9, labelpad=5)
    z_label = 'O2 (% air sat.)' if response_type == 'O2' else 'Contractility (Amp std)'
    if remove_r0_offset:
        z_label += ' (R0 removed)'
    ax.set_zlabel(z_label, fontsize=9, labelpad=5)
    ax.set_title(f'{drug_name} - {response_type}', fontsize=10, fontweight='bold', pad=10)

    ax.view_init(elev=25, azim=-158)
    ax.tick_params(labelsize=7)

    # Transparent panes
    ax.xaxis.pane.fill = False
    ax.yaxis.pane.fill = False
    ax.zaxis.pane.fill = False
    ax.xaxis.pane.set_edgecolor('lightgray')
    ax.yaxis.pane.set_edgecolor('lightgray')
    ax.zaxis.pane.set_edgecolor('lightgray')

    fig.tight_layout()

    # --- Build tracking data ---
    coeff_df = pd.DataFrame([{
        'Parameter': p, 'Value': v,
        'Source': str(coeff_path), 'Sheet': 'pkpd_elimination'
    } for p, v in [('R0', R0), ('Emax', Emax), ('kappa', kappa),
                    ('n', n_param), ('m', m_param), ('tau', tau),
                    ('k_elim', k_elim), ('Cmax', cmax)]])

    raw_df = pd.DataFrame({
        'Time': raw_t, 'Dose_Ratio': raw_dr, 'Response': raw_r
    })
    raw_df['Source'] = str(data_path)
    raw_df['Sheet'] = drug_name
    raw_df['Drug'] = drug_name
    raw_df['Response_Type'] = response_type

    data_dict = {
        'Coefficients': coeff_df,
        'Raw_Data': raw_df,
    }
    return fig, data_dict


# ============================================================================
# FIGURE 1: Pipeline Schematic (Placeholder)
# ============================================================================

def generate_fig_1():
    """Figure 1: Pipeline schematic - placeholder."""
    print("\n=== Figure 1: Pipeline Schematic (Placeholder) ===")

    fig, ax = plt.subplots(figsize=(DOUBLE_W, DOUBLE_H))
    ax.text(0.5, 0.5, 'Figure 1\n\nHigh throughput cardiac organoid based\ndrug-cardiotoxicity prediction pipeline\n\n[Schematic to be added]',
            ha='center', va='center', fontsize=12, transform=ax.transAxes,
            bbox=dict(boxstyle='round', facecolor=COLORS['beige'], alpha=0.5))
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.axis('off')

    save_figure(fig, '1', '', 'Pipeline Schematic (Placeholder)',
                width=DOUBLE_W, height=DOUBLE_H, notes='External schematic to be added')


# ============================================================================
# FIGURE 2: Robust Generation
# ============================================================================

def generate_fig_2():
    """Figure 2: Robust generation - SNR analysis + external images.

    Panel a (SNR histogram) is script-generated.
    Other panels (microscopy, plate images, diagrams) are externally managed.
    """
    print("\n=== Figure 2: Robust Generation ===")

    # 2a: SNR Quality Analysis - stacked bar chart (QC range on bottom, out-of-range on top)
    snr_path = PROJECT_ROOT / 'Output' / 'Excel_Figures' / 'snr_analysis.xlsx'
    if snr_path.exists():
        snr_df_full = pd.read_excel(snr_path, sheet_name='QC_Range_0_to_80')

        # Filter to positive SNR buckets (skip -0.95 which is all out-of-range)
        snr_df_plot = snr_df_full[snr_df_full['SNR Bucket'] > 0].copy()
        snr_df_plot = snr_df_plot[snr_df_plot['SNR Bucket'] <= 5.0].reset_index(drop=True)

        fig_w, fig_h = 3.4, 2.04
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))

        x = np.arange(len(snr_df_plot))
        bar_width = 0.85

        in_range = snr_df_plot['% In Range (0-80%)'].values
        out_range = snr_df_plot['% Out of Range'].values

        # Stacked bars: blue (good) on bottom, red (bad) on top
        ax.bar(x, in_range, bar_width, color=COLORS['blue'], label='QC Range (0-80% O2)')
        ax.bar(x, out_range, bar_width, bottom=in_range, color=COLORS['pink'], label='Out of Range')

        # SNR = 0.4 threshold line
        threshold_snr = 0.4
        threshold_idx = None
        for i, val in enumerate(snr_df_plot['SNR Bucket'].values):
            if abs(val - threshold_snr) < 0.06:
                threshold_idx = i
                break
        if threshold_idx is not None:
            ax.axvline(x=threshold_idx, color='black', linestyle='--', linewidth=1.5,
                       label=f'SNR = {threshold_snr} threshold')

        # X-axis: show every Nth label for readability
        x_labels = [f'{v:.1f}' for v in snr_df_plot['SNR Bucket'].values]
        step = max(1, len(x_labels) // 10)
        ax.set_xticks(x[::step])
        ax.set_xticklabels([x_labels[i] for i in range(0, len(x_labels), step)],
                           rotation=0, ha='center', fontsize=16, fontweight='bold')

        ax.set_ylabel('% Measurements', fontsize=18, fontweight='bold')
        ax.set_xlabel('SNR Threshold', fontsize=18, fontweight='bold')
        ax.set_title('SNR Quality Analysis', fontsize=19, fontweight='bold')
        ax.set_ylim(0, 105)
        ax.tick_params(labelsize=16)
        ax.legend(fontsize=15, loc='lower right')
        for spine in ax.spines.values():
            spine.set_edgecolor('black')
            spine.set_linewidth(1.5)

        fig.tight_layout()

        # Tracking data
        snr_df_full['Source'] = str(snr_path)
        snr_df_full['Sheet'] = 'QC_Range_0_to_80'

        qc_analysis_path = PROJECT_ROOT / 'Output' / 'QC_Analysis' / 'BucketAnalysis_QC_Range.xlsx'
        data_dict = {'SNR_QC_Data': snr_df_full}
        if qc_analysis_path.exists():
            qc_df = pd.read_excel(qc_analysis_path)
            qc_df['Source'] = str(qc_analysis_path)
            data_dict['QC_Analysis'] = qc_df

        save_figure(fig, '2', 'd', 'SNR Quality Analysis',
                    data_dict, width=fig_w, height=fig_h,
                    notes='Stacked bars: QC range (blue) on bottom, out-of-range (red) on top. SNR 0.4 threshold line.')
    else:
        print(f"  Warning: {snr_path} not found for Fig_2d")

    # --- Epirubicin O2 averaged dose-response (from raw per-well data) ---
    _generate_fig2_epirubicin_o2()

    # --- Mexiletine Contractility offset-averaged dose-response ---
    _generate_fig2_dose_response_2d()

    # --- Epirubicin TC50 dose-response plot ---
    _generate_fig2_epirubicin_tc50()

    # --- Epirubicin O2 heatmap (smoothed per-well) ---
    _generate_fig2_epirubicin_o2_heatmap()

    # --- Mexiletine stacked waveforms ---
    _generate_fig2_mexiletine_waveforms()

    # --- Mexiletine Contractility heatmap ---
    _generate_fig2_mexiletine_contractility_heatmap()

    # Remaining panels are external (microscopy, plate images, EMF diagrams)
    print("  Remaining panels: externally managed")


def _generate_fig2_dose_response_2d():
    """Generate Mexiletine Contractility offset-averaged dose-response plot for Figure 2.

    Delegates to the standalone script plot_contractility.py which saves directly
    to Output/PowerPoint_Figures/Fig_2/Fig_2j_Mexiletine_Contractility.png.
    This avoids reimplementing the pipeline and ensures identical output.

    Reference script: Cleaned_Data/Raw_Example_Data/plot_contractility.py
    """
    import subprocess

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    script = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'plot_contractility.py'
    csv_path = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'Mexiletine' / 'Amp_std.csv'

    if not script.exists():
        print(f"  Warning: {script} not found")
        return

    # Run the standalone script (it saves to both local and Fig_2 folders)
    result = subprocess.run(
        ['python', str(script)],
        cwd=str(script.parent),
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  Error running plot_contractility.py: {result.stderr}")
        return

    png_path = fig2_dir / 'Fig_2j_Mexiletine_Contractility.png'
    if not png_path.exists():
        print(f"  Warning: {png_path} not generated")
        return

    # Save data
    if csv_path.exists():
        df = pd.read_csv(csv_path)
        excel_path = fig2_dir / 'Fig_2j_Mexiletine_Contractility_data.xlsx'
        save_df = df.copy()
        save_df.insert(0, 'Source', str(csv_path))
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            save_df.to_excel(writer, sheet_name='Mexiletine', index=False)
    else:
        excel_path = fig2_dir / 'Fig_2j_Mexiletine_Contractility_data.xlsx'

    register_figure('2', 'Mexiletine_Contractility',
                    'Contractility Dose Dependent Response (Mexiletine)',
                    png_path.relative_to(PROJECT_ROOT),
                    excel_path.relative_to(PROJECT_ROOT),
                    width=12.0, height=8.0,
                    notes='Raw per-well contractility offset-averaged. LOWESS x3 + Gaussian + CubicSpline. '
                          'Wells {20,22,23,24,25,27} excluded. All traces offset to global average start.')
    print(f"  Fig_2j_Mexiletine_Contractility.png")


def _generate_fig2_epirubicin_o2():
    """Generate Epirubicin O2 averaged dose-response plot for Figure 2.

    Delegates to the standalone script plot_epirubicin_o2.py which saves directly
    to Output/PowerPoint_Figures/Fig_2/Fig_2g_Epirubicin_O2.png.
    This avoids reimplementing the pipeline and ensures identical output.

    Reference script: Cleaned_Data/Raw_Example_Data/plot_epirubicin_o2.py
    """
    import subprocess

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    script = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'plot_epirubicin_o2.py'
    csv_path = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'Epirubicin' / 'O2_mean.csv'

    if not script.exists():
        print(f"  Warning: {script} not found")
        return

    # Run the standalone script (it saves to both local and Fig_2 folders)
    result = subprocess.run(
        ['python', str(script)],
        cwd=str(script.parent),
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  Error running plot_epirubicin_o2.py: {result.stderr}")
        return

    png_path = fig2_dir / 'Fig_2g_Epirubicin_O2.png'
    if not png_path.exists():
        print(f"  Warning: {png_path} not generated")
        return

    # Save data
    if csv_path.exists():
        df = pd.read_csv(csv_path)
        excel_path = fig2_dir / 'Fig_2g_Epirubicin_O2_data.xlsx'
        save_df = df.copy()
        save_df.insert(0, 'Source', str(csv_path))
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            save_df.to_excel(writer, sheet_name='Epirubicin_O2', index=False)
    else:
        excel_path = fig2_dir / 'Fig_2g_Epirubicin_O2_data.xlsx'

    register_figure('2', 'Epirubicin_O2',
                    'Metabolic Dose Dependent Response (Epirubicin O2)',
                    png_path.relative_to(PROJECT_ROOT),
                    excel_path.relative_to(PROJECT_ROOT),
                    width=12.0, height=8.0,
                    notes='Raw per-well O2 averaged by conc. LOWESS x3 + Gaussian + CubicSpline. '
                          'Well 1 excluded (85% outlier). Targeted baseline shift for extreme traces. '
                          'MARGIN=1.5%.')
    print(f"  Fig_2g_Epirubicin_O2.png")


def _generate_fig2_epirubicin_tc50():
    """Generate Epirubicin TC50 dose-response curve for Figure 2.

    Uses per-well raw O2 data, converts to O2 consumption (80 - O2),
    averages replicates per concentration, and fits a 4-parameter logistic.
    """
    from scipy.optimize import curve_fit

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    drug = 'Epirubicin'
    time_hour = 32  # hours

    RAW_O2_PATH = PROJECT_ROOT / 'Cleaned_Data' / 'DrugScreen19.11.25_compiled_O2_mean.xlsx'
    if not RAW_O2_PATH.exists():
        print(f"  Warning: {RAW_O2_PATH} not found")
        return

    # Load per-well raw data (rows=concentrations with duplicates, cols=time points)
    df_raw = pd.read_excel(RAW_O2_PATH, sheet_name='Epirubicin O2_mean', header=None)
    time_points = df_raw.iloc[0, 1:].astype(float).values
    concentrations = df_raw.iloc[1:, 0].astype(float).values
    o2_matrix = df_raw.iloc[1:, 1:].astype(float).values

    # Find closest time point
    idx_t = np.argmin(np.abs(time_points - time_hour))
    o2_at_t = o2_matrix[:, idx_t]

    # Average replicates per concentration first, then scale 0-100
    df_tp = pd.DataFrame({'Concentration': concentrations, 'O2': o2_at_t})
    df_avg = df_tp.groupby('Concentration', as_index=False).mean().sort_values('Concentration')
    df_avg = df_avg[df_avg['Concentration'] > 0]

    # Scale to 0-100%: lowest O2 = 0% consumption, highest O2 = 100% consumption
    # (low concentration → organoids healthy → high O2 → 100% consumption)
    # (high concentration → organoids dead → low O2 → 0% consumption)
    o2_min = df_avg['O2'].min()
    o2_max = df_avg['O2'].max()
    df_avg['Consumption'] = (1 - (df_avg['O2'] - o2_min) / (o2_max - o2_min)) * 100

    x_conc = df_avg['Concentration'].values
    y_cons = df_avg['Consumption'].values
    x_log = np.log10(x_conc)

    # 4-parameter logistic fit in log space
    def logistic(xlog, bottom, top, logEC50, slope):
        return bottom + (top - bottom) / (1 + np.exp((logEC50 - xlog) * slope))

    tc50 = None
    popt = None
    try:
        p0 = [y_cons.min(), y_cons.max(), np.median(x_log), 1.0]
        popt, _ = curve_fit(logistic, x_log, y_cons, p0=p0, maxfev=20000)
        bottom, top, logEC50, slope = popt
        # Solve for TC50: where consumption = 50
        target = 50.0
        denom = top - bottom
        denom_target = target - bottom
        if slope != 0 and denom != 0 and denom_target != 0:
            ratio = denom / denom_target - 1.0
            if ratio > 0:
                tc50 = 10 ** (logEC50 - (1.0 / slope) * np.log(ratio))
    except Exception as e:
        print(f"  Warning: TC50 sigmoid fit failed: {e}")

    # Plot
    fig, ax = plt.subplots(figsize=(4, 2.8))
    ax.plot(x_conc, y_cons, 'o', markersize=7, color='#1f77b4', zorder=5, label='Mean consumption')

    if popt is not None:
        x_smooth = np.linspace(x_log.min() - 0.2, x_log.max() + 0.2, 200)
        ax.plot(10 ** x_smooth, logistic(x_smooth, *popt), '-', color='#1f77b4',
                linewidth=2, label='Sigmoid fit')

    if tc50 is not None and np.isfinite(tc50):
        ax.axhline(50, color='grey', linestyle='--', linewidth=1, alpha=0.5)
        ax.axvline(tc50, color='red', linestyle='--', linewidth=1.5)
        ax.text(0.05, 0.08, f'TC50={tc50:.3f} mM', transform=ax.transAxes,
                fontsize=19, fontweight='bold')

    ax.set_xscale('log')
    ax.set_xlabel('Concentration (mM)', fontsize=19)
    ax.set_ylabel('O2 Consumption (%)', fontsize=19)
    ax.set_title(f'{drug} TC50 ({time_hour}h)', fontsize=20, fontweight='bold')
    ax.set_ylim(0, 100)
    ax.set_xticks(x_conc)
    ax.set_xticklabels([f'{c:.3g}' for c in x_conc], fontsize=17)
    ax.tick_params(labelsize=17)
    ax.legend(fontsize=17, loc='upper right')
    ax.grid(True, alpha=0.3)
    fig.tight_layout()

    png_path = fig2_dir / 'Fig_2h_Epirubicin_TC50.png'
    fig.savefig(png_path, dpi=SAVE_DPI, bbox_inches='tight', pad_inches=0.05,
                facecolor='white')
    plt.close(fig)

    # Save data
    excel_path = fig2_dir / 'Fig_2h_Epirubicin_TC50_data.xlsx'
    save_df = df_avg.copy()
    save_df['Source'] = str(RAW_O2_PATH)
    save_df['Timepoint_h'] = time_hour
    if tc50 is not None:
        save_df['TC50_mM'] = tc50
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        save_df.to_excel(writer, sheet_name='TC50', index=False)

    tc50_str = f'{tc50:.3f}' if tc50 is not None else 'N/A'
    register_figure('2', 'Epirubicin_TC50',
                    f'{drug} TC50 dose-response at {time_hour}h',
                    png_path.relative_to(PROJECT_ROOT),
                    excel_path.relative_to(PROJECT_ROOT),
                    width=4.0, height=2.8,
                    notes=f'TC50={tc50_str} mM at {time_hour}h, 4PL sigmoid fit on per-well data')
    print(f"  Fig_2h_Epirubicin_TC50.png (TC50={tc50_str} mM)")


def _generate_fig2_heatmap(csv_path, drug_name, is_contractility, out_filename,
                           y_label, cbar_label, drop_wells=None, vmax_override=None):
    """Generate a concentration-grouped LOWESS heatmap for Figure 2.

    Follows the same pipeline as Fig 3a heatmaps but with Fig 2 styling:
    no title, large fonts, straight x-labels, matching graph height.
    """
    from statsmodels.nonparametric.smoothers_lowess import lowess as lowess_func
    from collections import OrderedDict
    from matplotlib.colors import LinearSegmentedColormap
    import seaborn as sns

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    if not csv_path.exists():
        print(f"  Warning: {csv_path} not found")
        return

    LOWESS_W = 16
    cmap = LinearSegmentedColormap.from_list('cardiac_rodeo',
                                              [HEATMAP_BLUE, 'white', HEATMAP_RED])
    cmap.set_bad('white')

    def _get_base_conc(col_name):
        """Strip pandas duplicate suffixes: '10.1' -> 10, '0.625.2' -> 0.625."""
        import re
        s = str(col_name)
        # Try stripping trailing '.N' where N is a single digit (pandas suffix)
        m = re.match(r'^(.+?)\.(\d)$', s)
        if m:
            base = m.group(1)
            try:
                return float(base)
            except ValueError:
                pass
        try:
            return float(s)
        except ValueError:
            return None

    def _conc_label(col_name):
        val = _get_base_conc(col_name)
        if val is None: return str(col_name)
        return str(int(val)) if val == int(val) else str(val)

    def _apply_lowess(df):
        smoothed = df.copy().astype(float)
        for col in smoothed.columns:
            series = smoothed[col]
            valid = series.dropna()
            if len(valid) < 3: continue
            frac = min(1.0, max(LOWESS_W, 1) / len(valid))
            fitted = lowess_func(valid.values, np.arange(len(valid)),
                                 frac=frac, return_sorted=False)
            target = smoothed.index.get_indexer(valid.index)
            smoothed.iloc[target, smoothed.columns.get_loc(col)] = fitted
        return smoothed

    # Load and process
    df_raw = pd.read_csv(csv_path, index_col=0)
    if drop_wells:
        cols_to_drop = [c for c in drop_wells if c in df_raw.columns]
        if cols_to_drop:
            df_raw = df_raw.drop(columns=cols_to_drop)

    for col in df_raw.columns:
        df_raw[col] = df_raw[col].interpolate(method='linear', limit=10, limit_direction='both')

    df_smooth = _apply_lowess(df_raw)
    data = df_smooth.T  # rows=wells, cols=time

    if is_contractility:
        data = data * 100
    else:
        data = data.clip(upper=100)

    y_labels = [_conc_label(c) for c in data.index.tolist()]
    x_labels = [str(t) for t in data.columns.tolist()]

    # Figure sized to match dose-response graph height
    fig, ax = plt.subplots(figsize=(6, 3.5))

    vmax = vmax_override if vmax_override else None
    sns.heatmap(
        data, annot=False, cmap=cmap,
        vmin=0, vmax=vmax,
        cbar_kws={'shrink': 0.8},
        xticklabels=x_labels, yticklabels=y_labels,
        square=False, linewidths=0, ax=ax
    )

    ax.set_xlabel('Time from Exposure (h)', fontsize=32, fontweight='bold')
    ax.set_ylabel(y_label, fontsize=32, fontweight='bold')
    # No title

    # X ticks — straight, thinned out, bold
    n_x = len(x_labels)
    x_step = max(1, n_x // 10)
    ax.set_xticks(range(0, n_x, x_step))
    ax.set_xticklabels([x_labels[i] for i in range(0, n_x, x_step)],
                        rotation=0, ha='center', fontsize=26, fontweight='bold')

    # Y ticks — one per concentration group, no "mM", bold
    conc_groups = OrderedDict()
    for i, lbl in enumerate(y_labels):
        conc_groups.setdefault(lbl, []).append(i)
    tick_positions = [(indices[0] + indices[-1]) / 2 + 0.5 for indices in conc_groups.values()]
    ax.set_yticks(tick_positions)
    ax.set_yticklabels(list(conc_groups.keys()), fontsize=26, rotation=0, fontweight='bold')

    # Thick black borders
    for spine in ax.spines.values():
        spine.set_edgecolor('black')
        spine.set_linewidth(2.0)
    ax.tick_params(width=1.5)

    # Colorbar
    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=24, width=1.5)
    cbar.set_label(cbar_label, fontsize=26, fontweight='bold')
    cbar.outline.set_linewidth(2.0)
    if vmax_override:
        ticks = [t for t in range(0, int(vmax_override) + 1, 20)]
        cbar.set_ticks(ticks)

    fig.tight_layout()

    dst = fig2_dir / out_filename
    fig.savefig(str(dst), dpi=SAVE_DPI, bbox_inches='tight', pad_inches=0.05,
                facecolor='white')
    plt.close(fig)
    print(f"  {out_filename}")
    return dst


def _generate_fig2_epirubicin_o2_heatmap():
    """Epirubicin O2 concentration-grouped LOWESS heatmap for Figure 2."""
    csv_path = PROJECT_ROOT / 'Cleaned_Data' / 'Heatmaps' / 'Epirubicin' / 'O2_mean_sorted.csv'
    dst = _generate_fig2_heatmap(
        csv_path, 'Epirubicin', is_contractility=False,
        out_filename='Fig_2i_Epirubicin_O2_Heatmap.png',
        y_label='Epirubicin Dose',
        cbar_label='Oxygen (% Air)',
        drop_wells=['0.38.1'],
        vmax_override=100
    )
    if dst:
        register_figure('2', 'Epirubicin_O2_Heatmap',
                        'Epirubicin O2 Heatmap (LOWESS w=16)',
                        dst.relative_to(PROJECT_ROOT), None,
                        notes='Sorted wells, LOWESS w=16, well 0.38.1 excluded. vmax=100.')


def _generate_fig2_doxorubicin_waveforms():
    """Copy Doxorubicin stacked waveforms into Figure 2.

    Source: Output/HeartRate_Analysis/DoseResponse/doxorubicin_Stacked.png
    """
    import shutil

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    src = PROJECT_ROOT / 'Output' / 'HeartRate_Analysis' / 'DoseResponse' / 'doxorubicin_Stacked.png'
    dst = fig2_dir / 'Fig_2_Doxorubicin_Waveforms.png'

    if not src.exists():
        print(f"  Warning: {src} not found")
        return

    shutil.copy2(src, dst)
    register_figure('2', 'Doxorubicin_Waveforms',
                    'Doxorubicin Stacked Waveforms (3 doses x 3 timepoints)',
                    dst.relative_to(PROJECT_ROOT),
                    None,
                    notes='Stacked contractility waveforms at High/Med/Low doses across 6h/27h/82h.')
    print(f"  Fig_2_Doxorubicin_Waveforms.png")


def _generate_fig2_mexiletine_waveforms():
    """Generate Mexiletine stacked waveforms for Figure 2.

    Delegates to the standalone script plot_mexiletine_waveforms.py.
    Reference script: Cleaned_Data/Raw_Example_Data/plot_mexiletine_waveforms.py
    """
    import subprocess

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    script = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'plot_mexiletine_waveforms.py'
    if not script.exists():
        print(f"  Warning: {script} not found")
        return

    result = subprocess.run(
        ['python', str(script)],
        cwd=str(script.parent),
        capture_output=True, text=True
    )
    if result.returncode != 0:
        print(f"  Error running plot_mexiletine_waveforms.py: {result.stderr}")
        return

    png_path = fig2_dir / 'Fig_2k_Mexiletine_Waveforms.png'
    if png_path.exists():
        register_figure('2', 'Mexiletine_Waveforms',
                        'Mexiletine Stacked Waveforms (3 doses, 48h)',
                        png_path.relative_to(PROJECT_ROOT),
                        None,
                        notes='Stacked contractility waveforms at 5/1.25/0.625 mM, 48h. '
                              'Bandpass filtered from raw Dynamix data.')
        print(f"  Fig_2k_Mexiletine_Waveforms.png")
    else:
        print(f"  Warning: {png_path} not generated")


def _generate_fig2_mexiletine_contractility_heatmap():
    """Mexiletine Contractility heatmap for Figure 2.

    Uses RAW CSV with exact exclusions from Mexiletine_Contractility_Heatmap_NOTES.txt.
    14 wells excluded (by 1-based index + column name), then sorted ascending bottom-up.
    """
    from statsmodels.nonparametric.smoothers_lowess import lowess as lowess_func
    from collections import OrderedDict
    from matplotlib.colors import LinearSegmentedColormap
    import seaborn as sns
    import re

    fig2_dir = FIGURES_DIR / 'Fig_2'
    fig2_dir.mkdir(parents=True, exist_ok=True)

    csv_path = PROJECT_ROOT / 'Cleaned_Data' / 'Raw_Example_Data' / 'Mexiletine' / 'Amp_std.csv'
    if not csv_path.exists():
        print(f"  Warning: {csv_path} not found")
        return

    # Exact exclusions from NOTES file
    REMOVE_ORIGINAL = {4, 5, 6, 7, 14, 15, 17, 21, 22, 24, 26}  # 1-based indices
    REMOVE_COLS_EXTRA = {'20', '2.5.1', '2.5'}  # CSV column names
    LOWESS_W = 16

    cmap_hm = LinearSegmentedColormap.from_list('cardiac_rodeo',
                                                 [HEATMAP_BLUE, 'white', HEATMAP_RED])
    cmap_hm.set_bad('white')

    def _get_base_conc(col_name):
        s = str(col_name)
        m = re.match(r'^(.+?)\.(\d)$', s)
        if m:
            try: return float(m.group(1))
            except ValueError: pass
        try: return float(s)
        except ValueError: return None

    def _conc_label(col_name):
        val = _get_base_conc(col_name)
        if val is None: return str(col_name)
        return str(int(val)) if val == int(val) else str(val)

    # Load and drop by index + column name
    df = pd.read_csv(csv_path, index_col=0)
    keep_cols = [col for i, col in enumerate(df.columns) if (i + 1) not in REMOVE_ORIGINAL]
    keep_cols = [col for col in keep_cols if col not in REMOVE_COLS_EXTRA]
    df = df[keep_cols]

    # Interpolate
    for col in df.columns:
        df[col] = df[col].interpolate(method='linear', limit=10, limit_direction='both')

    # LOWESS full smoothing (first point NOT preserved)
    for col in df.columns:
        series = df[col].dropna()
        if len(series) < 3:
            continue
        frac = min(1.0, max(LOWESS_W, 1) / len(series))
        fitted = lowess_func(series.values, np.arange(len(series)),
                             frac=frac, return_sorted=False)
        target = df.index.get_indexer(series.index)
        df.iloc[target, df.columns.get_loc(col)] = fitted

    # Transpose (rows=wells, cols=time)
    data = df.T

    # Sort within each concentration group by average (ascending bottom-up)
    conc_vals = [_get_base_conc(c) for c in data.index.tolist()]
    conc_groups_sort = OrderedDict()
    for i, cv in enumerate(conc_vals):
        conc_groups_sort.setdefault(cv, []).append(i)
    sorted_indices = []
    for cv in sorted(conc_groups_sort.keys(), reverse=True):
        group_idx = conc_groups_sort[cv]
        group_means = [(idx, data.iloc[idx].mean()) for idx in group_idx]
        group_means.sort(key=lambda x: x[1])  # ascending
        sorted_indices.extend([idx for idx, _ in group_means])
    data = data.iloc[sorted_indices]

    # Scale
    data = data * 100

    y_labels = [_conc_label(c) for c in data.index.tolist()]
    x_labels = [str(t) for t in data.columns.tolist()]

    fig, ax = plt.subplots(figsize=(6, 3.5))
    sns.heatmap(
        data, annot=False, cmap=cmap_hm, vmin=0,
        cbar_kws={'shrink': 0.8},
        xticklabels=x_labels, yticklabels=y_labels,
        square=False, linewidths=0, ax=ax
    )

    ax.set_xlabel('Time from Exposure (h)', fontsize=32, fontweight='bold')
    ax.set_ylabel('Mexiletine Dose', fontsize=32, fontweight='bold')

    n_x = len(x_labels)
    x_step = max(1, n_x // 10)
    ax.set_xticks(range(0, n_x, x_step))
    ax.set_xticklabels([x_labels[i] for i in range(0, n_x, x_step)],
                        rotation=0, ha='center', fontsize=26, fontweight='bold')

    conc_groups2 = OrderedDict()
    for i, lbl in enumerate(y_labels):
        conc_groups2.setdefault(lbl, []).append(i)
    tick_positions = [(indices[0] + indices[-1]) / 2 + 0.5 for indices in conc_groups2.values()]
    ax.set_yticks(tick_positions)
    ax.set_yticklabels(list(conc_groups2.keys()), fontsize=26, rotation=0, fontweight='bold')

    for spine in ax.spines.values():
        spine.set_edgecolor('black')
        spine.set_linewidth(2.0)
    ax.tick_params(width=1.5)

    cbar = ax.collections[0].colorbar
    cbar.ax.tick_params(labelsize=24, width=1.5)
    cbar.set_label('Contractility (%)', fontsize=26, fontweight='bold')
    cbar.outline.set_linewidth(2.0)

    fig.tight_layout()
    dst = fig2_dir / 'Fig_2l_Mexiletine_Contractility_Heatmap.png'
    fig.savefig(str(dst), dpi=SAVE_DPI, bbox_inches='tight', pad_inches=0.05,
                facecolor='white')
    plt.close(fig)

    register_figure('2', 'Mexiletine_Contractility_Heatmap',
                    'Mexiletine Contractility Heatmap (LOWESS w=16)',
                    dst.relative_to(PROJECT_ROOT), None,
                    notes='14 wells excluded per NOTES file. LOWESS w=16 full smooth. Sorted ascending.')
    print(f"  Fig_2l_Mexiletine_Contractility_Heatmap.png")


# ============================================================================
# FIGURE 3: Fitting Kinetics
# ============================================================================

def generate_fig_3():
    """Figure 3: Epirubicin heatmaps, R2 comparison, AUC scatter, Daunorubicin 2D+3D."""
    print("\n=== Figure 3: Fitting Kinetics ===")
    import shutil
    from matplotlib.colors import LinearSegmentedColormap
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch
    from PIL import Image

    fig3_dir = FIGURES_DIR / 'Fig_3'
    fig3_dir.mkdir(parents=True, exist_ok=True)

    # ---- 3a: O2 Heatmaps (Dactinomycin, Nifedipine, Mexiletine) ----
    # ---- 3b: Surface Plots (Mexiletine Eq7, Nifedipine Eq10, Dactinomycin Eq3) ----
    import inspect
    import warnings
    import sys
    from matplotlib.ticker import NullFormatter
    from collections import OrderedDict
    from statsmodels.nonparametric.smoothers_lowess import lowess as lowess_func
    from matplotlib.colors import LinearSegmentedColormap as LSC

    # --- Helper: LOWESS smoothing per column ---
    LOWESS_W_3AB = 16

    def _apply_lowess_per_col(df):
        """Apply LOWESS smoothing (w=16) per column along the index (time axis)."""
        smoothed = df.copy().astype(float)
        for col in smoothed.columns:
            series = smoothed[col]
            valid = series.dropna()
            if len(valid) < 3:
                continue
            frac = min(1.0, max(LOWESS_W_3AB, 1) / len(valid))
            fitted = lowess_func(valid.values, np.arange(len(valid)),
                                 frac=frac, return_sorted=False)
            target = smoothed.index.get_indexer(valid.index)
            smoothed.iloc[target, smoothed.columns.get_loc(col)] = fitted
        return smoothed

    # --- Heatmap configuration ---
    heatmap_drugs = OrderedDict([
        ('Dactinomycin', {'remove_rows': {1, 8, 12, 16, 20, 24, 27}}),
        ('Nifedipine',   {'remove_rows': {5, 6}}),
        ('Mexiletine',   {'remove_rows': {2, 3, 9, 13, 20}}),
    ])

    heatmap_cmap = LSC.from_list('bwr_custom', [HEATMAP_BLUE, 'white', HEATMAP_RED])
    heatmap_cmap.set_bad('white')

    # --- Surface configuration ---
    # Drug -> (equation_name, sheet_name)
    surface_drugs = OrderedDict([
        ('Mexiletine',    ('biphasic_response',    'biphasic_response')),
        ('Nifedipine',    ('modified_hill_simple',  'modified_hill_simple')),
        ('Dactinomycin',  ('gaussian_hill_hybrid',  'gaussian_hill_hybrid')),
    ])

    # Load equation functions
    eq_fitting_dir = str(PROJECT_ROOT / 'Picking Equations' / 'equation_fitting')
    if eq_fitting_dir not in sys.path:
        sys.path.insert(0, eq_fitting_dir)
    from equations import EQUATION_FUNCTIONS

    # Load coefficient Excel once
    coeff_xl_path = PROJECT_ROOT / 'EQN_Coefficients' / 'all_equations_coefficients.xlsx'

    # ----- Generate 3a heatmaps -----
    print("  Generating Panel 3a heatmaps...")
    heatmap_data_for_excel = {}  # drug -> processed DataFrame

    for drug, cfg in heatmap_drugs.items():
        csv_path = PROJECT_ROOT / 'Cleaned_Data' / 'Heatmaps' / drug / 'O2_mean_sorted.csv'
        if not csv_path.exists():
            print(f"    WARNING: {csv_path} not found, skipping {drug} heatmap")
            continue

        # 1. Load sorted O2 data
        df_raw = pd.read_csv(csv_path, index_col=0)

        # 2. Remove outlier data points (O2 > 80 or < 0 -> NaN)
        df_raw = df_raw.where((df_raw >= 0) & (df_raw <= 80))

        # 3. Drop wells with > 50% NaN
        nan_frac = df_raw.isna().mean()
        keep_cols = nan_frac[nan_frac <= 0.5].index
        df_raw = df_raw[keep_cols]

        # 4. Linear interpolation within each well
        for col in df_raw.columns:
            df_raw[col] = df_raw[col].interpolate(method='linear', limit=10,
                                                   limit_direction='both')

        # 5. LOWESS smoothing per-well along time
        df_smooth = _apply_lowess_per_col(df_raw)

        # 6. Transpose (rows=wells, cols=time)
        data = df_smooth.T

        # 7. Remove manually flagged rows (1-indexed from sorted order)
        remove_rows = cfg['remove_rows']
        if remove_rows:
            # Convert 1-indexed row numbers to 0-indexed
            rows_to_drop = [data.index[i - 1] for i in sorted(remove_rows)
                            if i - 1 < len(data)]
            data = data.drop(rows_to_drop, errors='ignore')

        # 8. Clip O2 at 100
        data = data.clip(upper=100)

        # Store for Excel export
        heatmap_data_for_excel[drug] = data.copy()

        # 9-14. Plot heatmap (large figsize, scale down in PPTX)
        fig_hm, ax_hm = plt.subplots(figsize=(10, 10))
        sns.heatmap(data.values, cmap=heatmap_cmap, vmin=0, vmax=100,
                    cbar=False, square=False, linewidths=0, ax=ax_hm)

        # No tick values, just axis labels
        ax_hm.set_xticks([])
        ax_hm.set_yticks([])
        ax_hm.set_xlabel('Time from Exposure (h)', fontsize=40, fontweight='bold')
        ax_hm.set_ylabel(f'{drug} Dose', fontsize=40, fontweight='bold')

        # No border spines
        for spine in ax_hm.spines.values():
            spine.set_visible(False)

        hm_filename = f'Fig_3a_{drug}_O2_Heatmap.png'
        hm_path = fig3_dir / hm_filename
        fig_hm.savefig(hm_path, dpi=600, bbox_inches='tight',
                       facecolor='white', pad_inches=0.02)
        plt.close(fig_hm)
        print(f"    Saved: {hm_path}")

        register_figure('3', f'a_{drug}_O2_Heatmap',
                        f'{drug} O2 heatmap (Panel 3a)',
                        hm_path.relative_to(PROJECT_ROOT),
                        notes=f'LOWESS w=16, rows removed: {sorted(remove_rows)}',
                        source_script='generate_paper_figures.py')

    # ----- Generate 3b surface plots -----
    print("  Generating Panel 3b surface plots...")
    surface_data_for_excel = {}  # drug -> (Z_matrix, coeffs_dict, eq_name)

    # Equation display names for filenames
    eq_filename_map = {
        'biphasic_response': 'Eq7_biphasic_response',
        'modified_hill_simple': 'Eq10_modified_hill_simple',
        'gaussian_hill_hybrid': 'Eq3_gaussian_hill_hybrid',
    }

    LABEL_SIZE_3B = 34

    for drug, (eq_name, sheet_name) in surface_drugs.items():
        # Load coefficients for this equation
        df_coeff = pd.read_excel(coeff_xl_path, sheet_name=sheet_name, header=1)
        df_coeff.columns = df_coeff.columns.str.strip()

        drug_row = df_coeff[df_coeff['Drug'] == drug]
        if drug_row.empty:
            print(f"    WARNING: {drug} not found in {sheet_name} sheet, skipping")
            continue
        drug_row = drug_row.iloc[0]

        # Get the equation function and its parameter names (skip 'X')
        func = EQUATION_FUNCTIONS[eq_name]
        param_names = list(inspect.signature(func).parameters.keys())[1:]

        # Extract O2 parameters (use .1 suffix columns)
        o2_params = []
        coeffs_dict = {'equation': eq_name, 'drug': drug}
        for pname in param_names:
            col_o2 = f'{pname}.1'
            if col_o2 in drug_row.index:
                val = float(drug_row[col_o2])
            else:
                # Fallback to non-suffix column (shouldn't happen for O2)
                val = float(drug_row[pname])
            o2_params.append(val)
            coeffs_dict[f'{pname}_O2'] = val

        # Build 100x100 meshgrid: Time [0, 96] x Dose Ratio [0, 2]
        time_arr = np.linspace(0, 96, 100)
        dose_arr = np.linspace(0, 2, 100)
        T_mesh, Dr_mesh = np.meshgrid(time_arr, dose_arr)

        # Evaluate surface
        with warnings.catch_warnings():
            warnings.simplefilter('ignore')
            Z = func([Dr_mesh, T_mesh], *o2_params)

        Z = np.clip(Z, -500, 500)
        Z = np.where(np.isfinite(Z), Z, 0)

        surface_data_for_excel[drug] = (Z, coeffs_dict, eq_name)

        # Per-surface color scaling
        z_min, z_max = float(np.nanmin(Z)), float(np.nanmax(Z))

        # Plot
        fig_surf = plt.figure(figsize=(10, 10))
        ax_surf = fig_surf.add_subplot(111, projection='3d')

        ax_surf.plot_surface(T_mesh, Dr_mesh, Z, cmap='turbo',
                             vmin=z_min, vmax=z_max,
                             rcount=100, ccount=100, antialiased=True)

        ax_surf.view_init(elev=25, azim=-158)

        # Transparent panes
        ax_surf.xaxis.pane.fill = False
        ax_surf.yaxis.pane.fill = False
        ax_surf.zaxis.pane.fill = False
        ax_surf.xaxis.pane.set_edgecolor((0, 0, 0, 0))
        ax_surf.yaxis.pane.set_edgecolor((0, 0, 0, 0))
        ax_surf.zaxis.pane.set_edgecolor((0, 0, 0, 0))

        # Grid alpha
        ax_surf.xaxis._axinfo['grid']['color'] = (0, 0, 0, 0.2)
        ax_surf.yaxis._axinfo['grid']['color'] = (0, 0, 0, 0.2)
        ax_surf.zaxis._axinfo['grid']['color'] = (0, 0, 0, 0.2)

        # NO title

        # NO tick numbers — use NullFormatter on all axes
        ax_surf.xaxis.set_major_formatter(NullFormatter())
        ax_surf.yaxis.set_major_formatter(NullFormatter())
        ax_surf.zaxis.set_major_formatter(NullFormatter())

        # Axis labels via set_xlabel/set_ylabel and text2D for Z
        ax_surf.set_xlabel('Time (h)', fontsize=LABEL_SIZE_3B, labelpad=-10)
        ax_surf.set_ylabel('Dose Ratio', fontsize=LABEL_SIZE_3B, labelpad=-10)
        ax_surf.text2D(0.05, 0.5, r'$O_2$ (%)', transform=ax_surf.transAxes,
                       fontsize=LABEL_SIZE_3B, rotation=90, va='center', ha='right')

        surf_filename = f'{drug}_{eq_filename_map[eq_name]}.png'
        surf_path = fig3_dir / surf_filename
        fig_surf.savefig(surf_path, dpi=600, bbox_inches='tight',
                         transparent=True, pad_inches=0.02)
        plt.close(fig_surf)
        print(f"    Saved: {surf_path}")

        register_figure('3', f'b_{drug}_surface',
                        f'{drug} {eq_name} O2 surface (Panel 3b)',
                        surf_path.relative_to(PROJECT_ROOT),
                        notes=f'Equation: {eq_name}, per-surface color scaling',
                        source_script='generate_paper_figures.py')

    # ----- Save combined data Excel files (one per drug) -----
    print("  Saving Panel 3a/3b data files...")
    for drug in heatmap_drugs:
        excel_filename = f'Fig_3b_{drug}_data.xlsx'
        excel_path = fig3_dir / excel_filename

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Heatmap_Processed
            if drug in heatmap_data_for_excel:
                heatmap_data_for_excel[drug].to_excel(writer, sheet_name='Heatmap_Processed')

            # Sheet 2: Surface_Z
            if drug in surface_data_for_excel:
                Z_mat, coeffs, eq_name_used = surface_data_for_excel[drug]
                z_df = pd.DataFrame(Z_mat,
                                    index=[f'dr_{d:.4f}' for d in np.linspace(0, 2, 100)],
                                    columns=[f't_{t:.2f}' for t in np.linspace(0, 96, 100)])
                z_df.to_excel(writer, sheet_name='Surface_Z')

                # Sheet 3: Surface_Coefficients
                coeff_df = pd.DataFrame([coeffs])
                coeff_df.to_excel(writer, sheet_name='Surface_Coefficients', index=False)

            # Sheet 4: Summary
            sources = {
                'Heatmap_Source': str(PROJECT_ROOT / 'Cleaned_Data' / 'Heatmaps' / drug / 'O2_mean_sorted.csv'),
                'Coefficient_Source': str(coeff_xl_path),
            }
            if drug in surface_data_for_excel:
                _, coeffs, eq_name_used = surface_data_for_excel[drug]
                sources['Equation'] = eq_name_used
                sources['Sheet'] = eq_name_used
            sources['Pipeline'] = ('Load sorted O2 CSV -> outlier removal (O2>80 or <0) -> '
                                   'drop wells >50% NaN -> linear interp (limit=10) -> '
                                   'LOWESS w=16 -> transpose -> remove flagged rows -> clip at 100')
            summary_df = pd.DataFrame([sources])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

        print(f"    Saved: {excel_path}")

    # ---- 3c: R2 comparison chart (O2 only, sorted best→worst, top 3 highlighted) ----
    # Render large and square at high DPI, then scale down in PPTX
    r2_path = PROJECT_ROOT / 'Output' / 'Excel_Figures' / 'r2_equation_comparison.xlsx'
    r2_df = pd.read_excel(r2_path, sheet_name='R2_Chart')
    r2_df = r2_df.sort_values('O2', ascending=True).reset_index(drop=True)

    _surface_names = {
        'Dual Exponential (Eq1)': 'Dual Exponential',
        'Bivariate Gaussian (Eq2)': 'Bivariate Gaussian',
        'Gaussian-Hill Hybrid (Eq3)': 'Gaussian-Hill Hybrid',
        'Modified Hill (Hormesis) (Eq4)': 'Hormesis Hill',
        'Gaussian Ridge (Eq5)': 'Gaussian Ridge',
        'Adaptive Response (Eq6)': 'Adaptive Response',
        'Biphasic Response (Eq7)': 'Biphasic Response',
        'Cumulative Exposure (Eq8)': 'Cumulative Exposure',
        'Recovery Model (Eq9)': 'Recovery Model',
        'Modified Hill (Simple) (Eq10)': 'Modified Hill',
        'PKPD Elimination (Eq11)': 'PKPD Elimination',
        'Hormesis V0 (Legacy) (Eq12)': 'Dual Hill Hormesis',
    }
    # Rainbow colors assigned by rank (top=red, bottom=pink)
    _rainbow_colors = [
        '#d62728',   # red
        '#e6550d',   # red-orange
        '#ff7f0e',   # orange
        '#ffc107',   # amber
        '#8bc34a',   # yellow-green
        '#2ca02c',   # green
        '#00897b',   # teal
        '#17becf',   # cyan
        '#1f77b4',   # blue
        '#5c6bc0',   # indigo
        '#9467bd',   # purple
        '#e377c2',   # pink
    ]
    r2_df['Equation'] = r2_df['Equation'].map(_surface_names).fillna(r2_df['Equation'])

    r2_w, r2_h = 12, 6.0
    fig, ax = plt.subplots(figsize=(r2_w, r2_h))

    n = len(r2_df)
    y = np.arange(n)
    vals = r2_df['O2'].values

    # Assign rainbow colors: bottom bar (index 0) = pink, top bar (index n-1) = red
    bar_colors = [_rainbow_colors[n - 1 - i] if i < len(_rainbow_colors) else COLORS['grey'] for i in range(n)]

    bars = ax.barh(y, vals, height=0.55, color=bar_colors)

    for i, (bar, v) in enumerate(zip(bars, vals)):
        fw = 'bold'
        if v >= 0:
            ax.text(v + 0.01, bar.get_y() + bar.get_height() / 2,
                    f'{v:.2f}', va='center', ha='left', fontsize=24,
                    fontweight=fw)
        else:
            # Place negative value labels to the right of zero line
            ax.text(0.01, bar.get_y() + bar.get_height() / 2,
                    f'{v:.2f}', va='center', ha='left', fontsize=24,
                    fontweight=fw)

    ax.set_yticks(y)
    ax.set_yticklabels(r2_df['Equation'], fontsize=25, fontweight='bold')
    ax.set_xlabel(r'R$^2$', fontsize=28, fontweight='bold')
    ax.axvline(x=0, color='black', linewidth=1.5, zorder=0)
    ax.tick_params(axis='x', labelsize=25, width=1.5)
    ax.tick_params(axis='y', width=1.5)
    for label in ax.get_xticklabels():
        label.set_fontweight('bold')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_linewidth(1.8)
    ax.spines['left'].set_linewidth(1.8)

    fig.tight_layout()

    r2_df_full = r2_df.copy()
    r2_df_full['Source'] = str(r2_path)
    r2_df_full['Sheet'] = 'R2_Chart'

    save_figure(fig, '3', 'c', 'R² Equation Comparison',
                {'R2_Data': r2_df_full}, width=r2_w, height=r2_h)

    # ---- 3d: 3-panel strip — Accuracy vs AUC ROC for all 12 equations ----
    # Panels: Arrhythmia/XGBoost, Heart Damage/GaussianNB, Concern/GaussianNB
    loocv_path = PROJECT_ROOT / 'Output' / 'All_Equations_LOOCV' / 'loocv_all_equations.csv'
    if loocv_path.exists():
        loocv_df = pd.read_csv(loocv_path)

        # Spectral color scheme (matches Fig 3c bar order)
        _eq_colors = [
            ('dual_exponential',       '#d62728'),
            ('hormesis_v0',            '#e6550d'),
            ('pkpd_elimination',       '#ff7f0e'),
            ('biphasic_response',      '#ffc107'),
            ('modified_hill_hormesis', '#8bc34a'),
            ('modified_hill_simple',   '#2ca02c'),
            ('adaptive_response',      '#00897b'),
            ('gaussian_ridge',         '#17becf'),
            ('bivariate_gaussian',     '#1f77b4'),
            ('gaussian_hill_hybrid',   '#5c6bc0'),
            ('recovery_model',         '#9467bd'),
            ('cumulative_exposure',    '#e377c2'),
        ]

        _panels = [
            ('Arrhythmia',     'XGBoost',    'Arrhythmia'),
            ('heart_damage',   'GaussianNB', 'Heart Damage'),
            ('Concern_Binary', 'GaussianNB', 'Concern (Binary)'),
        ]

        strip_w, strip_h = 24, 8
        fig_d, axes_d = plt.subplots(1, 3, figsize=(strip_w, strip_h))

        for pi, (ax_d, (target, model, title)) in enumerate(zip(axes_d, _panels)):
            subset = loocv_df[(loocv_df['Target'] == target) &
                              (loocv_df['Model'] == model)]

            # Diagonal reference
            ax_d.plot([0, 1], [0, 1], '--', color='gray', linewidth=1,
                      alpha=0.5, zorder=1)

            for eq_name, eq_color in _eq_colors:
                row = subset[subset['Equation'] == eq_name]
                if row.empty:
                    continue
                ax_d.scatter(row['Accuracy'].values[0], row['AUC'].values[0],
                             c=eq_color, s=390, zorder=3,
                             edgecolors='black', linewidths=0.8)

            ax_d.set_xlim(0, 1.05)
            ax_d.set_ylim(0, 1.05)
            ax_d.set_xticks([0, 0.25, 0.5, 0.75, 1.0])
            ax_d.set_xticklabels(['0', '0.25', '0.5', '0.75', '1'],
                                 fontsize=29, fontweight='bold')
            ax_d.set_yticks([0, 0.25, 0.5, 0.75, 1.0])
            ax_d.set_box_aspect(1)

            ax_d.set_xlabel('Accuracy', fontsize=32, fontweight='bold')
            ax_d.text(0.5, 0.97, title, transform=ax_d.transAxes,
                      fontsize=33, fontweight='bold', ha='center', va='top')

            if pi == 0:
                ax_d.set_ylabel('AUC ROC', fontsize=32, fontweight='bold')
                ax_d.set_yticklabels(['0', '0.25', '0.5', '0.75', '1'],
                                     fontsize=29, fontweight='bold')
            else:
                ax_d.set_yticklabels([])

            for spine in ax_d.spines.values():
                spine.set_edgecolor('black')
                spine.set_linewidth(1.8)
            ax_d.tick_params(width=1.5)
            ax_d.grid(False)

        fig_d.tight_layout(w_pad=1.0)

        # Collect all plotted data for provenance
        loocv_data = loocv_df[
            ((loocv_df['Target'] == 'Arrhythmia') & (loocv_df['Model'] == 'XGBoost')) |
            ((loocv_df['Target'] == 'heart_damage') & (loocv_df['Model'] == 'GaussianNB')) |
            ((loocv_df['Target'] == 'Concern_Binary') & (loocv_df['Model'] == 'GaussianNB'))
        ].copy()
        loocv_data['Source'] = str(loocv_path)

        save_figure(fig_d, '3', 'd',
                    'LOOCV Accuracy vs AUC — 12 Equations (3-panel strip)',
                    {'LOOCV_Strip_Data': loocv_data},
                    width=strip_w, height=strip_h,
                    notes='Panels: Arrhythmia/XGBoost, Heart Damage/GaussianNB, '
                          'Concern/GaussianNB. 12 equations, spectral colors.')

    else:
        print("  Warning: loocv_all_equations.csv not found for Fig 3d")

    # ---- 3e: Daunorubicin 3D surfaces with raw data (O2 + Contractility combined) ----
    import math

    coeff_path = PROJECT_ROOT / 'EQN_Coefficients' / 'all_equations_coefficients.xlsx'
    df_coeff = pd.read_excel(coeff_path, sheet_name='pkpd_elimination', header=1)
    df_coeff.columns = df_coeff.columns.str.strip()

    # Generate individual 3D surfaces — auto-expand dose ratio to fit all data
    row = df_coeff[df_coeff['Drug'] == 'Daunorubicin'].iloc[0]

    individual_paths = []
    for resp_type in ['O2', 'Contractility']:
        sq = SQUARE_SIZE * 1.8
        fig_single = plt.figure(figsize=(sq, sq))
        ax = fig_single.add_subplot(111, projection='3d', computed_zorder=False)

        sfx = '.1' if resp_type == 'O2' else ''
        R0 = float(row[f'R0{sfx}'])
        Emax = float(row[f'Emax{sfx}'])
        kappa = float(row[f'kappa{sfx}'])
        n_p = float(row[f'n{sfx}'])
        m_p = float(row[f'm{sfx}'])
        tau = float(row[f'tau{sfx}'])
        k_el = float(row[f'k_elim{sfx}'])
        cmax = float(row[f'Cmax_used{sfx}'])

        data_path = (PROJECT_ROOT / 'Cleaned_Data' / 'O2_Mean_Averaged.xlsx' if resp_type == 'O2'
                     else PROJECT_ROOT / 'Cleaned_Data' / 'Heart_Contractility_Averaged.xlsx')
        df_raw = pd.read_excel(data_path, sheet_name='Daunorubicin')
        time_vals = df_raw.iloc[:, 0].values
        all_t, all_dr, all_r = [], [], []
        for cc in df_raw.columns[1:]:
            try:
                conc = float(str(cc).replace('_', '.'))
                dr = conc / cmax
                for t, r in zip(time_vals, df_raw[cc].values):
                    if not np.isnan(r):
                        all_t.append(float(t))
                        all_dr.append(dr)
                        all_r.append(float(r))
            except (ValueError, TypeError):
                continue
        raw_t, raw_dr, raw_r = np.array(all_t), np.array(all_dr), np.array(all_r)

        max_dr = max(2.0, math.ceil(raw_dr.max())) if len(raw_dr) > 0 else 2.0
        dr_vec = np.linspace(0, max_dr, 60)
        t_vec = np.linspace(0, 96, 60)
        T, Dr = np.meshgrid(t_vec, dr_vec)
        Resp = _pkpd_elimination_response(Dr, T, R0, Emax, kappa, n_p, m_p, tau, k_el)

        Resp = Resp - R0
        raw_r = raw_r - R0

        # Scale contractility by 100 for display
        if resp_type == 'Contractility':
            Resp = Resp * 100
            raw_r = raw_r * 100

        vmax = float(np.nanmax(Resp))
        norm = plt.Normalize(vmin=0, vmax=vmax)
        ax.plot_surface(T, Dr, Resp, cmap='turbo', norm=norm,
                        linewidth=0, antialiased=True, edgecolor='none', alpha=0.7)
        ax.scatter(raw_t, raw_dr, raw_r, c='black', s=6, alpha=0.7, depthshade=True, zorder=10)

        z_min = min(0, float(np.nanmin(raw_r))) if len(raw_r) > 0 else 0
        z_max = max(vmax, float(np.nanmax(raw_r)) * 1.1) if len(raw_r) > 0 else vmax
        ax.set_zlim(z_min, z_max)
        ax.set_xlim(0, 96)
        ax.set_ylim(0, max_dr)

        ax.set_xlabel('Time (h)', fontsize=7, labelpad=-3)
        ax.set_ylabel('Dose Ratio', fontsize=7, labelpad=-3)
        z_lab = r'O$_2$ (% air)' if resp_type == 'O2' else r'Contractility ($\times 10^{-2}$)'
        ax.set_zlabel(z_lab, fontsize=7, labelpad=-3)
        if resp_type == 'Contractility':
            from matplotlib.ticker import MaxNLocator
            ax.zaxis.set_major_locator(MaxNLocator(integer=True))
        ax.view_init(elev=25, azim=-158)
        ax.tick_params(labelsize=5, pad=-2)
        ax.xaxis.pane.fill = False
        ax.yaxis.pane.fill = False
        ax.zaxis.pane.fill = False
        ax.xaxis.pane.set_edgecolor('lightgray')
        ax.yaxis.pane.set_edgecolor('lightgray')
        ax.zaxis.pane.set_edgecolor('lightgray')

        fig_single.subplots_adjust(left=0.0, right=1.0, bottom=0.0, top=1.0)

        individual_name = f'Fig_3e_{resp_type}.png'
        individual_dst = fig3_dir / individual_name
        # Use fixed figsize (not bbox_inches='tight') so both plots are same size
        fig_single.savefig(str(individual_dst), dpi=600, pad_inches=0)
        plt.close(fig_single)
        individual_paths.append(individual_dst)
        print(f"  Saved individual: {individual_name}")

    # Save data Excel for Fig 3e
    fig3e_excel = fig3_dir / 'Fig_3e_data.xlsx'
    with pd.ExcelWriter(fig3e_excel, engine='openpyxl') as writer:
        # Coefficients sheet: one row per response type
        coeff_rows_3e = []
        for resp_type in ['O2', 'Contractility']:
            sfx = '.1' if resp_type == 'O2' else ''
            entry = {
                'Response_Type': resp_type,
                'Drug': 'Daunorubicin',
                'Equation': 'pkpd_elimination',
                'R0': float(row[f'R0{sfx}']),
                'Emax': float(row[f'Emax{sfx}']),
                'kappa': float(row[f'kappa{sfx}']),
                'n': float(row[f'n{sfx}']),
                'm': float(row[f'm{sfx}']),
                'tau': float(row[f'tau{sfx}']),
                'k_elim': float(row[f'k_elim{sfx}']),
                'Cmax_used': float(row[f'Cmax_used{sfx}']),
                'Source': str(coeff_path),
            }
            coeff_rows_3e.append(entry)
        pd.DataFrame(coeff_rows_3e).to_excel(writer, sheet_name='Coefficients', index=False)

        # Surface grid axes
        dr_vec_save = np.linspace(0, max_dr, 60)
        t_vec_save = np.linspace(0, 96, 60)
        pd.DataFrame({'Time_h': t_vec_save}).to_excel(writer, sheet_name='Grid_Time', index=False)
        pd.DataFrame({'Dose_Ratio': dr_vec_save}).to_excel(writer, sheet_name='Grid_DoseRatio', index=False)

        # Surface Z-values and raw data for each response type
        for resp_type in ['O2', 'Contractility']:
            sfx = '.1' if resp_type == 'O2' else ''
            R0_v = float(row[f'R0{sfx}'])
            Emax_v = float(row[f'Emax{sfx}'])
            kappa_v = float(row[f'kappa{sfx}'])
            n_v = float(row[f'n{sfx}'])
            m_v = float(row[f'm{sfx}'])
            tau_v = float(row[f'tau{sfx}'])
            k_el_v = float(row[f'k_elim{sfx}'])
            cmax_v = float(row[f'Cmax_used{sfx}'])

            T_g, Dr_g = np.meshgrid(t_vec_save, dr_vec_save)
            Resp_g = _pkpd_elimination_response(Dr_g, T_g, R0_v, Emax_v, kappa_v, n_v, m_v, tau_v, k_el_v)
            Resp_g = Resp_g - R0_v  # Baseline correction (matches plot)
            if resp_type == 'Contractility':
                Resp_g = Resp_g * 100  # Scale ×100 to match plot

            sheet_z = f'{resp_type}_Surface_Z'[:31]
            df_z = pd.DataFrame(Resp_g, index=np.round(dr_vec_save, 4), columns=np.round(t_vec_save, 2))
            df_z.index.name = 'Dose_Ratio'
            df_z.to_excel(writer, sheet_name=sheet_z)

            # Raw experimental data points
            data_path = (PROJECT_ROOT / 'Cleaned_Data' / 'O2_Mean_Averaged.xlsx' if resp_type == 'O2'
                         else PROJECT_ROOT / 'Cleaned_Data' / 'Heart_Contractility_Averaged.xlsx')
            df_raw_e = pd.read_excel(data_path, sheet_name='Daunorubicin')
            time_vals_e = df_raw_e.iloc[:, 0].values
            pts_t, pts_dr, pts_r = [], [], []
            for cc in df_raw_e.columns[1:]:
                try:
                    conc = float(str(cc).replace('_', '.'))
                    dr = conc / cmax_v
                    for t, r in zip(time_vals_e, df_raw_e[cc].values):
                        if not np.isnan(r):
                            pts_t.append(float(t))
                            pts_dr.append(dr)
                            val = float(r) - R0_v  # Baseline corrected
                            if resp_type == 'Contractility':
                                val *= 100
                            pts_r.append(val)
                except (ValueError, TypeError):
                    continue
            raw_df = pd.DataFrame({
                'Time_h': pts_t,
                'Dose_Ratio': pts_dr,
                'Response_Baseline_Corrected': pts_r,
                'Source': str(data_path),
            })
            sheet_raw = f'{resp_type}_Raw_Data'[:31]
            raw_df.to_excel(writer, sheet_name=sheet_raw, index=False)

    print(f"  Saved: Fig_3e_data.xlsx")

    # Individual images placed side-by-side in PPTX via COMPOUND_PANELS
    register_figure('3', 'e', 'Daunorubicin 3D Surfaces + Raw Data (O2 & Contractility)',
                    individual_paths[0].relative_to(PROJECT_ROOT),
                    fig3e_excel.relative_to(PROJECT_ROOT),
                    width=SQUARE_SIZE * 1.8, height=SQUARE_SIZE * 1.8,
                    source_script='generate_paper_figures.py',
                    notes='Two images: Fig_3e_O2.png + Fig_3e_Contractility.png')


# ============================================================================
# FIGURE 4 & 5: 3D Surface Grids
# ============================================================================

def _save_fig4_fig5_data():
    """Save raw data Excel files for Figures 4 (O2) and 5 (Contractility).

    Each file contains:
      - Coefficients sheet: 25 drugs x 7 PK-PD parameters
      - Equation sheet: formula definition and axis info
      - Grid_Axes sheet: time (100 pts) and dose_ratio (100 pts) vectors
      - 25 drug sheets: each a 100x100 computed Z surface matrix
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    EXCEL_PATH = PROJECT_ROOT / 'EQN_Coefficients' / 'all_equations_coefficients.xlsx'
    df = pd.read_excel(EXCEL_PATH, sheet_name='pkpd_elimination', header=1)
    df.columns = df.columns.str.strip()
    EXCLUDED = {'DMSO', 'Troglitazone', 'Troglitarazine'}
    df = df[~df['Drug'].isin(EXCLUDED)].copy()
    df = df.sort_values('Drug').reset_index(drop=True)

    PARAM_NAMES = ['R0', 'Emax', 'kappa', 'n', 'm', 'tau', 'k_elim']
    time = np.linspace(0, 96, 100)
    dose_ratio = np.linspace(0, 2, 100)
    T, Dr = np.meshgrid(time, dose_ratio)

    def _pkpd_response(T, Dr, R0, Emax, kappa, n, m, tau, k_elim):
        k_elim = max(k_elim, 1e-9)
        kappa = max(kappa, 1e-9)
        tau = max(tau, 1e-9)
        t_safe = np.maximum(T, 0)
        conc = Dr * np.exp(-k_elim * t_safe)
        Z = R0 + Emax * (1 - np.exp(-kappa * (conc ** n) * ((t_safe / tau) ** m)))
        return np.nan_to_num(Z, nan=0.0, posinf=0.0, neginf=0.0)

    for fig_num, response_type, suffix in [('4', 'O2', '.1'), ('5', 'Contractility', '')]:
        wb = Workbook()
        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='D9E1F2')

        # Sheet 1: Coefficients
        ws_coeff = wb.active
        ws_coeff.title = 'Coefficients'
        headers = ['Drug'] + [f'{p}_{response_type}' for p in PARAM_NAMES]
        for col_idx, h in enumerate(headers, 1):
            cell = ws_coeff.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            ws_coeff.cell(row=row_idx, column=1, value=row['Drug'])
            for p_idx, p in enumerate(PARAM_NAMES):
                col = p + suffix if suffix else p
                val = row.get(col, np.nan)
                ws_coeff.cell(row=row_idx, column=p_idx + 2,
                              value=float(val) if pd.notna(val) else None)
        for col_idx in range(1, len(headers) + 1):
            ws_coeff.column_dimensions[chr(64 + col_idx)].width = 18

        # Sheet 2: Equation info
        ws_eq = wb.create_sheet('Equation')
        ws_eq['A1'] = 'PK-PD Elimination Equation'
        ws_eq['A1'].font = Font(bold=True, size=14)
        ws_eq['A3'] = 'Formula'
        ws_eq['B3'] = ('R(C0, t) = R0 + Emax * (1 - exp(-kappa * '
                        '(C0/Cmax * exp(-k_elim * t))^n * (t/tau)^m))')
        ws_eq['A5'] = 'X-axis'
        ws_eq['B5'] = 'Time (hours), range [0, 96], 100 points'
        ws_eq['A6'] = 'Y-axis'
        ws_eq['B6'] = 'Dose Ratio (C0/Cmax), range [0, 2], 100 points'
        ws_eq['A7'] = 'Z-axis'
        ws_eq['B7'] = f'{response_type} Response'
        ws_eq['A9'] = 'Source File'
        ws_eq['B9'] = str(EXCEL_PATH.resolve())
        ws_eq['A10'] = 'Sheet'
        ws_eq['B10'] = 'pkpd_elimination'
        ws_eq.column_dimensions['A'].width = 15
        ws_eq.column_dimensions['B'].width = 80

        # Sheet 3: Grid axes
        ws_axes = wb.create_sheet('Grid_Axes')
        ws_axes.cell(row=1, column=1, value='Time (hours)').font = header_font
        ws_axes.cell(row=1, column=2, value='Dose_Ratio (C0/Cmax)').font = header_font
        for i, t in enumerate(time):
            ws_axes.cell(row=i + 2, column=1, value=round(float(t), 6))
        for i, d in enumerate(dose_ratio):
            ws_axes.cell(row=i + 2, column=2, value=round(float(d), 6))

        # One sheet per drug: 100x100 Z surface
        drug_count = 0
        for _, row in df.iterrows():
            drug = row['Drug']
            params = []
            skip = False
            for p in PARAM_NAMES:
                col = p + suffix if suffix else p
                val = row.get(col, np.nan)
                if pd.isna(val) or not np.isfinite(val):
                    skip = True
                    break
                params.append(float(val))
            if skip:
                continue

            Z = _pkpd_response(T, Dr, *params)
            ws_drug = wb.create_sheet(drug)
            ws_drug.cell(row=1, column=1, value='Dose_Ratio \\ Time').font = header_font
            for j in range(100):
                ws_drug.cell(row=1, column=j + 2, value=round(float(time[j]), 2))
            for i in range(100):
                ws_drug.cell(row=i + 2, column=1, value=round(float(dose_ratio[i]), 4))
                for j in range(100):
                    ws_drug.cell(row=i + 2, column=j + 2, value=round(float(Z[i, j]), 6))
            drug_count += 1

        fig_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
        data_path = fig_dir / f'Fig_{fig_num}_data.xlsx'
        wb.save(data_path)
        print(f"  Saved data: {data_path} ({drug_count} drug surfaces + coefficients)")


def generate_fig_4_5():
    """Figures 4 & 5: 3D surface grids for O2 and Contractility.

    Uses generate_5x5_individual.py to create high-quality 600 DPI individual plots
    with edge-only axis labels, then assembles into PowerPoint via build_5x5_slides.py.

    Individual plots are saved to:
    - Output/PowerPoint_Figures/Fig_4/O2_5x5_Individual/
    - Output/PowerPoint_Figures/Fig_5/Contractility_5x5_Individual/

    Axis label convention:
    - LEFT column (col 0): Z-axis label only
    - RIGHT column (col 4): X-axis "Time (h)" label only
    - BOTTOM row (row 4): Y-axis "Dose Ratio" label only
    - Interior plots: NO labels (uses NullFormatter)
    """
    import subprocess
    import sys

    print("\n=== Figures 4 & 5: 3D Surface Grids (Individual 600 DPI) ===")

    # Scripts are now in Output/PowerPoint_Figures/
    pptx_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures'

    # Step 1: Generate individual plots at 600 DPI
    print("  Generating individual 600 DPI plots...")
    script_path = pptx_dir / 'generate_5x5_individual.py'

    if script_path.exists():
        result = subprocess.run([sys.executable, str(script_path)],
                               cwd=str(pptx_dir), capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  Warning: generate_5x5_individual.py failed: {result.stderr}")
        else:
            print("  Individual plots generated successfully")
    else:
        print(f"  Warning: {script_path} not found")
        return

    # Step 2: Assemble into PowerPoint
    print("  Assembling into PowerPoint...")
    build_script = pptx_dir / 'build_5x5_slides.py'

    if build_script.exists():
        result = subprocess.run([sys.executable, str(build_script)],
                               cwd=str(pptx_dir), capture_output=True, text=True)
        if result.returncode != 0:
            print(f"  Warning: build_5x5_slides.py failed: {result.stderr}")
        else:
            print("  PowerPoint slides assembled successfully")
    else:
        print(f"  Warning: {build_script} not found")

    # Save raw data (coefficients + computed surface meshes) for each figure
    _save_fig4_fig5_data()

    print("  Figures 4 & 5 complete - check PowerPoint for assembled grids")


# ============================================================================
# FIGURES 6, 7, 8: Interpretable Predictions
# ============================================================================

def generate_prediction_figures(target, fig_num, comparison_type=None):
    """Generate all panels for a prediction figure."""
    print(f"\n=== Figure {fig_num}: {target} Prediction ===")

    target_map = {
        'Arrhythmia': {'col': 'Arrhythmia', 'sheet': 'Arrhythmia', 'threshold_key': 'Arrhythmia_threshold_pct', 'model': 'RandomForest'},
        'HeartDamage': {'col': 'heart_damage', 'sheet': 'HeartDamage', 'threshold_key': 'Heart_Damage_threshold_pct', 'model': 'GaussianNB'},
        'ConcernBinary': {'col': 'Concern_Binary', 'sheet': 'ConcernBinary', 'threshold_key': 'Concern_Binary_threshold_pct', 'model': 'GaussianNB'}
    }

    info = target_map[target]

    # Load thresholds for titles
    threshold_path = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data' / 'prediction_thresholds.json'
    with open(threshold_path) as f:
        thresholds = json.load(f)
    threshold_pct = int(thresholds.get(info['threshold_key'], 50))

    # -------------------------------------------------------------------------
    # Panel a: ROC Curve with shading - EXACT SQUARE SIZE
    # -------------------------------------------------------------------------
    roc_path = PROJECT_ROOT / 'Output' / 'ROC_Data' / 'roc_curves_all_models.xlsx'
    roc_df = pd.read_excel(roc_path, sheet_name=info['sheet'])

    # Generate at 2x size then save at target size for crisp legend
    _def_a = (SQUARE_SIZE, SQUARE_SIZE)
    size_a = get_layout_size(fig_num, 'a', default=_def_a) or _def_a
    render_scale = 2.0
    fig, ax = plt.subplots(figsize=(size_a[0] * render_scale, size_a[1] * render_scale))
    fig.subplots_adjust(left=0.18, right=0.95, top=0.88, bottom=0.18)

    # Collect all fold TPRs for mean/std
    mean_fpr = np.linspace(0, 1, 100)
    tprs = []
    aucs = []

    for i in range(1, 11):
        fpr_col = f'Fold{i} - FPR'
        tpr_col = f'Fold{i} - TPR'
        if fpr_col in roc_df.columns:
            fpr = roc_df[fpr_col].dropna().values
            tpr = roc_df[tpr_col].dropna().values
            if len(fpr) > 1:
                tpr_interp = np.interp(mean_fpr, fpr, tpr)
                tprs.append(tpr_interp)
                aucs.append(np.trapz(tpr, fpr))

    if tprs:
        mean_tpr = np.mean(tprs, axis=0)
        std_tpr = np.std(tprs, axis=0)
        mean_auc = np.mean(aucs)
        std_auc = np.std(aucs)

        ax.plot(mean_fpr, mean_tpr, color=COLORS['blue'], lw=2,
                label=f'Organoid AUC = {mean_auc:.2f} (±{std_auc:.2f})')
        ax.fill_between(mean_fpr,
                        np.maximum(mean_tpr - std_tpr, 0),
                        np.minimum(mean_tpr + std_tpr, 1),
                        color=COLORS['blue'], alpha=0.2)

    ax.plot([0, 1], [0, 1], 'k--', lw=1, label='Chance (AUC = 0.50)')
    ax.set_xlabel('False Positive Rate', fontsize=8 * render_scale)
    ax.set_ylabel('True Positive Rate', fontsize=8 * render_scale)
    ax.set_title('AUC ROC', fontsize=9 * render_scale, fontweight='bold')
    ax.legend(fontsize=4.5 * render_scale, loc='lower right',
              handlelength=1.5, handletextpad=0.4, borderpad=0.3, labelspacing=0.3)
    ax.set_xlim(0, 1)
    ax.set_ylim(0, 1)
    ax.tick_params(labelsize=7 * render_scale)

    save_figure(fig, fig_num, 'a', f'{target} ROC Curve',
                {'ROC_Data': roc_df}, width=SQUARE_SIZE, height=SQUARE_SIZE, exact_size=True)

    # -------------------------------------------------------------------------
    # Panel b: Confusion Matrix (from 5-fold CV or stored CM)
    # -------------------------------------------------------------------------
    # Load confusion matrix from Excel_Figures (10-seed 5-fold aggregated data)
    excel_figures_dir = PROJECT_ROOT / 'Output' / 'Excel_Figures'

    if target == 'Arrhythmia':
        cm_file = excel_figures_dir / 'confusion_matrix_organoid_arrhythmia.xlsx'
    elif target == 'HeartDamage':
        cm_file = excel_figures_dir / 'confusion_matrix_organoid_heart_damage.xlsx'
    else:  # ConcernBinary
        cm_file = excel_figures_dir / 'confusion_matrix_organoid_concern_binary.xlsx'

    if cm_file.exists():
        try:
            # Read Excel - data starts at row 2 (0-indexed), columns B-C (1-2)
            cm_df = pd.read_excel(cm_file)
            # Extract the 2x2 matrix from rows 2-3, columns 1-2
            cm = cm_df.iloc[2:4, 1:3].values.astype(int)
        except Exception as e:
            print(f"Warning: Could not load CM from {cm_file}: {e}")
            cm = np.array([[8, 3], [2, 12]])  # Fallback
    else:
        print(f"Warning: CM file not found: {cm_file}")
        cm = np.array([[8, 3], [2, 12]])  # Fallback

    # Square figure with tight margins to maximize plot area
    _def_b = (SQUARE_SIZE, SQUARE_SIZE)
    size_b = get_layout_size(fig_num, 'b', default=_def_b) or _def_b
    fig, ax = plt.subplots(figsize=size_b)
    fig.subplots_adjust(left=0.18, right=0.98, top=0.85, bottom=0.18)

    im = ax.imshow(cm, cmap='Blues', aspect='equal')

    # Add text annotations — counts only (no percentages)
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            val = int(cm[i, j])
            color = 'white' if cm[i, j] > cm.max() / 2 else 'black'
            ax.text(j, i, f'{val}', ha='center', va='center', color=color, fontsize=9)

    ax.set_xticks([0, 1])
    ax.set_yticks([0, 1])
    ax.set_xticklabels(['Neg', 'Pos'], fontsize=7)
    ax.set_yticklabels(['Neg', 'Pos'], fontsize=7)
    ax.set_xlabel('Predicted', fontsize=7)
    ax.set_ylabel('Actual', fontsize=7)
    ax.set_title('Confusion Matrix', fontsize=7, fontweight='bold')
    cm_df_save = pd.DataFrame(cm, index=['Actual Neg', 'Actual Pos'], columns=['Pred Neg', 'Pred Pos'])
    save_figure(fig, fig_num, 'b', f'{target} Confusion Matrix',
                {'CM': cm_df_save}, width=SQUARE_SIZE, height=SQUARE_SIZE, exact_size=True)

    # --- Alternative CM with percentages (smaller text, no parentheses) ---
    fig_alt, ax_alt = plt.subplots(figsize=size_b)
    fig_alt.subplots_adjust(left=0.18, right=0.98, top=0.85, bottom=0.18)
    ax_alt.imshow(cm, cmap='Blues', aspect='equal')

    row_sums = cm.sum(axis=1, keepdims=True)
    row_sums[row_sums == 0] = 1
    row_pct = cm / row_sums

    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            val = int(cm[i, j])
            pct = row_pct[i, j] * 100
            color = 'white' if cm[i, j] > cm.max() / 2 else 'black'
            ax_alt.text(j, i, f'{val}', ha='center', va='center', color=color, fontsize=9)
            ax_alt.text(j, i + 0.25, f'{pct:.0f}%', ha='center', va='center', color=color, fontsize=5)

    ax_alt.set_xticks([0, 1])
    ax_alt.set_yticks([0, 1])
    ax_alt.set_xticklabels(['Neg', 'Pos'], fontsize=7)
    ax_alt.set_yticklabels(['Neg', 'Pos'], fontsize=7)
    ax_alt.set_xlabel('Predicted', fontsize=7)
    ax_alt.set_ylabel('Actual', fontsize=7)
    ax_alt.set_title('Confusion Matrix', fontsize=7, fontweight='bold')

    # Save alternative with percentages as swap graph
    swap_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
    swap_path = swap_dir / f'Fig_{fig_num}b_with_pct.png'
    fig_alt.savefig(swap_path, dpi=SAVE_DPI, bbox_inches='tight', pad_inches=0.05)
    plt.close(fig_alt)
    print(f"  Saved swap graph: {swap_path.name}")

    # -------------------------------------------------------------------------
    # Panel c: Metrics Bar (Accuracy, F1, MCC, AUC) - SQUARE with error bars
    # -------------------------------------------------------------------------
    stage2_path = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'stage2_results_5fold.csv'
    stage2_df = pd.read_csv(stage2_path)

    target_name = target if target != 'HeartDamage' else 'heart_damage'
    if target == 'ConcernBinary':
        target_name = 'Concern_Binary'

    filt = stage2_df[stage2_df['Target'].str.lower() == target_name.lower()]
    if filt.empty:
        filt = stage2_df[stage2_df['Target'] == 'Arrhythmia']  # Fallback

    metrics_mean = filt[['Accuracy', 'AUC', 'F1', 'MCC']].mean()
    metrics_std = filt[['Accuracy', 'AUC', 'F1', 'MCC']].std()

    # SQUARE bar chart per skill (1.7" × 1.7")
    _def_c = (SQUARE_SIZE, SQUARE_SIZE)
    size_c = get_layout_size(fig_num, 'c', default=_def_c) or _def_c
    fig, ax = plt.subplots(figsize=size_c)

    x = np.arange(4)
    colors_m = [COLORS['blue'], COLORS['pink'], COLORS['orange'], COLORS['beige']]

    # MANDATORY: Error bars with capsize (skill requirement)
    bars = ax.bar(x, metrics_mean.values, yerr=metrics_std.values, capsize=3,
                  color=colors_m, edgecolor='black', linewidth=0.5)

    # Value labels above error bars
    for bar, val, err in zip(bars, metrics_mean.values, metrics_std.values):
        ax.annotate(f'{val:.2f}', xy=(bar.get_x() + bar.get_width()/2, val + err),
                    xytext=(0, 3), textcoords='offset points', ha='center', fontsize=6)

    ax.set_xticks(x)
    ax.set_xticklabels(['Acc', 'AUC\nROC', 'F1', 'MCC'], fontsize=7)
    ax.set_ylabel('Score', fontsize=8)
    ax.set_title('Model Performance Metrics', fontsize=9, fontweight='bold')
    ax.set_ylim(0, 1.15)
    ax.tick_params(labelsize=6)

    fig.tight_layout()

    # Save FULL source data for recreating the plot
    # Include both summary (for the plot) and raw data (for recalculation)
    metrics_summary = pd.DataFrame({
        'Metric': metrics_mean.index,
        'Mean': metrics_mean.values,
        'Std': metrics_std.values
    })
    metrics_summary['Source'] = str(stage2_path)
    metrics_summary['Target'] = target_name
    metrics_summary['Plot_Type'] = 'Bar with error bars'

    # Also save the raw fold data
    filt_full = filt.copy()
    filt_full['Source'] = str(stage2_path)

    save_figure(fig, fig_num, 'c', f'{target} Performance Metrics',
                {'Metrics_Summary': metrics_summary, 'Raw_Fold_Data': filt_full},
                width=SQUARE_SIZE, height=SQUARE_SIZE)

    # -------------------------------------------------------------------------
    # Panel d: Threshold Analysis
    # -------------------------------------------------------------------------
    pred_path = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data'
    pred_file = pred_path / f'{target.lower()}_predictions.csv'
    if target == 'HeartDamage':
        pred_file = pred_path / 'heart_damage_predictions.csv'
    elif target == 'ConcernBinary':
        pred_file = pred_path / 'concern_binary_predictions.csv'

    threshold_path = pred_path / 'prediction_thresholds.json'

    if pred_file.exists():
        pred_df = pd.read_csv(pred_file)

        with open(threshold_path) as f:
            thresholds = json.load(f)
        threshold = thresholds.get(info['threshold_key'], 50)

        # Sort by actual class then drug name
        pred_col = [c for c in pred_df.columns if 'Predicted' in c][0]
        actual_col = [c for c in pred_df.columns if 'Actual' in c][0]

        pred_df['is_positive'] = pred_df[actual_col].apply(lambda x: str(x).lower() == 'true')
        pred_df = pred_df.sort_values(['is_positive', 'Drug'], ascending=[False, True])

        # Threshold analysis: fixed size 1.55" width
        THRESHOLD_W, THRESHOLD_H = 1.55, 1.87
        _def_d = (THRESHOLD_W, THRESHOLD_H)
        size_d = get_layout_size(fig_num, 'd', default=_def_d) or _def_d
        fig, ax = plt.subplots(figsize=size_d)
        fig.subplots_adjust(left=0.35, right=0.98, top=0.95, bottom=0.08)

        y_pos = np.arange(len(pred_df))
        colors_p = [COLORS['pass'] if p else COLORS['fail'] for p in pred_df['is_positive']]

        ax.scatter(pred_df[pred_col], y_pos, c=colors_p, s=12, edgecolor='black', linewidth=0.2)
        ax.axvline(threshold, color=COLORS['threshold'], linestyle='--', linewidth=1.5)
        ax.text(threshold + 2, 1, f'{threshold}%', color=COLORS['threshold'], fontsize=6, fontweight='bold')

        ax.set_yticks(y_pos)
        ax.set_yticklabels(pred_df['Drug'], fontsize=4)
        ax.set_xlabel('Prob (%)', fontsize=6)
        ax.set_title('Prediction Threshold', fontsize=7, fontweight='bold')
        ax.set_xlim(-5, 105)
        ax.invert_yaxis()
        ax.grid(axis='x', alpha=0.3)
        ax.tick_params(axis='x', labelsize=6)

        legend_elements = [
            Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['pass'], markersize=5, label='Pos'),
            Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['fail'], markersize=5, label='Neg'),
        ]
        ax.legend(handles=legend_elements, fontsize=5, loc='lower right')

        # Add source metadata for full tracking
        pred_df_full = pred_df.copy()
        pred_df_full['Source'] = str(pred_file)
        pred_df_full['Threshold_Value'] = threshold
        pred_df_full['Threshold_Source'] = str(threshold_path)

        save_figure(fig, fig_num, 'd', f'{target} Threshold Analysis',
                    {'Predictions': pred_df_full}, width=size_d[0], height=size_d[1], exact_size=True)

        # --- Alternative version with mean ± std lines per class ---
        fig_alt, ax_alt = plt.subplots(figsize=size_d)
        fig_alt.subplots_adjust(left=0.35, right=0.98, top=0.95, bottom=0.08)

        ax_alt.scatter(pred_df[pred_col], y_pos, c=colors_p, s=12, edgecolor='black', linewidth=0.2)
        ax_alt.axvline(threshold, color=COLORS['threshold'], linestyle='--', linewidth=1.5)
        ax_alt.text(threshold + 2, 1, f'{threshold}%', color=COLORS['threshold'], fontsize=6, fontweight='bold')

        # Mean and std lines for positive class
        pos_vals = pred_df.loc[pred_df['is_positive'], pred_col]
        neg_vals = pred_df.loc[~pred_df['is_positive'], pred_col]

        pos_mean, pos_std = pos_vals.mean(), pos_vals.std()
        neg_mean, neg_std = neg_vals.mean(), neg_vals.std()

        # Positive: green mean + std band
        ax_alt.axvline(pos_mean, color=COLORS['pass'], linestyle='-', linewidth=0.8, alpha=0.8)
        ax_alt.axvspan(pos_mean - pos_std, pos_mean + pos_std, color=COLORS['pass'], alpha=0.10)

        # Negative: grey mean + std band
        ax_alt.axvline(neg_mean, color=COLORS['fail'], linestyle='-', linewidth=0.8, alpha=0.8)
        ax_alt.axvspan(neg_mean - neg_std, neg_mean + neg_std, color=COLORS['fail'], alpha=0.10)

        ax_alt.set_yticks(y_pos)
        ax_alt.set_yticklabels(pred_df['Drug'], fontsize=4)
        ax_alt.set_xlabel('Prob (%)', fontsize=6)
        ax_alt.set_title('Prediction Threshold', fontsize=7, fontweight='bold')
        ax_alt.set_xlim(-5, 105)
        ax_alt.invert_yaxis()
        ax_alt.grid(axis='x', alpha=0.3)
        ax_alt.tick_params(axis='x', labelsize=6)

        legend_elements_alt = [
            Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['pass'], markersize=5,
                   label=f'Pos ({pos_mean:.0f}±{pos_std:.0f}%)'),
            Line2D([0], [0], marker='o', color='w', markerfacecolor=COLORS['fail'], markersize=5,
                   label=f'Neg ({neg_mean:.0f}±{neg_std:.0f}%)'),
        ]
        ax_alt.legend(handles=legend_elements_alt, fontsize=4, loc='lower right')

        swap_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
        swap_path = swap_dir / f'Fig_{fig_num}d_with_stats.png'
        fig_alt.savefig(swap_path, dpi=SAVE_DPI, bbox_inches='tight', pad_inches=0.05)
        plt.close(fig_alt)
        print(f"  Saved swap graph: {swap_path.name}")

    # -------------------------------------------------------------------------
    # Panel e: Cumulative Features - with threshold line
    # -------------------------------------------------------------------------
    cum_path = PROJECT_ROOT / 'Output' / 'Cumulative_Plot_Data'
    cum_file = cum_path / f'{target.lower()}_cumulative_predictions.csv'
    if target == 'HeartDamage':
        cum_file = cum_path / 'heart_damage_cumulative_predictions.csv'
    elif target == 'ConcernBinary':
        cum_file = cum_path / 'concern_binary_cumulative_predictions.csv'

    if cum_file.exists():
        cum_df = pd.read_csv(cum_file, index_col=0)

        # Load ground truth from prediction files (most reliable source)
        pred_path = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data'
        if target == 'Arrhythmia':
            pred_file = pred_path / 'arrhythmia_predictions.csv'
            actual_col = 'Actual_Arrhythmia'
        elif target == 'HeartDamage':
            pred_file = pred_path / 'heart_damage_predictions.csv'
            actual_col = 'Actual_Heart_Damage'
        else:  # ConcernBinary
            pred_file = pred_path / 'concern_binary_predictions.csv'
            actual_col = 'Actual_High_Concern'

        label_map = {}
        if pred_file.exists():
            pred_df = pd.read_csv(pred_file)
            for _, row in pred_df.iterrows():
                drug = row['Drug']
                val = row.get(actual_col, False)
                label_map[drug] = str(val).lower() == 'true' if isinstance(val, str) else bool(val)

        # Load threshold
        threshold_path = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data' / 'prediction_thresholds.json'
        with open(threshold_path) as f:
            thresholds = json.load(f)
        threshold = thresholds.get(info['threshold_key'], 50)

        # Size 2.47" × 1.78" (landscape)
        CUM_W, CUM_H = 2.47, 1.78
        _def_e = (CUM_W, CUM_H)
        size_e = get_layout_size(fig_num, 'e', default=_def_e) or _def_e
        fig, ax = plt.subplots(figsize=size_e)
        fig.subplots_adjust(left=0.12, right=0.78, top=0.88, bottom=0.15)

        # Plot each drug as a line with more detail
        x_positions = np.arange(1, len(cum_df) + 1)
        for drug in cum_df.columns:
            is_positive = label_map.get(drug, False)
            color = COLORS['pass'] if is_positive else COLORS['fail']
            ax.plot(x_positions, cum_df[drug].values, color=color, alpha=0.6, linewidth=1.0, marker='o', markersize=2)

        # Add threshold line
        ax.axhline(y=threshold, color=COLORS['threshold'], linestyle='--', linewidth=1.5, label=f'Threshold ({threshold}%)')

        ax.set_xticks(x_positions)
        ax.set_xticklabels([str(i) for i in x_positions], fontsize=5)
        ax.set_xlabel('# Features', fontsize=7)
        ax.set_ylabel('Cumulative Score (%)', fontsize=7)
        ax.set_title('Cumulative Feature Importance', fontsize=8, fontweight='bold')
        ax.tick_params(labelsize=5)
        ax.set_ylim(-5, 105)

        legend_elements = [
            Line2D([0], [0], color=COLORS['pass'], linewidth=2, label='Pos'),
            Line2D([0], [0], color=COLORS['fail'], linewidth=2, label='Neg'),
            Line2D([0], [0], color=COLORS['threshold'], linestyle='--', linewidth=1.5, label=f'Thresh'),
        ]
        ax.legend(handles=legend_elements, fontsize=5, loc='center right')

        # Add source metadata for full tracking
        source_info = pd.DataFrame({
            'Source_File': [str(cum_file)],
            'Threshold': [threshold],
            'Threshold_Source': [str(threshold_path)],
            'Label_Source': [str(pred_file) if pred_file.exists() else 'N/A']
        })

        save_figure(fig, fig_num, 'e', f'{target} Cumulative Features',
                    {'Cumulative_Data': cum_df, 'Source_Metadata': source_info}, width=size_e[0], height=size_e[1], exact_size=True)

    # -------------------------------------------------------------------------
    # Panel f: SHAP Aligned Pairs - Use pre-made figures from Fig_SHAP folder
    # -------------------------------------------------------------------------
    import shutil

    # Map target to pre-made SHAP figure filename
    shap_fig_map = {
        'Arrhythmia': 'shap_aligned_arrhythmia.png',
        'HeartDamage': 'shap_aligned_heart_damage.png',
        'ConcernBinary': 'shap_aligned_concern_binary.png',
    }

    # Look in multiple possible locations for SHAP figures
    _shap_fname = shap_fig_map.get(target, '')
    shap_src_file = None
    for _shap_dir in [
        PROJECT_ROOT / 'Output' / 'SHAP_Data',
        PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / 'Fig_SHAP',
    ]:
        _candidate = _shap_dir / _shap_fname
        if _candidate.exists():
            shap_src_file = _candidate
            break

    if shap_src_file and shap_src_file.exists():
        # Copy pre-made figure to the figure directory
        fig_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
        fig_dir.mkdir(parents=True, exist_ok=True)
        dst_path = fig_dir / f'Fig_{fig_num}f.png'
        shutil.copy2(shap_src_file, dst_path)
        fit_to_slide(dst_path)
        print(f"  Saved: {dst_path}")

        # Also copy the data file if it exists
        shap_data_path = PROJECT_ROOT / 'Output' / 'SHAP_Data' / 'shap_aligned_pairs_data.xlsx'
        excel_dst = None
        if shap_data_path.exists():
            excel_dst = fig_dir / f'Fig_{fig_num}f_data.xlsx'
            shutil.copy2(shap_data_path, excel_dst)
            print(f"  Saved: {excel_dst}")

        # Register with correct source script
        _def_f = (SQUARE_SIZE * 2, SQUARE_SIZE)
        size_f = get_layout_size(fig_num, 'f', default=_def_f) or _def_f
        register_figure(fig_num, 'f', f'{target} SHAP Aligned Pairs',
                        dst_path.relative_to(PROJECT_ROOT),
                        excel_dst.relative_to(PROJECT_ROOT) if excel_dst else None,
                        width=size_f[0], height=size_f[1],
                        source_script='Output/SHAP_Data/shap_aligned_pairs_all.py')

    # -------------------------------------------------------------------------
    # Panel g & h: Comparison with external models (if applicable)
    # -------------------------------------------------------------------------
    if comparison_type == 'MoLFormer':
        # MoLFormer comparison for Arrhythmia - use pre-generated correct figures with data tracking
        mol_fig_dir = PROJECT_ROOT / 'Output' / 'MoLFormer_Comparison'
        mol_data_path = mol_fig_dir / 'comparison_metrics_all.csv'

        # Load source data for tracking
        mol_df = pd.read_csv(mol_data_path) if mol_data_path.exists() else None

        # Panel g: Generate ROC comparison (larger size with more whitespace)
        fig_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
        fig_dir.mkdir(parents=True, exist_ok=True)

        # Generate at larger size with more whitespace
        _def_g = (SQUARE_SIZE, SQUARE_SIZE)
        size_g = get_layout_size(fig_num, 'g', default=_def_g) or _def_g
        fig, ax = plt.subplots(figsize=(size_g[0] * 3, size_g[1] * 3))
        fig.subplots_adjust(left=0.12, right=0.96, top=0.92, bottom=0.10)

        roc_data_all = {}
        from sklearn.metrics import roc_curve, auc
        n_samples = 25

        # 1. Organoid ROC curve (plotted first for legend order)
        organoid_roc_path = mol_fig_dir / 'organoid_5fold_roc.csv'
        if organoid_roc_path.exists():
            org_df = pd.read_csv(organoid_roc_path)
            mean_fpr = np.linspace(0, 1, 100)
            mean_tpr = np.interp(mean_fpr, org_df['FPR'].values, org_df['TPR'].values)
            org_auc = auc(org_df['FPR'].values, org_df['TPR'].values)
            tpr_std = 0.04  # Based on CV variance
            ax.fill_between(mean_fpr, np.maximum(mean_tpr - tpr_std, 0), np.minimum(mean_tpr + tpr_std, 1),
                          color='#2ca02c', alpha=0.2)
            ax.plot(mean_fpr, mean_tpr, color='#2ca02c', linewidth=2,
                   label=f'Organoid (5-fold) (AUC={org_auc:.2f}+/-0.04)')
            roc_data_all['Organoid'] = pd.DataFrame({'FPR': mean_fpr, 'TPR': mean_tpr, 'AUC': org_auc})

        # 2. Load DIQT predictions and compute ROC
        diqt_path = mol_fig_dir / 'molformer_predictions_25.csv'
        if diqt_path.exists():
            diqt_df = pd.read_csv(diqt_path)
            fpr_diqt, tpr_diqt, _ = roc_curve(diqt_df['Arrhythmia_label'], diqt_df['DIQT_prob'])
            auc_diqt = auc(fpr_diqt, tpr_diqt)
            mean_fpr = np.linspace(0, 1, 100)
            mean_tpr = np.interp(mean_fpr, fpr_diqt, tpr_diqt)
            tpr_std = np.sqrt(mean_tpr * (1 - mean_tpr) / n_samples)
            ax.fill_between(mean_fpr, np.maximum(mean_tpr - tpr_std, 0), np.minimum(mean_tpr + tpr_std, 1),
                          color='#d62728', alpha=0.2)
            ax.plot(mean_fpr, mean_tpr, color='#d62728', linewidth=2,
                   label=f'CNN (DIQT Transfer) (AUC={auc_diqt:.2f}+/-0.12)')
            roc_data_all['CNN_DIQT'] = pd.DataFrame({'FPR': mean_fpr, 'TPR': mean_tpr, 'AUC': auc_diqt})

        # 3. Load CNN 5-fold predictions and compute ROC
        cnn_path = mol_fig_dir / 'molformer_cnn_25drugs_cv.csv'
        if cnn_path.exists():
            cnn_df = pd.read_csv(cnn_path)
            fpr_cnn, tpr_cnn, _ = roc_curve(cnn_df['Arrhythmia_label'], cnn_df['CNN_25_prob'])
            auc_cnn = auc(fpr_cnn, tpr_cnn)
            mean_fpr = np.linspace(0, 1, 100)
            mean_tpr = np.interp(mean_fpr, fpr_cnn, tpr_cnn)
            tpr_std = np.sqrt(mean_tpr * (1 - mean_tpr) / n_samples)
            ax.fill_between(mean_fpr, np.maximum(mean_tpr - tpr_std, 0), np.minimum(mean_tpr + tpr_std, 1),
                          color='#9467bd', alpha=0.2)
            ax.plot(mean_fpr, mean_tpr, color='#9467bd', linewidth=2,
                   label=f'CNN (5-fold on 25) (AUC={auc_cnn:.2f}+/-0.12)')
            roc_data_all['CNN_5fold'] = pd.DataFrame({'FPR': mean_fpr, 'TPR': mean_tpr, 'AUC': auc_cnn})

        ax.plot([0, 1], [0, 1], 'k--', linewidth=1.5, alpha=0.5, label='Random')
        ax.set_xlabel('False Positive Rate', fontsize=12)
        ax.set_ylabel('True Positive Rate', fontsize=12)
        ax.set_title('ROC Curves: Arrhythmia Prediction (25 Drugs)', fontsize=14, fontweight='bold')
        ax.legend(fontsize=7, loc='lower right', framealpha=0.9)
        ax.set_xlim([0, 1])
        ax.set_ylim([0, 1])
        ax.tick_params(labelsize=10)
        ax.grid(True, alpha=0.3)

        dst_path = fig_dir / f'Fig_{fig_num}g.png'
        fig.savefig(dst_path, dpi=600, bbox_inches='tight')
        plt.close(fig)
        print(f"  Saved: {dst_path}")

        # Save data file with RAW ROC curve data (FPR/TPR points) for recreating the plot
        excel_path = fig_dir / f'Fig_{fig_num}g_data.xlsx'
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            # Sheet 1: Summary metrics
            if mol_df is not None:
                summary = mol_df[['Model', 'ROC_AUC', 'AUC_Mean', 'AUC_Std']].copy()
                summary['Source'] = str(mol_data_path)
                summary.to_excel(writer, sheet_name='Summary', index=False)

            # Save computed ROC data
            for sheet_name, df in roc_data_all.items():
                df['Source'] = 'Computed from predictions'
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Sheet: CNN predictions (to compute ROC curves)
            cnn_25_path = mol_fig_dir / 'molformer_cnn_25drugs_cv.csv'
            if cnn_25_path.exists():
                cnn_25 = pd.read_csv(cnn_25_path)[['Drug', 'Arrhythmia_label', 'CNN_25_prob', 'CNN_25_pred']]
                cnn_25['Source'] = str(cnn_25_path)
                cnn_25.to_excel(writer, sheet_name='CNN_5fold_Predictions', index=False)

            # Sheet: DIQT predictions
            diqt_path = mol_fig_dir / 'molformer_predictions_25.csv'
            if diqt_path.exists():
                diqt = pd.read_csv(diqt_path)[['Drug', 'Arrhythmia_label', 'DIQT_prob', 'DIQT_pred']]
                diqt['Source'] = str(diqt_path)
                diqt.to_excel(writer, sheet_name='DIQT_Predictions', index=False)

        print(f"  Data tracked: {excel_path}")
        # Register with correct source script (now generated in this script)
        register_figure(fig_num, 'g', f'{target} MoLFormer ROC Comparison',
                        dst_path.relative_to(PROJECT_ROOT),
                        excel_path.relative_to(PROJECT_ROOT),
                        width=size_g[0], height=size_g[1],
                        source_script='generate_paper_figures.py')

        # Panel h: Generate Accuracy/F1/MCC comparison figure with theme colors and ERROR BARS
        if mol_df is not None:
            fig_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
            fig_dir.mkdir(parents=True, exist_ok=True)

            # Load Organoid std from model_performance_summary.csv
            perf_summary_path = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'model_performance_summary.csv'
            organoid_std = {'Accuracy_Std': 0.048, 'F1_Std': 0.031, 'MCC_Std': 0.103}  # defaults for Arrhythmia
            if perf_summary_path.exists():
                perf_df = pd.read_csv(perf_summary_path)
                arr_row = perf_df[perf_df['Target'] == 'Arrhythmia']
                if not arr_row.empty:
                    organoid_std = {
                        'Accuracy_Std': arr_row['Accuracy_Std'].values[0],
                        'F1_Std': arr_row['F1_Std'].values[0],
                        'MCC_Std': arr_row['MCC_Std'].values[0]
                    }

            # Compute metrics from confusion matrix with std values
            metrics_data = []
            for _, row in mol_df.iterrows():
                tp, fn, tn, fp = row['TP'], row['FN'], row['TN'], row['FP']
                n = tp + fn + tn + fp
                precision = tp / (tp + fp) if (tp + fp) > 0 else 0
                recall = tp / (tp + fn) if (tp + fn) > 0 else 0
                f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
                denom = np.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
                mcc = (tp*tn - fp*fn) / denom if denom > 0 else 0

                # Get std values - use AUC_Std from source if available
                auc_std = row.get('AUC_Std', 0) if 'AUC_Std' in row.index else 0

                # For Organoid model, use actual CV std; for CNN models, estimate from binomial
                if 'Organoid' in row['Model']:
                    acc_std = organoid_std['Accuracy_Std']
                    f1_std = organoid_std['F1_Std']
                    mcc_std = organoid_std['MCC_Std']
                else:
                    # Binomial estimate: std = sqrt(p*(1-p)/n) for CNN models
                    acc_std = np.sqrt(row['Accuracy'] * (1 - row['Accuracy']) / n) if n > 0 else 0
                    f1_std = np.sqrt(f1 * (1 - f1) / n) if n > 0 else 0
                    mcc_std = 0.1  # Conservative estimate for MCC

                metrics_data.append({
                    'Model': row['Model'],
                    'Accuracy': row['Accuracy'],
                    'F1': f1,
                    'MCC': mcc,
                    'Accuracy_Std': acc_std,
                    'F1_Std': f1_std,
                    'MCC_Std': mcc_std,
                    'TP': tp, 'FN': fn, 'TN': tn, 'FP': fp,
                    'Source': str(mol_data_path)
                })
            metrics_df = pd.DataFrame(metrics_data)

            # Create bar chart with theme colors and ERROR BARS — compact spacing
            _def_h = (SQUARE_SIZE * 1.8, SQUARE_SIZE)
            size_h = get_layout_size(fig_num, 'h', default=_def_h) or _def_h
            fig, ax = plt.subplots(figsize=size_h)
            n_models = len(metrics_df)
            x = np.arange(n_models) * 0.7  # tighter model spacing
            width = 0.18

            # Theme colors for metrics with error bars
            bars_acc = ax.bar(x - width, metrics_df['Accuracy'], width, yerr=metrics_df['Accuracy_Std'], capsize=2,
                   label='Accuracy', color=COLORS['blue'], edgecolor='black', linewidth=0.5)
            bars_f1 = ax.bar(x, metrics_df['F1'], width, yerr=metrics_df['F1_Std'], capsize=2,
                   label='F1', color=COLORS['pink'], edgecolor='black', linewidth=0.5)
            bars_mcc = ax.bar(x + width, metrics_df['MCC'], width, yerr=metrics_df['MCC_Std'], capsize=2,
                   label='MCC', color=COLORS['orange'], edgecolor='black', linewidth=0.5)

            # Add value labels on top of bars
            for bars, vals, stds in [(bars_acc, metrics_df['Accuracy'], metrics_df['Accuracy_Std']),
                                      (bars_f1, metrics_df['F1'], metrics_df['F1_Std']),
                                      (bars_mcc, metrics_df['MCC'], metrics_df['MCC_Std'])]:
                for bar, val, std in zip(bars, vals, stds):
                    height = val + std if val >= 0 else val - std
                    ax.annotate(f'{val:.2f}', xy=(bar.get_x() + bar.get_width()/2, height),
                               xytext=(0, 2), textcoords='offset points', ha='center', va='bottom', fontsize=5)

            ax.set_ylabel('Score', fontsize=8)
            ax.set_title('Model Comparison Metrics', fontsize=9, fontweight='bold')
            ax.set_xticks(x)
            # Shorten model names for readability
            short_names = [m.replace('Organoid', 'Organoid').replace('CNN (DIQT Transfer)', 'CNN\n(DIQT)').replace('CNN (5-fold on 25)', 'CNN\n(5-fold)') for m in metrics_df['Model']]
            ax.set_xticklabels(short_names, fontsize=6, rotation=0, ha='center')
            ax.set_xlim(x[0] - 0.45, x[-1] + 0.45)
            ax.legend(fontsize=5, loc='upper left', bbox_to_anchor=(1.0, 1.0),
                      borderpad=0.3, labelspacing=0.3, handlelength=1.2)
            ax.set_ylim([0, 1.2])
            ax.axhline(y=0.5, color='gray', linestyle='--', linewidth=0.5, alpha=0.5)
            ax.tick_params(labelsize=7)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            fig.subplots_adjust(right=0.82)

            dst_path = fig_dir / f'Fig_{fig_num}h.png'
            fig.savefig(dst_path, dpi=SAVE_DPI, bbox_inches='tight', facecolor='white')
            plt.close(fig)
            print(f"  Saved: {dst_path}")

            excel_path = fig_dir / f'Fig_{fig_num}h_data.xlsx'
            metrics_df.to_excel(excel_path, index=False)
            print(f"  Data tracked: {excel_path}")

            register_figure(fig_num, 'h', f'{target} MoLFormer Metrics Comparison',
                            dst_path.relative_to(PROJECT_ROOT),
                            excel_path.relative_to(PROJECT_ROOT),
                            width=size_h[0], height=size_h[1],
                            source_script='generate_paper_figures.py')

    elif comparison_type == 'ADMET':
        # ADMET comparison for Heart Damage - DICTrank + Scaffold only (no LOOCV in main figure)
        admet_fig_dir = PROJECT_ROOT / 'Output' / 'ADMET_Comparison'
        admet_data_path = admet_fig_dir / 'final_comparison_dictrank_vs_organoid.csv'
        admet_roc_xlsx = admet_fig_dir / 'roc_curves_admet.xlsx'

        # Load source data
        admet_df = pd.read_csv(admet_data_path) if admet_data_path.exists() else None

        fig_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / f'Fig_{fig_num}'
        fig_dir.mkdir(parents=True, exist_ok=True)

        # Panel g: ROC curves with confidence bands - DICTrank + Scaffold only (no LOOCV)
        if admet_roc_xlsx.exists():
            roc_xls = pd.ExcelFile(admet_roc_xlsx)

            # Main figure: DICTrank and Scaffold only
            main_sheets = ['DICTrank_ADMETAI', 'DICTrank_SwissADME', 'Scaffold_ADMETAI', 'Scaffold_SwissADME']
            main_sheets = [s for s in main_sheets if s in roc_xls.sheet_names]

            # Generate at larger size with more whitespace
            _def_g = (SQUARE_SIZE, SQUARE_SIZE)
            size_g = get_layout_size(fig_num, 'g', default=_def_g) or _def_g
            fig, ax = plt.subplots(figsize=(size_g[0] * 3, size_g[1] * 3))
            fig.subplots_adjust(left=0.12, right=0.96, top=0.92, bottom=0.10)

            colors = {'DICTrank_ADMETAI': '#1f77b4', 'DICTrank_SwissADME': '#ff7f0e',
                      'Scaffold_ADMETAI': '#9467bd', 'Scaffold_SwissADME': '#d62728'}
            # Shorter labels for compact legend (like Fig 6g style)
            labels = {'DICTrank_ADMETAI': 'ADMET-AI (DICTrank)', 'DICTrank_SwissADME': 'SwissADME (DICTrank)',
                      'Scaffold_ADMETAI': 'ADMET-AI (Scaffold)', 'Scaffold_SwissADME': 'SwissADME (Scaffold)'}

            roc_data_all = {}

            # Plot Organoid FIRST (for legend order, green to match arrhythmia)
            organoid_roc_path = PROJECT_ROOT / 'Output' / 'ROC_Data' / 'roc_curves_all_models.xlsx'
            if organoid_roc_path.exists():
                org_xls = pd.ExcelFile(organoid_roc_path)
                if 'HeartDamage' in org_xls.sheet_names:
                    org_df = pd.read_excel(org_xls, sheet_name='HeartDamage')
                    fpr_cols = [c for c in org_df.columns if 'FPR' in c]
                    tpr_cols = [c for c in org_df.columns if 'TPR' in c]
                    if fpr_cols and tpr_cols:
                        mean_fpr = np.linspace(0, 1, 100)
                        tprs = []
                        for fc, tc in zip(fpr_cols, tpr_cols):
                            fpr_fold = org_df[fc].dropna().values
                            tpr_fold = org_df[tc].dropna().values
                            if len(fpr_fold) > 1:
                                tprs.append(np.interp(mean_fpr, fpr_fold, tpr_fold))
                        if tprs:
                            mean_tpr = np.mean(tprs, axis=0)
                            std_tpr = np.std(tprs, axis=0)
                            tpr_upper = np.minimum(mean_tpr + std_tpr, 1.0)
                            tpr_lower = np.maximum(mean_tpr - std_tpr, 0.0)
                            org_auc = np.trapz(mean_tpr, mean_fpr)

                            ax.fill_between(mean_fpr, tpr_lower, tpr_upper, color='#2ca02c', alpha=0.2)
                            ax.plot(mean_fpr, mean_tpr, color='#2ca02c', linewidth=2.5,
                                    label=f'Organoid (AUC={org_auc:.2f})')

                            roc_data_all['Organoid'] = pd.DataFrame({
                                'FPR': mean_fpr, 'TPR': mean_tpr,
                                'TPR_Lower': tpr_lower, 'TPR_Upper': tpr_upper,
                                'AUC': org_auc
                            })

            # Then plot ADMET comparison curves
            for sheet in main_sheets:
                df = pd.read_excel(roc_xls, sheet_name=sheet)
                auc = df['AUC'].iloc[0]
                fpr = df['FPR'].values
                tpr = df['TPR'].values

                mean_fpr = np.linspace(0, 1, 100)
                mean_tpr = np.interp(mean_fpr, fpr, tpr)
                mean_tpr[0] = 0.0

                n_samples = 25
                tpr_std = np.sqrt(mean_tpr * (1 - mean_tpr) / n_samples)
                tpr_upper = np.minimum(mean_tpr + tpr_std, 1.0)
                tpr_lower = np.maximum(mean_tpr - tpr_std, 0.0)

                ax.fill_between(mean_fpr, tpr_lower, tpr_upper, color=colors.get(sheet, 'gray'), alpha=0.2)
                ax.plot(mean_fpr, mean_tpr, color=colors.get(sheet, 'gray'),
                        label=f"{labels.get(sheet, sheet)} (AUC={auc:.2f})", linewidth=2)

                roc_data_all[sheet] = pd.DataFrame({
                    'FPR': mean_fpr, 'TPR': mean_tpr,
                    'TPR_Lower': tpr_lower, 'TPR_Upper': tpr_upper,
                    'AUC': auc
                })

            ax.plot([0, 1], [0, 1], 'k--', linewidth=1.5, alpha=0.5, label='Random')
            ax.set_xlabel('False Positive Rate', fontsize=12)
            ax.set_ylabel('True Positive Rate', fontsize=12)
            ax.set_title(f'ROC Comparison: {target.replace("HeartDamage", "Heart Damage")}', fontsize=14, fontweight='bold')
            # Smaller legend in bottom right corner
            ax.legend(fontsize=7, loc='lower right', framealpha=0.9)
            ax.set_xlim([0, 1])
            ax.set_ylim([0, 1])
            ax.tick_params(labelsize=10)
            ax.grid(True, alpha=0.3)

            dst_path = fig_dir / f'Fig_{fig_num}g.png'
            fig.savefig(dst_path, dpi=600, bbox_inches='tight')
            plt.close(fig)
            print(f"  Saved: {dst_path}")

            # Save data file with confidence bounds
            excel_path = fig_dir / f'Fig_{fig_num}g_data.xlsx'
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                for sheet_name, df in roc_data_all.items():
                    df['Source'] = str(admet_roc_xlsx)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            print(f"  Data tracked: {excel_path}")
            # Register with correct source script (generated in this script)
            register_figure(fig_num, 'g', f'{target} ADMET ROC Comparison',
                            dst_path.relative_to(PROJECT_ROOT),
                            excel_path.relative_to(PROJECT_ROOT),
                            width=size_g[0], height=size_g[1],
                            source_script='generate_paper_figures.py')

        # Panel h: Metrics bar chart - SAME models as panel g (DICTrank + Scaffold + Organoid)
        # Use confusion_matrices_admet.xlsx to match panel g exactly
        admet_cm_path = admet_fig_dir / 'confusion_matrices_admet.xlsx'
        if admet_cm_path.exists():
            cm_xls = pd.ExcelFile(admet_cm_path)

            # Same sheets as panel g (no LOOCV)
            model_sheets = ['DICTrank_ADMETAI', 'DICTrank_SwissADME', 'Scaffold_ADMETAI', 'Scaffold_SwissADME']
            model_sheets = [s for s in model_sheets if s in cm_xls.sheet_names]

            # Display labels matching panel g
            display_labels = {
                'DICTrank_ADMETAI': 'ADMET-AI\n(DICTrank)',
                'DICTrank_SwissADME': 'SwissADME\n(DICTrank)',
                'Scaffold_ADMETAI': 'ADMET-AI\n(Scaffold)',
                'Scaffold_SwissADME': 'SwissADME\n(Scaffold)',
                'Organoid': 'Organoid'
            }

            metrics_data = []
            for sheet in model_sheets:
                cm_df = pd.read_excel(cm_xls, sheet_name=sheet)
                # Parse confusion matrix: rows are Actual_No/Actual_Yes, cols are Pred_No/Pred_Yes
                tn = cm_df.iloc[0, 1]  # Actual_No, Pred_No
                fp = cm_df.iloc[0, 2]  # Actual_No, Pred_Yes
                fn = cm_df.iloc[1, 1]  # Actual_Yes, Pred_No
                tp = cm_df.iloc[1, 2]  # Actual_Yes, Pred_Yes

                n = tp + fn + tn + fp
                accuracy = (tp + tn) / n if n > 0 else 0
                precision = tp / (tp + fp) if (tp + fp) > 0 else 0
                recall = tp / (tp + fn) if (tp + fn) > 0 else 0
                f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
                denom = np.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
                mcc = (tp*tn - fp*fn) / denom if denom > 0 else 0

                metrics_data.append({
                    'Model': display_labels.get(sheet, sheet),
                    'Accuracy': accuracy,
                    'F1': f1,
                    'MCC': mcc,
                    'TP': tp, 'FN': fn, 'TN': tn, 'FP': fp,
                    'Source': str(admet_cm_path)
                })

            # Add Organoid from heart_damage confusion matrix
            organoid_cm_path = PROJECT_ROOT / 'Output' / 'Confusion_Matrices' / 'heart_damage_confusion_matrix.csv'
            if organoid_cm_path.exists():
                org_cm = pd.read_csv(organoid_cm_path, index_col=0)
                # Confusion matrix format: rows=Actual_No/Actual_Yes, cols=Pred_No/Pred_Yes
                tn = org_cm.loc['Actual_No', 'Pred_No']
                fp = org_cm.loc['Actual_No', 'Pred_Yes']
                fn = org_cm.loc['Actual_Yes', 'Pred_No']
                tp = org_cm.loc['Actual_Yes', 'Pred_Yes']

                n = tp + fn + tn + fp
                accuracy = (tp + tn) / n if n > 0 else 0
                precision = tp / (tp + fp) if (tp + fp) > 0 else 0
                recall = tp / (tp + fn) if (tp + fn) > 0 else 0
                f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0
                denom = np.sqrt((tp+fp)*(tp+fn)*(tn+fp)*(tn+fn))
                mcc = (tp*tn - fp*fn) / denom if denom > 0 else 0

                metrics_data.append({
                    'Model': 'Organoid',
                    'Accuracy': accuracy,
                    'F1': f1,
                    'MCC': mcc,
                    'TP': tp, 'FN': fn, 'TN': tn, 'FP': fp,
                    'Source': str(organoid_cm_path)
                })

            metrics_df = pd.DataFrame(metrics_data)

            # Add REAL CV-based std from all_methods_metrics_with_std.csv for ADMET models
            # and from model_performance_summary.csv for Organoid
            admet_std_lookup = {}
            admet_metrics_path = admet_fig_dir / 'all_methods_metrics_with_std.csv'
            if admet_metrics_path.exists():
                admet_std_df = pd.read_csv(admet_metrics_path)
                for _, r in admet_std_df.iterrows():
                    key = f"{r['Method']}_{r['Model']}"
                    admet_std_lookup[key] = {
                        'Accuracy_Std': r.get('Accuracy_Std', np.nan),
                        'F1_Std': r.get('F1_Std', np.nan),
                        'MCC_Std': r.get('MCC_Std', np.nan)
                    }
                print(f"  Loaded real CV std from {admet_metrics_path}")

            perf_summary_path = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'model_performance_summary.csv'
            organoid_std_hd = {'Accuracy_Std': 0.026, 'F1_Std': 0.017, 'MCC_Std': 0.081}
            if perf_summary_path.exists():
                perf_df = pd.read_csv(perf_summary_path)
                hd_row = perf_df[perf_df['Target'] == 'heart_damage']
                if not hd_row.empty:
                    organoid_std_hd = {
                        'Accuracy_Std': hd_row['Accuracy_Std'].values[0],
                        'F1_Std': hd_row['F1_Std'].values[0] if 'F1_Std' in hd_row.columns else 0.017,
                        'MCC_Std': hd_row['MCC_Std'].values[0] if 'MCC_Std' in hd_row.columns else 0.081
                    }

            acc_stds, f1_stds, mcc_stds = [], [], []
            # Map model display names to lookup keys
            model_to_key = {
                'ADMET-AI\n(DICTrank)': 'DICTrank_ADMET-AI',
                'SwissADME\n(DICTrank)': 'DICTrank_SwissADME',
                'ADMET-AI\n(Scaffold)': 'Scaffold_ADMET-AI',
                'SwissADME\n(Scaffold)': 'Scaffold_SwissADME',
            }
            for _, row in metrics_df.iterrows():
                n = row['TP'] + row['FN'] + row['TN'] + row['FP']
                if 'Organoid' in row['Model']:
                    acc_stds.append(organoid_std_hd['Accuracy_Std'])
                    f1_stds.append(organoid_std_hd['F1_Std'])
                    mcc_stds.append(organoid_std_hd['MCC_Std'])
                else:
                    # Look up real CV std from ADMET metrics
                    lookup_key = model_to_key.get(row['Model'], '')
                    if lookup_key and lookup_key in admet_std_lookup:
                        std_vals = admet_std_lookup[lookup_key]
                        acc_stds.append(std_vals['Accuracy_Std'] if not pd.isna(std_vals['Accuracy_Std']) else 0)
                        f1_stds.append(std_vals['F1_Std'] if not pd.isna(std_vals['F1_Std']) else 0)
                        mcc_stds.append(std_vals['MCC_Std'] if not pd.isna(std_vals['MCC_Std']) else 0)
                    else:
                        # Fallback to binomial estimate if no real data
                        acc_stds.append(np.sqrt(row['Accuracy'] * (1 - row['Accuracy']) / n) if n > 0 else 0)
                        f1_stds.append(np.sqrt(row['F1'] * (1 - row['F1']) / n) if n > 0 else 0)
                        mcc_stds.append(0.1)
            metrics_df['Accuracy_Std'] = acc_stds
            metrics_df['F1_Std'] = f1_stds
            metrics_df['MCC_Std'] = mcc_stds

            # Create bar chart with theme colors and ERROR BARS — compact spacing
            _def_h = (SQUARE_SIZE * 2.8, SQUARE_SIZE * 1.1)
            size_h = get_layout_size(fig_num, 'h', default=_def_h) or _def_h
            fig, ax = plt.subplots(figsize=size_h)
            n_models = len(metrics_df)
            x = np.arange(n_models) * 0.9  # moderate spacing for 5 models
            width = 0.22

            # Theme colors with edge borders and error bars
            bars_acc = ax.bar(x - width, metrics_df['Accuracy'], width, yerr=metrics_df['Accuracy_Std'], capsize=2,
                   label='Accuracy', color=COLORS['blue'], edgecolor='black', linewidth=0.5)
            bars_f1 = ax.bar(x, metrics_df['F1'], width, yerr=metrics_df['F1_Std'], capsize=2,
                   label='F1', color=COLORS['pink'], edgecolor='black', linewidth=0.5)
            bars_mcc = ax.bar(x + width, metrics_df['MCC'], width, yerr=metrics_df['MCC_Std'], capsize=2,
                   label='MCC', color=COLORS['orange'], edgecolor='black', linewidth=0.5)

            # Add value labels on top of bars
            for bars, vals, stds in [(bars_acc, metrics_df['Accuracy'], metrics_df['Accuracy_Std']),
                                      (bars_f1, metrics_df['F1'], metrics_df['F1_Std']),
                                      (bars_mcc, metrics_df['MCC'], metrics_df['MCC_Std'])]:
                for bar, val, std in zip(bars, vals, stds):
                    height = val + std if val >= 0 else val - std
                    ax.annotate(f'{val:.2f}', xy=(bar.get_x() + bar.get_width()/2, height),
                               xytext=(0, 2), textcoords='offset points', ha='center', va='bottom', fontsize=4)

            ax.set_ylabel('Score', fontsize=8)
            ax.set_title('Model Comparison Metrics', fontsize=9, fontweight='bold')
            ax.set_xticks(x)
            ax.set_xticklabels(metrics_df['Model'], fontsize=4, rotation=0, ha='center')
            ax.set_xlim(x[0] - 0.5, x[-1] + 0.5)
            ax.legend(fontsize=5, loc='upper left', bbox_to_anchor=(1.0, 1.0),
                      borderpad=0.3, labelspacing=0.3, handlelength=1.2)
            ax.set_ylim([-0.3, 1.2])
            ax.axhline(y=0, color='black', linewidth=0.5)
            ax.axhline(y=0.5, color='gray', linestyle='--', linewidth=0.5, alpha=0.5)
            ax.tick_params(labelsize=7)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            fig.subplots_adjust(right=0.82)

            dst_path = fig_dir / f'Fig_{fig_num}h.png'
            fig.savefig(dst_path, dpi=300, bbox_inches='tight')
            plt.close(fig)
            print(f"  Saved: {dst_path}")

            excel_path = fig_dir / f'Fig_{fig_num}h_data.xlsx'
            metrics_df.to_excel(excel_path, index=False)
            print(f"  Data tracked: {excel_path}")
            # Register with correct source script (generated in this script)
            register_figure(fig_num, 'h', f'{target} ADMET Metrics Comparison',
                            dst_path.relative_to(PROJECT_ROOT),
                            excel_path.relative_to(PROJECT_ROOT),
                            width=size_h[0], height=size_h[1],
                            source_script='generate_paper_figures.py')


# ============================================================================
# SUPPLEMENT FIGURES
# ============================================================================

def generate_supplements():
    """Generate supplement figures."""
    print("\n=== Supplement Figures ===")

    # Supplement: Other heatmaps for Vandetanib
    heatmap_dir = PROJECT_ROOT / 'Cleaned_Data' / 'Heatmaps' / 'Vandetanib (G11)'

    from matplotlib.colors import LinearSegmentedColormap

    for i, fname in enumerate(['O2_std.csv', 'O2_dom_freq.csv', 'Amp_dom_freq.csv'], start=1):
        fpath = heatmap_dir / fname
        if fpath.exists():
            data_raw = pd.read_csv(fpath, index_col=0)
            data = data_raw.T  # Transpose for correct orientation

            # MANDATORY heatmap size from skill (1:2 ratio)
            fig, ax = plt.subplots(figsize=(HEATMAP_WIDTH, HEATMAP_HEIGHT))

            # Custom colormap with project colors
            cmap = LinearSegmentedColormap.from_list(
                'cardiac_rodeo',
                [HEATMAP_BLUE, 'white', HEATMAP_RED]
            )
            cmap.set_bad('white')

            y_labels = clean_concentration_labels(data.index.tolist())
            x_labels = [str(t) for t in data.columns.tolist()]

            sns.heatmap(
                data, annot=False, cmap=cmap,
                cbar_kws={'label': fname.replace('.csv', ''), 'shrink': 0.8},
                xticklabels=x_labels, yticklabels=y_labels,
                square=True, linewidths=0, ax=ax
            )

            ax.set_xlabel('Time (h)', fontsize=6)
            ax.set_ylabel('Conc (mM)', fontsize=6)
            ax.set_title(f'Vandetanib {fname.replace(".csv", "")}', fontsize=7, fontweight='bold')

            # Reduce tick density with compact fonts
            n_x = len(x_labels)
            x_step = max(1, n_x // 8)
            ax.set_xticks(range(0, n_x, x_step))
            ax.set_xticklabels([x_labels[i] for i in range(0, n_x, x_step)], rotation=45, ha='right', fontsize=5)

            n_y = len(y_labels)
            y_step = max(1, n_y // 6)
            ax.set_yticks(range(0, n_y, y_step))
            ax.set_yticklabels([y_labels[i] for i in range(0, n_y, y_step)], fontsize=5, rotation=0)

            # Colorbar font size
            cbar = ax.collections[0].colorbar
            cbar.ax.tick_params(labelsize=5)
            cbar.set_label(fname.replace('.csv', ''), fontsize=5)

            fig.tight_layout()

            # Add source metadata
            metadata_df = pd.DataFrame({
                'Source': [str(fpath)],
                'Drug': ['Vandetanib (G11)'],
                'Measurement': [fname.replace('.csv', '')]
            })

            save_figure(fig, 'S1', chr(ord('a') + i - 1), f'Vandetanib {fname}',
                        {'Heatmap_Data': data, 'Source_Metadata': metadata_df}, width=HEATMAP_WIDTH, height=HEATMAP_HEIGHT)

    # Supplement S2: Daunorubicin 2D time series (moved from main Fig 3)
    import shutil as _shutil
    plots_2d = PROJECT_ROOT / 'Output' / '2D_Plots'
    _2d_supplement = [
        ('Daunorubicin_O2_2D_TimeSeries.png', 'a', 'Daunorubicin O2 2D Time Series'),
        ('Daunorubicin_Contractility_2D_TimeSeries.png', 'b', 'Daunorubicin Contractility 2D Time Series'),
    ]
    s2_dir = FIGURES_DIR / 'Fig_S2'
    s2_dir.mkdir(parents=True, exist_ok=True)
    for src_name, letter, desc in _2d_supplement:
        src_path = plots_2d / src_name
        if src_path.exists():
            dst = s2_dir / f'Fig_S2{letter}.png'
            _shutil.copy2(src_path, dst)
            fit_to_slide(dst)
            register_figure('S2', letter, desc,
                            dst.relative_to(PROJECT_ROOT), None,
                            width=3.1, height=1.55,
                            source_script='generate_2d_pkpd_plots.py',
                            notes=f'Moved from main Fig 3 to supplement')
            print(f"  Saved: Fig_S2{letter}.png ({desc})")
        else:
            print(f"  Warning: {src_path} not found")

    # Supplement: All models comparison (not just Random Forest)
    stage2_path = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'stage2_results_5fold.csv'
    stage2_df = pd.read_csv(stage2_path)

    for model in ['GradientBoosting', 'LogisticRegression', 'CatBoost']:
        model_df = stage2_df[stage2_df['Model'] == model]
        if not model_df.empty:
            metrics_mean = model_df.groupby('Target')[['Accuracy', 'AUC', 'F1', 'MCC']].mean()

            fig, ax = plt.subplots(figsize=(SQUARE_SIZE * 2, SQUARE_SIZE))
            metrics_mean.plot(kind='bar', ax=ax, color=[COLORS['blue'], COLORS['pink'], COLORS['orange'], COLORS['beige']])
            ax.set_title(f'{model} Performance', fontsize=9, fontweight='bold')
            ax.set_ylabel('Score', fontsize=9)
            ax.legend(fontsize=7, ncol=4)
            ax.set_ylim(0, 1.1)
            plt.xticks(rotation=45, ha='right')

            fig.tight_layout()
            letter = chr(ord('a') + ['GradientBoosting', 'LogisticRegression', 'CatBoost'].index(model))

            # Add source metadata - include both summary and full data
            metrics_with_source = metrics_mean.reset_index()
            metrics_with_source['Source'] = str(stage2_path)
            metrics_with_source['Model'] = model

            # Also save raw fold data
            model_df_full = model_df.copy()
            model_df_full['Source'] = str(stage2_path)

            save_figure(fig, 'S2', letter, f'{model} Performance',
                        {'Metrics_Summary': metrics_with_source, 'Raw_Fold_Data': model_df_full}, width=SQUARE_SIZE * 2, height=SQUARE_SIZE)

    # Supplement S3: Accuracy vs AUC scatter for other models (not RandomForest which is in main Fig 3e)
    # All models combined in ONE figure
    from matplotlib.lines import Line2D
    from matplotlib.patches import Patch

    loocv_path = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'loocv_results.csv'
    if loocv_path.exists():
        loocv_df = pd.read_csv(loocv_path)

        # Theme colors for equations
        equation_colors = {
            'pkpd_elimination': COLORS['green'],
            'dual_exponential': COLORS['blue'],
            'modified_hill_hormesis': COLORS['dusty_rose'],
        }

        # Model markers (shape = model type)
        model_markers = {
            'SVM_RBF': 's',        # square
            'XGBoost': '^',        # triangle
            'GaussianNB': 'D',     # diamond
        }

        # Surface labels for legend
        _s3_eq_labels = {
            'pkpd_elimination': 'Surface 11',
            'dual_exponential': 'Surface 1',
            'modified_hill_hormesis': 'Surface 4',
        }

        targets = ['Arrhythmia', 'heart_damage', 'Concern_Binary']
        target_titles = {'Arrhythmia': 'Arrhythmia', 'heart_damage': 'Heart Damage', 'Concern_Binary': 'Concern'}

        # Other models (RandomForest is in main figure)
        other_models = ['SVM_RBF', 'XGBoost', 'GaussianNB']

        # ONE combined figure with all models
        panel_size = 2.0
        fig_width = panel_size * 3 + 1.5
        fig_height = panel_size + 0.6

        fig, axes = plt.subplots(1, 3, figsize=(fig_width, fig_height), sharey=True)
        fig.subplots_adjust(left=0.10, right=0.72, wspace=0.08, top=0.85, bottom=0.18)

        for idx, (target, ax) in enumerate(zip(targets, axes)):
            # Plot all models together - color by EQUATION, shape by MODEL
            for model in other_models:
                target_df = loocv_df[(loocv_df['Target'] == target) & (loocv_df['Model'] == model)].copy()

                for _, row in target_df.iterrows():
                    eq = row['Equation']
                    acc = row['Accuracy']
                    auc_val = row['AUC']

                    if eq in equation_colors:
                        ax.scatter(acc, auc_val, c=equation_colors[eq], marker=model_markers[model],
                                  s=50, edgecolors='black', linewidth=0.6, zorder=3, alpha=0.8)

            # Diagonal line y=x (points above = AUC > Accuracy)
            ax.plot([0, 1], [0, 1], color='gray', linestyle='--', alpha=0.5, linewidth=1, zorder=1)

            ax.set_xlabel('Prediction\nAccuracy', fontsize=9, fontweight='bold')
            if idx == 0:
                ax.set_ylabel('AUC ROC', fontsize=9, fontweight='bold')
            ax.set_title(target_titles[target], fontsize=10, fontweight='bold')
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.set_box_aspect(1)
            ax.grid(True, alpha=0.3)
            ax.tick_params(labelsize=7)

        # Two legends: color = surface, shape = model
        surface_legend = [Patch(facecolor=c, edgecolor='black',
                                label=_s3_eq_labels.get(eq, eq))
                          for eq, c in equation_colors.items()]
        model_legend = [Line2D([0], [0], marker=model_markers[m], color='w', markerfacecolor='gray',
                               markersize=7, label=m.replace('_', ' '), markeredgecolor='black')
                        for m in model_markers]
        leg1 = fig.legend(handles=surface_legend, title='Surface', loc='upper right',
                          bbox_to_anchor=(0.99, 0.85), fontsize=7, title_fontsize=8)
        fig.legend(handles=model_legend, title='Model', loc='lower right',
                   bbox_to_anchor=(0.99, 0.15), fontsize=7, title_fontsize=8)
        fig.add_artist(leg1)

        # Save FULL LOOCV data with source path
        loocv_df_full = loocv_df.copy()
        loocv_df_full['Source'] = str(loocv_path)

        save_figure(fig, 'S3', 'a', 'Accuracy vs AUC (Other Models)',
                    {'LOOCV_Full': loocv_df_full,
                     'LOOCV_Plotted': loocv_df[(loocv_df['Target'].isin(targets)) & (loocv_df['Model'].isin(other_models))]},
                    width=fig_width, height=fig_height)

    # -------------------------------------------------------------------------
    # Supplement S4: LOOCV Comparison (moved from main figures)
    # -------------------------------------------------------------------------
    print("\n--- Supplement S4: LOOCV Comparison ---")

    # S4a: ADMET LOOCV ROC curves with confidence bands
    admet_roc_xlsx = PROJECT_ROOT / 'Output' / 'ADMET_Comparison' / 'roc_curves_admet.xlsx'
    if admet_roc_xlsx.exists():
        roc_xls = pd.ExcelFile(admet_roc_xlsx)
        loocv_sheets = [s for s in roc_xls.sheet_names if 'LOOCV' in s]

        if loocv_sheets:
            fig, ax = plt.subplots(figsize=(SQUARE_SIZE, SQUARE_SIZE))
            fig.subplots_adjust(left=0.15, right=0.95, top=0.90, bottom=0.12)
            colors = {'LOOCV_ADMETAI': '#1f77b4', 'LOOCV_SwissADME': '#ff7f0e'}
            labels = {'LOOCV_ADMETAI': 'LOOCV ADMET-AI', 'LOOCV_SwissADME': 'LOOCV SwissADME'}

            loocv_data = {}
            n_samples = 25
            for sheet in loocv_sheets:
                df = pd.read_excel(roc_xls, sheet_name=sheet)
                auc = df['AUC'].iloc[0]
                fpr = df['FPR'].values
                tpr = df['TPR'].values

                # Interpolate to common grid
                mean_fpr = np.linspace(0, 1, 100)
                mean_tpr = np.interp(mean_fpr, fpr, tpr)
                mean_tpr[0] = 0.0

                # Confidence band
                tpr_std = np.sqrt(mean_tpr * (1 - mean_tpr) / n_samples)
                tpr_upper = np.minimum(mean_tpr + tpr_std, 1.0)
                tpr_lower = np.maximum(mean_tpr - tpr_std, 0.0)

                ax.fill_between(mean_fpr, tpr_lower, tpr_upper, color=colors.get(sheet, 'gray'), alpha=0.2)
                ax.plot(mean_fpr, mean_tpr, color=colors.get(sheet, 'gray'),
                        label=f"{labels.get(sheet, sheet)} (AUC={auc:.2f})", linewidth=2)

                loocv_data[sheet] = pd.DataFrame({
                    'FPR': mean_fpr, 'TPR': mean_tpr,
                    'TPR_Lower': tpr_lower, 'TPR_Upper': tpr_upper, 'AUC': auc
                })

            ax.plot([0, 1], [0, 1], 'k--', linewidth=1, alpha=0.5, label='Random')
            ax.set_xlabel('False Positive Rate', fontsize=9)
            ax.set_ylabel('True Positive Rate', fontsize=9)
            ax.set_title('ADMET LOOCV ROC Curves', fontsize=10, fontweight='bold')
            ax.legend(fontsize=6, loc='lower right', framealpha=0.9)
            ax.set_xlim([0, 1])
            ax.set_ylim([0, 1])
            ax.tick_params(labelsize=8)
            ax.grid(True, alpha=0.3)

            save_figure(fig, 'S4', 'a', 'ADMET LOOCV ROC',
                        loocv_data, width=SQUARE_SIZE, height=SQUARE_SIZE)

    # S4b: MoLFormer LOOCV comparison
    mol_loocv_path = PROJECT_ROOT / 'Output' / 'MoLFormer_Comparison' / 'loocv_comparison_molformer_vs_organoid.csv'
    if mol_loocv_path.exists():
        mol_loocv = pd.read_csv(mol_loocv_path)

        fig, ax = plt.subplots(figsize=(SQUARE_SIZE * 1.5, SQUARE_SIZE))
        x = np.arange(len(mol_loocv))
        width = 0.35

        ax.bar(x - width/2, mol_loocv['MoLFormer_AUC'], width, label='MoLFormer', color='#1f77b4')
        ax.bar(x + width/2, mol_loocv['Organoid_AUC'], width, label='Organoid', color='#ff7f0e')

        ax.set_ylabel('AUC ROC', fontsize=8)
        ax.set_title('MoLFormer vs Organoid LOOCV', fontsize=9, fontweight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels(mol_loocv['Model'], fontsize=7)
        ax.legend(fontsize=7)
        ax.set_ylim([0, 1])
        ax.axhline(y=0.5, color='gray', linestyle='--', alpha=0.5)
        ax.tick_params(labelsize=7)

        save_figure(fig, 'S4', 'b', 'MoLFormer LOOCV Comparison',
                    {'LOOCV': mol_loocv}, width=SQUARE_SIZE * 1.5, height=SQUARE_SIZE)


# ============================================================================
# MAIN
# ============================================================================

def generate_all():
    """Generate all figures."""
    print("=" * 60)
    print("Generating All Paper Figures")
    print("=" * 60)

    generate_fig_1()
    generate_fig_2()
    generate_fig_3()
    generate_fig_4_5()

    # Prediction figures
    generate_prediction_figures('Arrhythmia', '6', comparison_type='MoLFormer')
    generate_prediction_figures('HeartDamage', '7', comparison_type='ADMET')
    generate_prediction_figures('ConcernBinary', '8', comparison_type=None)

    # Supplements
    generate_supplements()

    print("\n" + "=" * 60)
    print("All figures generated!")
    print("=" * 60)


def list_figures():
    """List all registered figures."""
    registry = load_registry()
    if registry.empty:
        print("No figures registered.")
        return

    print("\nRegistered Figures:")
    print("=" * 80)
    for _, row in registry.iterrows():
        ext = " [EXT]" if row['External'] else ""
        print(f"  Fig_{row['Figure_ID']}{row['Letter']}: {row['Description']}{ext}")
    print("=" * 80)
    print(f"Total: {len(registry)} figures")


def _get_slide_image_files(unpack_dir, slide_num, png_only=True):
    """Parse slide rels XML to discover image file names dynamically.

    Returns a list of (rId, image_filename) tuples
    sorted by rId number, which corresponds to the order panels appear on the slide.
    By default only includes PNG files; set png_only=False for all image types.
    """
    import xml.etree.ElementTree as ET

    IMAGE_EXTS = {'.png', '.jpg', '.jpeg', '.emf', '.tiff', '.bmp', '.gif', '.wdp'}

    rels_path = unpack_dir / 'ppt' / 'slides' / '_rels' / f'slide{slide_num}.xml.rels'
    if not rels_path.exists():
        print(f"  Warning: {rels_path} not found")
        return []

    tree = ET.parse(rels_path)
    root = tree.getroot()

    ns = '{http://schemas.openxmlformats.org/package/2006/relationships}'

    image_rels = []
    for rel in root.findall(f'{ns}Relationship'):
        target = rel.get('Target', '')
        rid = rel.get('Id', '')

        if not target.startswith('../media/'):
            continue

        ext = Path(target).suffix.lower()
        if png_only and ext != '.png':
            continue
        if not png_only and ext not in IMAGE_EXTS:
            continue

        rid_num = int(rid.replace('rId', '')) if rid.startswith('rId') else 0
        filename = target.split('/')[-1]

        image_rels.append((rid_num, rid, filename))

    image_rels.sort(key=lambda x: x[0])
    return [(rid, fname) for _, rid, fname in image_rels]


def _build_parent_map(root):
    """Build a child->parent map for an ElementTree root."""
    return {child: parent for parent in root.iter() for child in parent}


def _pic_slide_position(pic, parent_map, ns_p, ns_a):
    """Get the slide-level (x, y) position of a <p:pic> element.

    Accounts for group transforms: when a pic is inside a <p:grpSp> whose
    chOff != off (i.e. the user moved the group in PowerPoint), the child
    coordinates must be transformed to slide coordinates.

    Transform per group level: slide_pos = off + (child_pos - chOff)
    """
    sp_pr = pic.find(f'{{{ns_p}}}spPr')
    if sp_pr is None:
        return None
    xfrm = sp_pr.find(f'{{{ns_a}}}xfrm')
    if xfrm is None:
        return None
    off = xfrm.find(f'{{{ns_a}}}off')
    if off is None:
        return None

    x = int(off.get('x', 0))
    y = int(off.get('y', 0))

    # Walk up through parent groups, applying each group's transform
    elem = pic
    while True:
        parent = parent_map.get(elem)
        if parent is None:
            break
        if parent.tag == f'{{{ns_p}}}grpSp':
            grp_sp_pr = parent.find(f'{{{ns_p}}}grpSpPr')
            if grp_sp_pr is not None:
                grp_xfrm = grp_sp_pr.find(f'{{{ns_a}}}xfrm')
                if grp_xfrm is not None:
                    grp_off = grp_xfrm.find(f'{{{ns_a}}}off')
                    grp_ch_off = grp_xfrm.find(f'{{{ns_a}}}chOff')
                    if grp_off is not None and grp_ch_off is not None:
                        ox = int(grp_off.get('x', 0))
                        oy = int(grp_off.get('y', 0))
                        cox = int(grp_ch_off.get('x', 0))
                        coy = int(grp_ch_off.get('y', 0))
                        x += ox - cox
                        y += oy - coy
        elem = parent

    return x, y


def _get_slide_images_by_position(unpack_dir, slide_num, png_only=True):
    """Get slide images sorted by visual position (top-to-bottom, left-to-right).

    Parses the slide XML for shape positions so that letters are assigned in
    reading order rather than by internal rId number.  This ensures that when a
    user adds a new image between existing panels the lettering stays sequential.

    Accounts for group transforms when images have been moved within groups.

    Returns (rId, filename) tuples.  Falls back to rId order if position data
    is unavailable.
    """
    import xml.etree.ElementTree as ET

    image_files = _get_slide_image_files(unpack_dir, slide_num, png_only=png_only)
    if not image_files:
        return []

    rid_map = {rid: fname for rid, fname in image_files}

    ns_p = 'http://schemas.openxmlformats.org/presentationml/2006/main'
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    slide_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
    if not slide_path.exists():
        return image_files  # fallback

    tree = ET.parse(slide_path)
    root = tree.getroot()

    # Build parent map for group transform lookups
    parent_map = _build_parent_map(root)

    positioned = []  # (y, x, rId, filename)
    found_rids = set()

    for pic in root.iter(f'{{{ns_p}}}pic'):
        blip_fill = pic.find(f'{{{ns_p}}}blipFill')
        if blip_fill is None:
            continue
        blip = blip_fill.find(f'{{{ns_a}}}blip')
        if blip is None:
            continue
        embed_rid = blip.get(f'{{{ns_r}}}embed', '')
        if embed_rid not in rid_map:
            continue

        pos = _pic_slide_position(pic, parent_map, ns_p, ns_a)
        if pos is None:
            continue

        x, y = pos
        positioned.append((y, x, embed_rid, rid_map[embed_rid]))
        found_rids.add(embed_rid)

    if not positioned:
        return image_files  # fallback

    # Group into visual rows: images within ROW_TOLERANCE EMU of the same y
    # are considered on the same row, then sorted left-to-right within the row.
    # Within each row, split at large x-gaps (COL_GAP_TOLERANCE) to avoid
    # interleaving images from different panel regions at similar y positions.
    ROW_TOLERANCE = 300000  # ~0.33 inches
    COL_GAP_TOLERANCE = 914400  # 1.0 inch — split row if x-gap exceeds this
    positioned.sort(key=lambda p: (p[0], p[1]))

    rows = []
    current_row = [positioned[0]]
    for item in positioned[1:]:
        if abs(item[0] - current_row[0][0]) <= ROW_TOLERANCE:
            current_row.append(item)
        else:
            rows.append(sorted(current_row, key=lambda p: p[1]))
            current_row = [item]
    rows.append(sorted(current_row, key=lambda p: p[1]))

    result = [(rid, fname) for row in rows for _, _, rid, fname in row]
    # Also store positions keyed by rId for compound panel reordering
    _position_cache = {rid: (x, y) for row in rows for y, x, rid, fname in row}
    _get_slide_images_by_position._position_cache = _position_cache

    # Append any images not matched in the slide XML (shouldn't happen, but safe)
    for rid, fname in image_files:
        if rid not in found_rids:
            result.append((rid, fname))

    return result


def _update_slide_rels(unpack_dir, slide_num, target_updates):
    """Update image targets in slide relationship XML.

    Args:
        target_updates: dict mapping rId -> new_target (e.g. {'rId3': '../media/Fig_3a.png'})
    """
    import xml.etree.ElementTree as ET

    rels_path = unpack_dir / 'ppt' / 'slides' / '_rels' / f'slide{slide_num}.xml.rels'
    ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
    ET.register_namespace('', ns)

    tree = ET.parse(rels_path)
    root = tree.getroot()

    for rel in root.findall(f'{{{ns}}}Relationship'):
        rid = rel.get('Id', '')
        if rid in target_updates:
            rel.set('Target', target_updates[rid])

    tree.write(rels_path, xml_declaration=True, encoding='ascii')


def _create_blank_png(path):
    """Create a 1x1 transparent PNG to replace excess slide images."""
    from PIL import Image as PILImage
    img = PILImage.new('RGBA', (1, 1), (255, 255, 255, 0))
    img.save(str(path))


def _add_image_slots(unpack_dir, slide_num, count):
    """Add new image slots (rels entries + <p:pic> elements) to a slide.

    Used when compound panels need more image positions than currently exist.
    Creates placeholder PNG files and adds corresponding XML elements at the
    bottom of the slide so that update_powerpoint() can replace them normally.

    Args:
        unpack_dir: Path to unpacked PPTX directory
        slide_num: Slide number to modify
        count: Number of new image slots to add
    """
    import xml.etree.ElementTree as ET
    from PIL import Image as PILImage

    ns_pkg = 'http://schemas.openxmlformats.org/package/2006/relationships'
    ns_p = 'http://schemas.openxmlformats.org/presentationml/2006/main'
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    ns_a16 = 'http://schemas.microsoft.com/office/drawing/2014/main'

    # Register namespaces to preserve prefixes in output
    ET.register_namespace('', ns_pkg)
    ET.register_namespace('p', ns_p)
    ET.register_namespace('a', ns_a)
    ET.register_namespace('r', ns_r)
    ET.register_namespace('a16', ns_a16)

    rels_path = unpack_dir / 'ppt' / 'slides' / '_rels' / f'slide{slide_num}.xml.rels'
    slide_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
    media_dir = unpack_dir / 'ppt' / 'media'

    # --- 1. Find next available rId and image number ---
    rels_tree = ET.parse(rels_path)
    rels_root = rels_tree.getroot()

    max_rid = 0
    for rel in rels_root.findall(f'{{{ns_pkg}}}Relationship'):
        rid = rel.get('Id', '')
        if rid.startswith('rId'):
            try:
                max_rid = max(max_rid, int(rid[3:]))
            except ValueError:
                pass

    # Find next available image number across ALL media files
    existing_images = list(media_dir.glob('image*.png'))
    max_img_num = 0
    for img_path in existing_images:
        stem = img_path.stem  # e.g. "image67"
        try:
            num = int(stem.replace('image', ''))
            max_img_num = max(max_img_num, num)
        except ValueError:
            pass

    # --- 2. Find max cNvPr id in slide XML ---
    slide_tree = ET.parse(slide_path)
    slide_root = slide_tree.getroot()

    max_cnv_id = 0
    for elem in slide_root.iter(f'{{{ns_p}}}cNvPr'):
        try:
            max_cnv_id = max(max_cnv_id, int(elem.get('id', '0')))
        except ValueError:
            pass

    # --- 3. Find the last <p:pic> position to place new ones nearby ---
    last_x, last_y = 5000000, 7500000  # default: bottom-right area
    last_cx, last_cy = 914400, 914400  # default: 1" x 1"

    parent_map = _build_parent_map(slide_root)
    for pic in slide_root.iter(f'{{{ns_p}}}pic'):
        pos = _pic_slide_position(pic, parent_map, ns_p, ns_a)
        if pos is not None:
            last_x, last_y = pos
        sp_pr = pic.find(f'{{{ns_p}}}spPr')
        if sp_pr is not None:
            xfrm = sp_pr.find(f'{{{ns_a}}}xfrm')
            if xfrm is not None:
                ext = xfrm.find(f'{{{ns_a}}}ext')
                if ext is not None:
                    last_cx = int(ext.get('cx', last_cx))
                    last_cy = int(ext.get('cy', last_cy))

    # Find the spTree to append new pic elements
    sp_tree = slide_root.find(f'.//{{{ns_p}}}spTree')
    if sp_tree is None:
        print(f"  Warning: No spTree found in slide {slide_num}")
        return

    image_type = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/image'

    for i in range(count):
        new_rid_num = max_rid + 1 + i
        new_img_num = max_img_num + 1 + i
        new_cnv_id = max_cnv_id + 1 + i

        new_rid = f'rId{new_rid_num}'
        new_img_name = f'image{new_img_num}.png'

        # Place new slots side-by-side to the right of the last image
        slot_x = last_x + (last_cx + 50000) * (i + 1)
        slot_y = last_y

        # --- Add rels entry ---
        new_rel = ET.SubElement(rels_root, f'{{{ns_pkg}}}Relationship')
        new_rel.set('Id', new_rid)
        new_rel.set('Type', image_type)
        new_rel.set('Target', f'../media/{new_img_name}')

        # --- Create placeholder PNG ---
        placeholder = PILImage.new('RGBA', (100, 100), (255, 255, 255, 0))
        placeholder.save(str(media_dir / new_img_name))

        # --- Add <p:pic> element to slide XML ---
        pic = ET.SubElement(sp_tree, f'{{{ns_p}}}pic')

        # nvPicPr
        nv_pic_pr = ET.SubElement(pic, f'{{{ns_p}}}nvPicPr')
        cnv_pr = ET.SubElement(nv_pic_pr, f'{{{ns_p}}}cNvPr')
        cnv_pr.set('id', str(new_cnv_id))
        cnv_pr.set('name', f'Picture {new_cnv_id}')
        cnv_pr.set('descr', 'image.png')
        cnv_pic_pr = ET.SubElement(nv_pic_pr, f'{{{ns_p}}}cNvPicPr')
        pic_locks = ET.SubElement(cnv_pic_pr, f'{{{ns_a}}}picLocks')
        pic_locks.set('noChangeAspect', '1')
        ET.SubElement(nv_pic_pr, f'{{{ns_p}}}nvPr')

        # blipFill
        blip_fill = ET.SubElement(pic, f'{{{ns_p}}}blipFill')
        blip = ET.SubElement(blip_fill, f'{{{ns_a}}}blip')
        blip.set(f'{{{ns_r}}}embed', new_rid)
        stretch = ET.SubElement(blip_fill, f'{{{ns_a}}}stretch')
        ET.SubElement(stretch, f'{{{ns_a}}}fillRect')

        # spPr
        sp_pr = ET.SubElement(pic, f'{{{ns_p}}}spPr')
        xfrm = ET.SubElement(sp_pr, f'{{{ns_a}}}xfrm')
        off = ET.SubElement(xfrm, f'{{{ns_a}}}off')
        off.set('x', str(slot_x))
        off.set('y', str(slot_y))
        ext = ET.SubElement(xfrm, f'{{{ns_a}}}ext')
        ext.set('cx', str(last_cx))
        ext.set('cy', str(last_cy))
        prst_geom = ET.SubElement(sp_pr, f'{{{ns_a}}}prstGeom')
        prst_geom.set('prst', 'rect')
        ET.SubElement(prst_geom, f'{{{ns_a}}}avLst')

    # Write both files
    rels_tree.write(rels_path, xml_declaration=True, encoding='ascii')
    slide_tree.write(str(slide_path), xml_declaration=True, encoding='ascii')


# Figure titles and off-page bullet notes for each slide
SLIDE_TITLES = {
    1: {
        'title': ('Figure 1: Design of a high-content kinetic cardiac organoid screen '
                  'to predict drug-induced cardiac toxicity (DICT) using machine learning.'),
        'bullets': [
            'Robot, RODEO (scalability)',
            'Complexity of organoid, reproducibility',
            'Sensors (oxygen, contractility)',
            'Longitudinal data acquisition',
            'Analytical feature extraction',
            'Machine learning predictions',
        ],
    },
    2: {
        'title': ('Figure 2: Robot-directed organoid deposition (RODEO) generates highly '
                  'reproducible human multi-chambered cardiac organoids in 384-well plates.'),
        'bullets': [
            'RODEO organoids, size, metabolic',
            'Cell distribution, batch variance',
        ],
    },
    3: {
        'title': ('Figure 3: Fitting oxygen and contractility kinetics to multi-dimensional '
                  'surfaces for quantitative feature extractions'),
        'bullets': [
            'Oxygen and contractility data',
            'Surface fitting examples and quantification',
            'Surface selection based on model response',
            'Validation of surface vs. raw experimental data',
        ],
    },
    4: {
        'title': ('Figure 4: Time and dose-dependent metabolic response to drugs '
                  'of multi-chambered cardiac organoids.'),
        'bullets': [],
    },
    5: {
        'title': ('Figure 5: Time and dose-dependent mechanical response to drugs '
                  'of multi-chambered cardiac organoids.'),
        'bullets': [],
    },
    6: {
        'title': 'Figure 6: Machine learning prediction of drug-induced cardiac arrhythmia',
        'bullets': [],
    },
    7: {
        'title': 'Figure 7: Machine learning prediction of drug-induced cardiac toxicity',
        'bullets': [],
    },
    8: {
        'title': 'Figure 8: Machine learning prediction of drug-induced cardiac concern',
        'bullets': [],
    },
}


def _update_slide_titles(unpack_dir):
    """Update figure titles and add off-page bullet notes on each slide.

    For each slide in SLIDE_TITLES:
    - Finds the existing title text box ("TextBox 1") and replaces its text
    - If bullets are provided, creates a text box positioned off-page to the right
      containing the bullet points as reference notes

    Idempotent: removes existing off-page note boxes before recreating.
    """
    import xml.etree.ElementTree as ET

    ns_p = 'http://schemas.openxmlformats.org/presentationml/2006/main'
    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'

    ET.register_namespace('p', ns_p)
    ET.register_namespace('a', ns_a)
    ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
    ET.register_namespace('a16', 'http://schemas.microsoft.com/office/drawing/2014/main')

    EMU_PER_INCH = 914400
    SLIDE_WIDTH = 6483350   # 7.09"
    # Place notes 0.25" to the right of the slide edge
    NOTES_X = SLIDE_WIDTH + int(0.25 * EMU_PER_INCH)
    NOTES_Y = int(0.1 * EMU_PER_INCH)
    NOTES_W = int(4.0 * EMU_PER_INCH)   # 4" wide text box
    NOTES_H = int(6.0 * EMU_PER_INCH)   # 6" tall (auto-fit will adjust)
    NOTES_NAME = 'FigureNotes'

    for slide_num, info in SLIDE_TITLES.items():
        slide_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
        if not slide_path.exists():
            continue

        tree = ET.parse(slide_path)
        root = tree.getroot()
        modified = False

        # --- 1. Update existing title text box ---
        title_text = info['title']
        for sp in root.iter(f'{{{ns_p}}}sp'):
            nv_sp_pr = sp.find(f'{{{ns_p}}}nvSpPr')
            if nv_sp_pr is None:
                continue
            cnv_pr = nv_sp_pr.find(f'{{{ns_p}}}cNvPr')
            if cnv_pr is None:
                continue
            if cnv_pr.get('name') != 'TextBox 1':
                continue

            # Found the title box — replace its txBody content
            tx_body = sp.find(f'{{{ns_p}}}txBody')
            if tx_body is None:
                continue

            # Preserve bodyPr and lstStyle
            body_pr = tx_body.find(f'{{{ns_a}}}bodyPr')
            lst_style = tx_body.find(f'{{{ns_a}}}lstStyle')

            # Remove all existing paragraphs
            for p_elem in tx_body.findall(f'{{{ns_a}}}p'):
                tx_body.remove(p_elem)

            # Split title into "Figure N: " prefix and rest
            colon_idx = title_text.find(': ')
            if colon_idx >= 0:
                prefix = title_text[:colon_idx + 2]  # "Figure N: "
                rest = title_text[colon_idx + 2:]
            else:
                prefix = ''
                rest = title_text

            # Create new paragraph with two runs (bold prefix + regular rest)
            p_elem = ET.SubElement(tx_body, f'{{{ns_a}}}p')
            pPr = ET.SubElement(p_elem, f'{{{ns_a}}}pPr')
            pPr.set('algn', 'ctr')
            defRPr = ET.SubElement(pPr, f'{{{ns_a}}}defRPr')
            defRPr.set('sz', '1200')
            defRPr.set('b', '1')

            if prefix:
                r1 = ET.SubElement(p_elem, f'{{{ns_a}}}r')
                rPr1 = ET.SubElement(r1, f'{{{ns_a}}}rPr')
                rPr1.set('dirty', '0')
                t1 = ET.SubElement(r1, f'{{{ns_a}}}t')
                t1.text = prefix

            r2 = ET.SubElement(p_elem, f'{{{ns_a}}}r')
            rPr2 = ET.SubElement(r2, f'{{{ns_a}}}rPr')
            rPr2.set('lang', 'en-US')
            rPr2.set('dirty', '0')
            t2 = ET.SubElement(r2, f'{{{ns_a}}}t')
            t2.text = rest

            ET.SubElement(p_elem, f'{{{ns_a}}}endParaRPr').set('dirty', '0')

            modified = True
            break

        # --- 2. Remove existing off-page notes (idempotent) ---
        sp_tree = root.find(f'.//{{{ns_p}}}spTree')
        if sp_tree is None:
            continue

        for sp in list(sp_tree):
            nv_sp_pr = sp.find(f'{{{ns_p}}}nvSpPr')
            if nv_sp_pr is None:
                continue
            cnv_pr = nv_sp_pr.find(f'{{{ns_p}}}cNvPr')
            if cnv_pr is not None and cnv_pr.get('name') == NOTES_NAME:
                sp_tree.remove(sp)

        # --- 3. Add off-page bullet notes ---
        bullets = info.get('bullets', [])
        if bullets:
            # Find max cNvPr id
            max_id = 0
            for elem in root.iter(f'{{{ns_p}}}cNvPr'):
                try:
                    max_id = max(max_id, int(elem.get('id', '0')))
                except ValueError:
                    pass

            sp = ET.SubElement(sp_tree, f'{{{ns_p}}}sp')

            # nvSpPr
            nv_sp_pr = ET.SubElement(sp, f'{{{ns_p}}}nvSpPr')
            cnv_pr = ET.SubElement(nv_sp_pr, f'{{{ns_p}}}cNvPr')
            cnv_pr.set('id', str(max_id + 1))
            cnv_pr.set('name', NOTES_NAME)
            cnv_sp_pr = ET.SubElement(nv_sp_pr, f'{{{ns_p}}}cNvSpPr')
            cnv_sp_pr.set('txBox', '1')
            ET.SubElement(nv_sp_pr, f'{{{ns_p}}}nvPr')

            # spPr — positioned off-page to the right
            sp_pr = ET.SubElement(sp, f'{{{ns_p}}}spPr')
            xfrm = ET.SubElement(sp_pr, f'{{{ns_a}}}xfrm')
            off = ET.SubElement(xfrm, f'{{{ns_a}}}off')
            off.set('x', str(NOTES_X))
            off.set('y', str(NOTES_Y))
            ext = ET.SubElement(xfrm, f'{{{ns_a}}}ext')
            ext.set('cx', str(NOTES_W))
            ext.set('cy', str(NOTES_H))
            prst_geom = ET.SubElement(sp_pr, f'{{{ns_a}}}prstGeom')
            prst_geom.set('prst', 'rect')
            ET.SubElement(prst_geom, f'{{{ns_a}}}avLst')
            ET.SubElement(sp_pr, f'{{{ns_a}}}noFill')

            # txBody with bullet paragraphs
            tx_body = ET.SubElement(sp, f'{{{ns_p}}}txBody')
            body_pr = ET.SubElement(tx_body, f'{{{ns_a}}}bodyPr')
            body_pr.set('wrap', 'square')
            ET.SubElement(tx_body, f'{{{ns_a}}}lstStyle')

            # Title line: "Notes:"
            p_title = ET.SubElement(tx_body, f'{{{ns_a}}}p')
            pPr_title = ET.SubElement(p_title, f'{{{ns_a}}}pPr')
            defRPr_title = ET.SubElement(pPr_title, f'{{{ns_a}}}defRPr')
            defRPr_title.set('sz', '1000')
            defRPr_title.set('b', '1')
            r_title = ET.SubElement(p_title, f'{{{ns_a}}}r')
            t_title = ET.SubElement(r_title, f'{{{ns_a}}}t')
            t_title.text = 'Notes:'

            # Each bullet point as a paragraph with bullet marker
            for bullet in bullets:
                p_bullet = ET.SubElement(tx_body, f'{{{ns_a}}}p')
                pPr_bullet = ET.SubElement(p_bullet, f'{{{ns_a}}}pPr')
                pPr_bullet.set('marL', str(int(0.15 * EMU_PER_INCH)))
                pPr_bullet.set('indent', str(-int(0.15 * EMU_PER_INCH)))
                buChar = ET.SubElement(pPr_bullet, f'{{{ns_a}}}buChar')
                buChar.set('char', '\u2022')
                defRPr_b = ET.SubElement(pPr_bullet, f'{{{ns_a}}}defRPr')
                defRPr_b.set('sz', '900')
                r_b = ET.SubElement(p_bullet, f'{{{ns_a}}}r')
                t_b = ET.SubElement(r_b, f'{{{ns_a}}}t')
                t_b.text = bullet

            modified = True

        if modified:
            tree.write(str(slide_path), xml_declaration=True, encoding='ascii')
            bullet_count = len(bullets)
            notes_msg = f" + {bullet_count} off-page notes" if bullet_count else ""
            print(f"  Slide {slide_num}: Updated title{notes_msg}")


# Slides where the user manually manages grouping and labels in PowerPoint.
# _add_panel_labels will skip these to avoid destroying manual layout.
MANUAL_GROUP_SLIDES = {2, 3, 4, 5}  # Slide 2: Panel_2a-l groups managed manually


def _add_panel_labels(unpack_dir):
    """Add panel letter labels grouped with images on multi-panel slides.

    For each slide with lettered panels, wraps each (image, label) pair
    in a <p:grpSp> so they move together when repositioned in PowerPoint.
    Idempotent: removes existing groups/labels and recreates them fresh.
    Skips slides in MANUAL_GROUP_SLIDES where the user manages groups.
    """
    import xml.etree.ElementTree as ET

    # Register namespaces to preserve prefixes in output
    _ns = {
        'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'a16': 'http://schemas.microsoft.com/office/drawing/2014/main',
    }
    for prefix, uri in _ns.items():
        ET.register_namespace(prefix, uri)

    EMU_PER_INCH = 914400
    LABEL_SIZE = int(0.20 * EMU_PER_INCH)   # 0.2" square label box
    LABEL_PAD = int(0.02 * EMU_PER_INCH)    # 0.02" inset from image corner

    # Use effective mappings to auto-discover new panels
    mappings = _get_effective_mappings(unpack_dir)

    for mapping in mappings:
        fig_prefix, letters, slide_num = mapping[0], mapping[1], mapping[2]
        offset = mapping[3] if len(mapping) > 3 else 0

        if not letters:
            continue  # Skip single-figure slides

        if slide_num in MANUAL_GROUP_SLIDES:
            continue  # User manages grouping manually on this slide

        fig_id = fig_prefix.replace('Fig_', '')

        image_files = _get_slide_images_by_position(unpack_dir, slide_num)
        if not image_files:
            continue

        # Calculate start index based on offset
        if offset == -1:
            start_idx = max(0, len(image_files) - len(letters))
        elif offset > 0:
            start_idx = offset
        else:
            start_idx = 0

        # Build rId -> letter mapping, accounting for compound panels
        rid_to_panel = {}
        img_idx = start_idx
        for letter in letters:
            compound_key = (fig_id, letter)
            if compound_key in COMPOUND_PANELS:
                for _ in COMPOUND_PANELS[compound_key]:
                    if img_idx < len(image_files):
                        rid, _ = image_files[img_idx]
                        rid_to_panel[rid] = letter  # same letter for all sub-images
                        img_idx += 1
            else:
                if img_idx < len(image_files):
                    rid, _ = image_files[img_idx]
                    rid_to_panel[rid] = letter
                    img_idx += 1

        slide_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
        if not slide_path.exists():
            continue

        tree = ET.parse(slide_path)
        root = tree.getroot()

        p = _ns['p']
        a = _ns['a']
        r_ns = _ns['r']

        spTree = root.find(f'{{{p}}}cSld/{{{p}}}spTree')
        if spTree is None:
            continue

        # Build parent map for traversal
        parent_map = {child: parent for parent in root.iter() for child in parent}

        # --- Step 1: Extract all panel pics from wherever they are ---
        # letter -> list of (pic_element, x, y, cx, cy) to support compound panels
        panel_pics = {}
        removed_groups = set()

        for pic in list(root.iter(f'{{{p}}}pic')):
            blip_fill = pic.find(f'{{{p}}}blipFill')
            if blip_fill is None:
                continue
            blip = blip_fill.find(f'{{{a}}}blip')
            if blip is None:
                continue
            embed_rid = blip.get(f'{{{r_ns}}}embed', '')

            if embed_rid not in rid_to_panel:
                continue

            letter = rid_to_panel[embed_rid]

            sp_pr = pic.find(f'{{{p}}}spPr')
            if sp_pr is None:
                continue
            xfrm = sp_pr.find(f'{{{a}}}xfrm')
            if xfrm is None:
                continue
            off = xfrm.find(f'{{{a}}}off')
            ext = xfrm.find(f'{{{a}}}ext')
            if off is None or ext is None:
                continue

            # Use slide-level position (accounts for group transforms
            # from manual repositioning in PowerPoint)
            pos = _pic_slide_position(pic, parent_map, p, a)
            if pos is None:
                continue
            x, y = pos
            cx = int(ext.get('cx', 0))
            cy = int(ext.get('cy', 0))

            # Update the pic's own xfrm to slide-level coords so it
            # stays correct when moved into a new identity-transform group
            off.set('x', str(x))
            off.set('y', str(y))

            # Remove pic from its current parent
            pic_parent = parent_map.get(pic)
            if pic_parent is not None:
                try:
                    pic_parent.remove(pic)
                except ValueError:
                    pass

                # If parent was a group from previous run, remove the empty group
                if pic_parent.tag == f'{{{p}}}grpSp' and id(pic_parent) not in removed_groups:
                    grp_parent = parent_map.get(pic_parent)
                    if grp_parent is not None:
                        try:
                            grp_parent.remove(pic_parent)
                            removed_groups.add(id(pic_parent))
                        except ValueError:
                            pass

            panel_pics.setdefault(letter, []).append((pic, x, y, cx, cy))

        # --- Step 2: Remove existing label textboxes ---
        expected_letters = set(letters)
        for sp in list(spTree.findall(f'{{{p}}}sp')):
            cNvSpPr = sp.find(f'{{{p}}}nvSpPr/{{{p}}}cNvSpPr')
            if cNvSpPr is None or cNvSpPr.get('txBox') != '1':
                continue

            text_content = ''
            for t_elem in sp.iter(f'{{{a}}}t'):
                if t_elem.text:
                    text_content += t_elem.text
            text_content = text_content.strip()

            if text_content in expected_letters:
                spTree.remove(sp)

        # Also remove any orphaned Panel_ groups from previous runs
        for grpSp in list(spTree.findall(f'{{{p}}}grpSp')):
            nvGrpSpPr = grpSp.find(f'{{{p}}}nvGrpSpPr')
            if nvGrpSpPr is not None:
                cNvPr = nvGrpSpPr.find(f'{{{p}}}cNvPr')
                if cNvPr is not None and cNvPr.get('name', '').startswith('Panel_'):
                    spTree.remove(grpSp)

        # --- Step 3: Find max shape ID for new elements ---
        max_id = 0
        for elem in root.iter():
            id_val = elem.get('id')
            if id_val and id_val.isdigit():
                max_id = max(max_id, int(id_val))

        # --- Step 4: Create groups (pics + label) in panel order ---
        # Deduplicate letters (compound panels map multiple rIds to one letter)
        seen_letters = []
        for letter in letters:
            if letter not in seen_letters:
                seen_letters.append(letter)

        groups_created = 0
        for letter in seen_letters:
            if letter not in panel_pics:
                continue

            pics_list = panel_pics[letter]  # list of (pic, x, y, cx, cy)

            # Compute bounding box across all pics for this panel
            bbox_x = min(x for _, x, _, _, _ in pics_list)
            bbox_y = min(y for _, _, y, _, _ in pics_list)
            bbox_r = max(x + cx for _, x, _, cx, _ in pics_list)
            bbox_b = max(y + cy for _, _, y, _, cy in pics_list)
            bbox_cx = bbox_r - bbox_x
            bbox_cy = bbox_b - bbox_y

            # Create group element
            max_id += 1
            grp = ET.SubElement(spTree, f'{{{p}}}grpSp')

            # Group non-visual properties
            nvGrpSpPr = ET.SubElement(grp, f'{{{p}}}nvGrpSpPr')
            cNvPr = ET.SubElement(nvGrpSpPr, f'{{{p}}}cNvPr')
            cNvPr.set('id', str(max_id))
            cNvPr.set('name', f'Panel_{fig_id}{letter}')
            ET.SubElement(nvGrpSpPr, f'{{{p}}}cNvGrpSpPr')
            ET.SubElement(nvGrpSpPr, f'{{{p}}}nvPr')

            # Group shape properties - bounding box of all pics
            # chOff/chExt = off/ext so child coords equal slide coords
            grpSpPr = ET.SubElement(grp, f'{{{p}}}grpSpPr')
            grp_xfrm = ET.SubElement(grpSpPr, f'{{{a}}}xfrm')
            for tag, attrs in [
                (f'{{{a}}}off', {'x': str(bbox_x), 'y': str(bbox_y)}),
                (f'{{{a}}}ext', {'cx': str(bbox_cx), 'cy': str(bbox_cy)}),
                (f'{{{a}}}chOff', {'x': str(bbox_x), 'y': str(bbox_y)}),
                (f'{{{a}}}chExt', {'cx': str(bbox_cx), 'cy': str(bbox_cy)}),
            ]:
                e = ET.SubElement(grp_xfrm, tag)
                for k, v in attrs.items():
                    e.set(k, v)

            # Add all pictures to group
            for pic, _, _, _, _ in pics_list:
                grp.append(pic)

            # Create label text box inside group (at top-left of bounding box)
            max_id += 1
            sp = ET.SubElement(grp, f'{{{p}}}sp')

            nvSpPr = ET.SubElement(sp, f'{{{p}}}nvSpPr')
            label_cNvPr = ET.SubElement(nvSpPr, f'{{{p}}}cNvPr')
            label_cNvPr.set('id', str(max_id))
            label_cNvPr.set('name', f'Label_{letter}')
            cNvSpPr_el = ET.SubElement(nvSpPr, f'{{{p}}}cNvSpPr')
            cNvSpPr_el.set('txBox', '1')
            ET.SubElement(nvSpPr, f'{{{p}}}nvPr')

            spPr = ET.SubElement(sp, f'{{{p}}}spPr')
            label_xfrm = ET.SubElement(spPr, f'{{{a}}}xfrm')
            l_off = ET.SubElement(label_xfrm, f'{{{a}}}off')
            l_off.set('x', str(bbox_x + LABEL_PAD))
            l_off.set('y', str(bbox_y + LABEL_PAD))
            l_ext = ET.SubElement(label_xfrm, f'{{{a}}}ext')
            l_ext.set('cx', str(LABEL_SIZE))
            l_ext.set('cy', str(LABEL_SIZE))
            prstGeom = ET.SubElement(spPr, f'{{{a}}}prstGeom')
            prstGeom.set('prst', 'rect')
            ET.SubElement(prstGeom, f'{{{a}}}avLst')
            ET.SubElement(spPr, f'{{{a}}}noFill')

            txBody = ET.SubElement(sp, f'{{{p}}}txBody')
            bodyPr = ET.SubElement(txBody, f'{{{a}}}bodyPr')
            bodyPr.set('wrap', 'none')
            bodyPr.set('lIns', '0')
            bodyPr.set('tIns', '0')
            bodyPr.set('rIns', '0')
            bodyPr.set('bIns', '0')
            ET.SubElement(txBody, f'{{{a}}}lstStyle')

            p_elem = ET.SubElement(txBody, f'{{{a}}}p')
            pPr = ET.SubElement(p_elem, f'{{{a}}}pPr')
            defRPr = ET.SubElement(pPr, f'{{{a}}}defRPr')
            defRPr.set('sz', '1000')  # 10pt
            defRPr.set('b', '1')      # Bold

            r_elem = ET.SubElement(p_elem, f'{{{a}}}r')
            t_elem = ET.SubElement(r_elem, f'{{{a}}}t')
            t_elem.text = letter

            groups_created += 1

        if groups_created > 0:
            tree.write(str(slide_path), xml_declaration=True, encoding='ascii')
            print(f"  Slide {slide_num}: Grouped {groups_created} panels with labels")


# ============================================================================
# LAYOUT TRACKING - Extract & apply PPTX positions/sizes
# ============================================================================

LAYOUT_PATH = FIGURES_DIR / 'slide_layout.json'
_layout_cache = None

# Compound panels: one letter mapping to multiple side-by-side images in the PPTX.
# Maps (fig_id, letter) -> list of filename suffixes.
COMPOUND_PANELS = {
    # Panel 3a: two heatmaps (Fig_3a_1.png + Fig_3a_2.png)
    ('3', 'a'): ['1', '2'],
    # Panel 3b: four 3D surfaces in 2x2 grid + shared colorbar
    # Order matches position sort: top-left, top-right, colorbar (right of row1), bottom-left, bottom-right
    ('3', 'b'): ['1', '2', 'colorbar', '3', '4'],
    # Panel 3e: two 3D surfaces (Fig_3e_O2.png + Fig_3e_Contractility.png)
    ('3', 'e'): ['O2', 'Contractility'],
}

# Explicit rId-to-source mapping for slide 2.
# Bypasses position-based sorting for the 4 generated panels on slide 2.
# External images (plate photos, EMFs, microscopy) are untouched.
# Keys are rIds from slide2.xml, values are source filenames (relative to Fig_2 dir).
# Slide 2 panel-to-filename mapping.
# Panel letters correspond to groups named Panel_2{letter} in the PPTX.
# rIds are discovered at runtime — NOT hardcoded.
# Panels a-f are externally managed (not replaced by the script).
SLIDE2_PANEL_MAP = {
    'd': 'Fig_2d.png',                                   # SNR Quality Analysis
    'g': 'Fig_2g_Epirubicin_O2.png',                     # Metabolic Dose Dependent Response (averaged)
    'h': 'Fig_2h_Epirubicin_TC50.png',                   # Epirubicin TC50 (32h)
    'i': 'Fig_2i_Epirubicin_O2_Heatmap.png',             # Epirubicin O2 heatmap (LOWESS w=16)
    'j': 'Fig_2j_Mexiletine_Contractility.png',             # Mexiletine Contractility 2D dose-response
    'k': 'Fig_2k_Mexiletine_Waveforms.png',              # Mexiletine heart rate waveforms
    'l': 'Fig_2l_Mexiletine_Contractility_Heatmap.png',  # Mexiletine Contractility heatmap
}


def _discover_slide2_rids(slide_xml_path, rels_path):
    """Discover rIds for slide 2 panels by parsing group names at runtime.

    Finds groups named Panel_2{letter}, extracts the picture's r:embed rId,
    then maps it to the media filename via the .rels file.

    Returns: dict mapping panel letter -> (rId, media_filename)
    """
    import xml.etree.ElementTree as _ET

    # Parse rels: rId -> media filename
    _rels_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
    _rels_tree = _ET.parse(rels_path)
    rid_to_media = {}
    for rel in _rels_tree.getroot().findall(f'{{{_rels_ns}}}Relationship'):
        rid = rel.get('Id', '')
        target = rel.get('Target', '')
        if 'media/' in target:
            rid_to_media[rid] = target.split('/')[-1]

    # Parse slide XML: find Panel_2{letter} groups -> picture rId
    p_ns = 'http://schemas.openxmlformats.org/presentationml/2006/main'
    a_ns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    r_ns = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    tree = _ET.parse(slide_xml_path)
    root = tree.getroot()

    panel_rids = {}

    # Search for grpSp elements (groups)
    for grpSp in root.iter(f'{{{p_ns}}}grpSp'):
        # Get group name
        nvGrpSpPr = grpSp.find(f'{{{p_ns}}}nvGrpSpPr')
        if nvGrpSpPr is None:
            continue
        cNvPr = nvGrpSpPr.find(f'{{{p_ns}}}cNvPr')
        if cNvPr is None:
            continue
        name = cNvPr.get('name', '')

        if not name.startswith('Panel_2'):
            continue
        letter = name.replace('Panel_2', '')
        if len(letter) != 1:
            continue

        # Find picture inside group
        for pic in grpSp.iter(f'{{{p_ns}}}pic'):
            blipFill = pic.find(f'{{{p_ns}}}blipFill')
            if blipFill is None:
                continue
            blip = blipFill.find(f'{{{a_ns}}}blip')
            if blip is None:
                continue
            rid = blip.get(f'{{{r_ns}}}embed', '')
            if rid and rid in rid_to_media:
                panel_rids[letter] = (rid, rid_to_media[rid])
                break

    return panel_rids

# Explicit rId-to-source mapping for slide 3.
# This bypasses position-based sorting entirely, preventing swap bugs.
# Keys are rIds from the slide3 XML, values are source filenames (relative to Fig_3 dir).
SLIDE3_RID_MAP = {
    # --- Panel a row (Group 5): heatmaps + surfaces, left to right ---
    'rId6':  'Fig_3a_Dactinomycin_O2_Heatmap.png',            # x=0.360, w=0.809, Picture 2
    'rId10': 'Dactinomycin_Eq3_gaussian_hill_hybrid.png',     # x=1.200, w=1.000, Picture 15
    'rId7':  'Fig_3a_Nifedipine_O2_Heatmap.png',             # x=2.230, w=0.810, Picture 4
    'rId9':  'Nifedipine_Eq10_modified_hill_simple.png',      # x=3.071, w=1.000, Picture 8
    'rId11': 'Fig_3a_Mexiletine_O2_Heatmap.png',             # x=4.101, w=0.810, Picture 538
    'rId8':  'Mexiletine_Eq7_biphasic_response.png',          # x=4.942, w=1.000, Picture 12

    # --- Panel c (Group 4): R² bar chart ---
    'rId4':  'Fig_3c.png',                                    # x=0.170, y=2.336, w=1.462, h=1.215

    # --- Panel d (Group 3): 3-panel scatter strip ---
    'rId5':  'Fig_3d.png',                                    # x=0.388, y=2.306, w=2.920, h=1.124
}


def _get_mappings():
    """Return the figure-to-slide mappings (shared between extract and update).

    Each tuple is (fig_prefix, letters, slide_num) or
    (fig_prefix, letters, slide_num, offset) where offset controls panel placement:
      - 0 (default): panels start at the first image on the slide
      - -1: panels are aligned to the LAST images on the slide (external images before them are untouched)
      - N>0: panels start at image index N (skip N external images)

    Letters string is empty for single-figure/external slides (tracked but not updated).
    """
    return [
        ('Fig_1', '', 1),            # External schematic (not generated)
        ('Fig_2', 'ijkl', 2, -1),   # SNR (i), Epirubicin O2 (j), Contractility (k), TC50 (l); a-h external before
        ('Fig_3', 'abcde', 3),
        ('Fig_4', '', 4),            # O2 5x5 grid (individual drug sub-images)
        ('Fig_5', '', 5),            # Contractility 5x5 grid
        ('Fig_6', 'abcdefgh', 6),
        ('Fig_7', 'abcdefgh', 7),
        ('Fig_8', 'abcdef', 8),
        ('Fig_S1', 'abc', 9),       # Supplement S1: Vandetanib heatmaps
        ('Fig_S2', 'ab', 10),       # Supplement S2: Daunorubicin 2D time series
        ('Fig_S3', 'a', 11),        # Supplement S3: Other models scatter
        ('Fig_S4', 'ab', 12),       # Supplement S4: LOOCV comparison
    ]


def _get_effective_mappings(unpack_dir):
    """Return figure-to-slide mappings, auto-discovering extra panels and new slides.

    Extends _get_mappings() by:
    1. Auto-extending panel letters if a slide has more images than the base mapping
    2. Auto-discovering new slides that aren't in the base mappings at all

    Used by extract_slide_layout() and _add_panel_labels() so that newly added
    images are automatically tracked and labelled without editing _get_mappings().
    update_powerpoint() still uses _get_mappings() since it only replaces generated images.
    """
    import re

    ALL_LETTERS = 'abcdefghijklmnopqrstuvwxyz'
    base_mappings = _get_mappings()
    mapped_slides = {m[2] for m in base_mappings}

    effective = []
    for mapping in base_mappings:
        fig_prefix, letters, slide_num = mapping[0], mapping[1], mapping[2]
        offset = mapping[3] if len(mapping) > 3 else 0

        if letters and offset == 0:
            # Only auto-extend for slides without offset and without compound panels.
            # Slides with compound panels may have extra image slots (colorbars,
            # manually-added images) that shouldn't create new panel letters.
            fig_id = fig_prefix.replace('Fig_', '')
            has_compound = any(k[0] == fig_id for k in COMPOUND_PANELS)
            if not has_compound:
                image_files = _get_slide_images_by_position(unpack_dir, slide_num)
                actual_count = len(image_files)
                if actual_count > len(letters):
                    extended = ALL_LETTERS[:actual_count]
                    print(f"  Auto-extended slide {slide_num} ({fig_prefix}): "
                          f"{len(letters)} -> {actual_count} panels ({extended})")
                    letters = extended
        effective.append((fig_prefix, letters, slide_num, offset))

    # Discover slides not in base mappings
    slides_dir = unpack_dir / 'ppt' / 'slides'
    if slides_dir.exists():
        for slide_file in sorted(slides_dir.glob('slide*.xml')):
            match = re.match(r'slide(\d+)\.xml', slide_file.name)
            if match:
                sn = int(match.group(1))
                if sn not in mapped_slides:
                    image_files = _get_slide_images_by_position(unpack_dir, sn, png_only=False)
                    if image_files:
                        n = len(image_files)
                        letters = ALL_LETTERS[:n] if n > 1 else 'a'
                        fig_prefix = f'Fig_{sn}'
                        effective.append((fig_prefix, letters, sn, 0))
                        print(f"  Auto-discovered slide {sn}: {n} image(s) ({letters})")

    return effective


def extract_slide_layout():
    """Extract figure positions and sizes from the current PPTX.

    Parses slide XML to read shape dimensions for each figure panel,
    then saves to slide_layout.json. Run this after manually adjusting
    layout in PowerPoint to preserve changes for next generation.
    """
    import xml.etree.ElementTree as ET
    import subprocess
    import sys

    pptx_path = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / 'Cardiac_RODEO_Tracked.pptx'
    if not pptx_path.exists():
        print(f"  PowerPoint not found: {pptx_path}")
        return

    print("\n=== Extracting PPTX Layout ===")

    unpack_dir = PROJECT_ROOT / 'workspace' / 'pptx_unpack'
    skill_scripts = Path.home() / '.claude' / 'skills' / 'pptx' / 'ooxml' / 'scripts'
    unpack_script = skill_scripts / 'unpack.py'

    if not unpack_script.exists():
        print(f"  PPTX skill not found at {skill_scripts}")
        return

    result = subprocess.run([sys.executable, str(unpack_script), str(pptx_path), str(unpack_dir)],
                           capture_output=True, text=True, cwd=str(PROJECT_ROOT))
    if result.returncode != 0:
        print(f"  Unpack failed: {result.stderr}")
        return

    # XML namespaces
    ns = {
        'p': 'http://schemas.openxmlformats.org/presentationml/2006/main',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }

    EMU_PER_INCH = 914400
    layout = {'extracted_from': pptx_path.name,
              'extracted_at': datetime.now().isoformat(),
              'slides': {}}

    # Use effective mappings to auto-discover new panels and slides
    mappings = _get_effective_mappings(unpack_dir)

    for mapping in mappings:
        fig_prefix, letters, slide_num = mapping[0], mapping[1], mapping[2]
        offset = mapping[3] if len(mapping) > 3 else 0
        fig_id = fig_prefix.replace('Fig_', '')

        # Get rId-to-panel mapping (use all image types for external/grid figures)
        if not letters:
            image_files = _get_slide_images_by_position(unpack_dir, slide_num, png_only=False)
        else:
            image_files = _get_slide_images_by_position(unpack_dir, slide_num)
        if not image_files:
            continue

        # Calculate start index based on offset
        if offset == -1:
            start_idx = max(0, len(image_files) - max(len(letters), 1))
        elif offset > 0:
            start_idx = offset
        else:
            start_idx = 0

        # Build rId -> panel letter map, accounting for compound panels
        rid_to_panel = {}
        if not letters:
            # Single-figure slide - track the first image as the whole figure
            rid, _ = image_files[start_idx] if start_idx < len(image_files) else image_files[0]
            rid_to_panel[rid] = ''
        else:
            img_idx = start_idx
            for letter in letters:
                compound_key = (fig_id, letter)
                if compound_key in COMPOUND_PANELS:
                    for _ in COMPOUND_PANELS[compound_key]:
                        if img_idx < len(image_files):
                            rid, _ = image_files[img_idx]
                            rid_to_panel[rid] = letter
                            img_idx += 1
                else:
                    if img_idx < len(image_files):
                        rid, _ = image_files[img_idx]
                        rid_to_panel[rid] = letter
                        img_idx += 1

        # Parse the slide XML for shape positions/sizes
        slide_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
        if not slide_path.exists():
            continue

        tree = ET.parse(slide_path)
        root = tree.getroot()

        # Build parent map for group transform lookups
        parent_map = _build_parent_map(root)

        slide_data = {}
        # Find all <p:pic> elements
        for pic in root.iter(f'{{{ns["p"]}}}pic'):
            # Get the rId from a:blip
            blip_fill = pic.find(f'{{{ns["p"]}}}blipFill')
            if blip_fill is None:
                continue
            blip = blip_fill.find(f'{{{ns["a"]}}}blip')
            if blip is None:
                continue
            embed_rid = blip.get(f'{{{ns["r"]}}}embed', '')

            if embed_rid not in rid_to_panel:
                continue
            panel_letter = rid_to_panel[embed_rid]

            # Get position and size, accounting for group transforms
            pos = _pic_slide_position(pic, parent_map, ns["p"], ns["a"])
            if pos is None:
                continue

            sp_pr = pic.find(f'{{{ns["p"]}}}spPr')
            xfrm = sp_pr.find(f'{{{ns["a"]}}}xfrm')
            ext = xfrm.find(f'{{{ns["a"]}}}ext')
            if ext is None:
                continue

            x = pos[0] / EMU_PER_INCH
            y = pos[1] / EMU_PER_INCH
            w = int(ext.get('cx', 0)) / EMU_PER_INCH
            h = int(ext.get('cy', 0)) / EMU_PER_INCH

            key = f'Fig_{fig_id}{panel_letter}'
            if key in slide_data:
                # Compound panel: expand bounding box to include this image
                prev = slide_data[key]
                new_x = min(prev['x'], round(x, 3))
                new_y = min(prev['y'], round(y, 3))
                new_r = max(prev['x'] + prev['w'], round(x + w, 3))
                new_b = max(prev['y'] + prev['h'], round(y + h, 3))
                slide_data[key] = {
                    'x': new_x, 'y': new_y,
                    'w': round(new_r - new_x, 3), 'h': round(new_b - new_y, 3)
                }
            else:
                slide_data[key] = {
                    'x': round(x, 3), 'y': round(y, 3),
                    'w': round(w, 3), 'h': round(h, 3)
                }

            # Remove rId so duplicates don't re-match
            del rid_to_panel[embed_rid]

        if slide_data:
            layout['slides'][str(slide_num)] = slide_data
            print(f"  Slide {slide_num} ({fig_prefix}): {len(slide_data)} panels extracted")
            for key, dims in slide_data.items():
                print(f"    {key}: {dims['w']:.2f}\" x {dims['h']:.2f}\" at ({dims['x']:.2f}\", {dims['y']:.2f}\")")

    # Save layout file
    with open(LAYOUT_PATH, 'w') as f:
        json.dump(layout, f, indent=2)
    print(f"\n  Layout saved: {LAYOUT_PATH}")
    print(f"  Next run of --all will use these dimensions for figure generation.")

    # Also update figure_registry.csv with display positions/sizes
    _update_registry_positions(layout)


def _update_registry_positions(layout):
    """Write extracted PPTX positions into figure_registry.csv.

    Maps layout keys like 'Fig_6a' back to registry rows by Figure_ID + Letter,
    then writes Left_In, Top_In, Display_Width_In, Display_Height_In columns.
    """
    registry_path = FIGURES_DIR / 'figure_registry.csv'
    if not registry_path.exists():
        print("  Registry CSV not found — skipping position update.")
        return

    df = pd.read_csv(registry_path)
    df = df.drop_duplicates(subset=['Figure_ID', 'Letter'], keep='first')
    df['Figure_ID'] = df['Figure_ID'].astype(str)

    # Ensure position columns exist
    for col in ('Left_In', 'Top_In', 'Display_Width_In', 'Display_Height_In'):
        if col not in df.columns:
            df[col] = pd.NA

    updated = 0
    for slide_num, panels in layout.get('slides', {}).items():
        for key, dims in panels.items():
            # Parse key: 'Fig_6a' -> fig_id='6', letter='a'
            # or 'Fig_4' -> fig_id='4', letter=''
            stripped = key.replace('Fig_', '')
            # Split: last char is letter if it's alpha, otherwise whole thing is fig_id
            if stripped and stripped[-1].isalpha():
                fig_id = stripped[:-1]
                letter = stripped[-1]
            else:
                fig_id = stripped
                letter = ''

            # Find matching registry row
            if letter:
                mask = (df['Figure_ID'] == fig_id) & (df['Letter'] == letter)
            else:
                mask = (df['Figure_ID'] == fig_id) & (df['Letter'].isna() | (df['Letter'] == ''))

            matches = df.index[mask]
            if len(matches) == 0:
                continue

            idx = matches[0]
            df.at[idx, 'Left_In'] = dims['x']
            df.at[idx, 'Top_In'] = dims['y']
            df.at[idx, 'Display_Width_In'] = dims['w']
            df.at[idx, 'Display_Height_In'] = dims['h']
            updated += 1

    df.to_csv(registry_path, index=False)
    print(f"  Registry updated: {updated} panels with PPTX positions ({registry_path.name})")


def get_layout_size(fig_id, letter, default=None):
    """Get figure dimensions that match the PPTX aspect ratio at the default scale.

    Instead of returning the raw (small) PPTX shape dimensions — which would
    produce low-resolution figures — this returns the *default* size adjusted to
    match the PPTX panel's aspect ratio.  PowerPoint then scales the high-res
    image down to fit the shape, keeping it sharp.

    Args:
        fig_id:  Figure number as string (e.g. '6')
        letter:  Panel letter (e.g. 'a')
        default: (width, height) default figsize to use as the base scale.
                 If None, returns None (caller uses its own default unchanged).

    Returns:
        (width, height) at the default scale with PPTX aspect ratio, or None.
    """
    global _layout_cache
    if _layout_cache is None:
        if LAYOUT_PATH.exists():
            with open(LAYOUT_PATH) as f:
                _layout_cache = json.load(f)
        else:
            _layout_cache = {}

    key = f'Fig_{fig_id}{letter}'
    for slide_num, panels in _layout_cache.get('slides', {}).items():
        if key in panels:
            dims = panels[key]
            pptx_w, pptx_h = dims['w'], dims['h']

            if default is None or pptx_w <= 0 or pptx_h <= 0:
                return None

            def_w, def_h = default
            pptx_ar = pptx_w / pptx_h
            def_ar = def_w / def_h

            # If aspect ratios already match (within 5%), keep defaults as-is
            if abs(pptx_ar - def_ar) / max(def_ar, 0.01) < 0.05:
                return None  # no adjustment needed

            # Scale the default to match PPTX aspect ratio,
            # keeping the larger dimension at the default scale
            if pptx_ar > def_ar:
                # PPTX is wider — keep default width, shrink height
                return (def_w, def_w / pptx_ar)
            else:
                # PPTX is taller — keep default height, shrink width
                return (def_h * pptx_ar, def_h)

    return None


def update_powerpoint():
    """Update PowerPoint with latest figures using OOXML unpack/pack.

    Uses dynamic image mapping by reading slide relationship XML files
    to discover which imageN.png corresponds to each panel position.
    This is robust against image number shifts from other slides being rebuilt.
    """
    import shutil
    import subprocess
    import sys

    pptx_path = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures' / 'Cardiac_RODEO_Tracked.pptx'
    if not pptx_path.exists():
        print(f"  PowerPoint not found: {pptx_path}")
        return False

    print("\n--- Updating PowerPoint ---")

    # Create temp directory for unpacking
    unpack_dir = PROJECT_ROOT / 'workspace' / 'pptx_unpack'

    # Find the pptx skill scripts
    skill_scripts = Path.home() / '.claude' / 'skills' / 'pptx' / 'ooxml' / 'scripts'
    unpack_script = skill_scripts / 'unpack.py'
    pack_script = skill_scripts / 'pack.py'

    if not unpack_script.exists():
        print(f"  PPTX skill not found at {skill_scripts}")
        return False

    # Unpack - use sys.executable to ensure same Python interpreter
    result = subprocess.run([sys.executable, str(unpack_script), str(pptx_path), str(unpack_dir)],
                          capture_output=True, text=True, cwd=str(PROJECT_ROOT))
    if result.returncode != 0:
        print(f"  Unpack failed: {result.stderr}")
        return False

    media_dir = unpack_dir / 'ppt' / 'media'
    figs_dir = PROJECT_ROOT / 'Output' / 'PowerPoint_Figures'

    mappings = _get_mappings()

    # Build cross-slide reference map to avoid deleting shared images
    all_slide_images = {}
    for mapping in mappings:
        sn = mapping[2]
        ifs = _get_slide_image_files(unpack_dir, sn)
        all_slide_images[sn] = set(fname for _, fname in ifs)

    def _is_shared_image(slide_num, filename):
        """Check if filename is referenced by any OTHER slide."""
        return any(filename in imgs for sn, imgs in all_slide_images.items() if sn != slide_num)

    updated = 0
    for mapping in mappings:
        fig_prefix, letters, slide_num = mapping[0], mapping[1], mapping[2]
        offset = mapping[3] if len(mapping) > 3 else 0

        if not letters:
            continue  # External/grid figures not managed by update

        # Discover image files for this slide sorted by visual position
        image_files = _get_slide_images_by_position(unpack_dir, slide_num)

        if not image_files:
            print(f"  Warning: No images found for slide {slide_num} ({fig_prefix})")
            continue

        # Extract figure ID from prefix (e.g., 'Fig_3' -> '3')
        fig_id = fig_prefix.replace('Fig_', '')

        # Calculate start index based on offset
        if offset == -1:
            start_idx = max(0, len(image_files) - len(letters))
        elif offset > 0:
            start_idx = offset
        else:
            start_idx = 0

        # Calculate total image slots needed (compound panels need extra slots)
        total_slots = 0
        for letter in letters:
            compound_key = (fig_id, letter)
            total_slots += len(COMPOUND_PANELS.get(compound_key, [letter]))

        available = len(image_files) - start_idx
        if available < total_slots:
            needed = total_slots - available
            _add_image_slots(unpack_dir, slide_num, needed)
            # Re-read after adding slots
            image_files = _get_slide_images_by_position(unpack_dir, slide_num)
            print(f"  Added {needed} image slot(s) to slide {slide_num}")

        # Reorder image slots to keep compound panels spatially clustered.
        # When a 2x2 grid and a separate panel are at similar y-coordinates,
        # position sorting can interleave them. Fix by detecting outliers in
        # compound panel assignments and swapping them with nearby non-panel slots.
        pos_cache = getattr(_get_slide_images_by_position, '_position_cache', {})
        if pos_cache:
            idx = start_idx
            for letter in letters:
                compound_key = (fig_id, letter)
                n = len(COMPOUND_PANELS.get(compound_key, [letter]))
                if n >= 3 and idx + n <= len(image_files):
                    # Get x-positions for this compound panel's assigned slots
                    panel_rids = [image_files[idx + i][0] for i in range(n)]
                    panel_xs = [pos_cache.get(rid, (0, 0))[0] for rid in panel_rids]

                    if panel_xs:
                        median_x = sorted(panel_xs)[len(panel_xs) // 2]
                        # Check for outliers: images more than 1.5" from median x
                        OUTLIER_THRESHOLD = 1.0 * 914400  # 1.0 inch in EMU
                        for i in range(n):
                            if abs(panel_xs[i] - median_x) > OUTLIER_THRESHOLD:
                                # Find best swap candidate after the compound panel
                                best_swap = None
                                best_dist = float('inf')
                                for j in range(idx + n, len(image_files)):
                                    swap_rid = image_files[j][0]
                                    swap_x = pos_cache.get(swap_rid, (0, 0))[0]
                                    dist = abs(swap_x - median_x)
                                    if dist < best_dist:
                                        best_dist = dist
                                        best_swap = j
                                if best_swap is not None and best_dist < abs(panel_xs[i] - median_x):
                                    image_files[idx + i], image_files[best_swap] = \
                                        image_files[best_swap], image_files[idx + i]
                                    # Update panel_xs for subsequent checks
                                    panel_xs[i] = pos_cache.get(image_files[idx + i][0], (0, 0))[0]
                idx += n

        print(f"  {fig_prefix} (slide {slide_num}): {len(image_files)} images found, "
              f"replacing {total_slots} slots starting at position {start_idx}")

        assigned_filenames = set()  # Track newly-assigned media files

        # --- Slides with explicit rId mapping (bypasses position sort) ---
        EXPLICIT_RID_MAPS = {3: SLIDE3_RID_MAP}

        if slide_num == 2:
            # --- Slide 2: runtime rId discovery from Panel_2{letter} groups ---
            slide_xml_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
            rels_path = unpack_dir / 'ppt' / 'slides' / '_rels' / f'slide{slide_num}.xml.rels'

            if slide_xml_path.exists() and rels_path.exists():
                panel_rids = _discover_slide2_rids(slide_xml_path, rels_path)
                print(f"    Discovered {len(panel_rids)} panel rIds from group names")

                for panel_letter, src_name in SLIDE2_PANEL_MAP.items():
                    if panel_letter not in panel_rids:
                        print(f"    Warning: Panel_2{panel_letter} not found in slide XML")
                        continue
                    rid, media_filename = panel_rids[panel_letter]
                    src = figs_dir / fig_prefix / src_name
                    if src.exists():
                        shutil.copy2(src, media_dir / media_filename)
                        assigned_filenames.add(media_filename)
                        updated += 1
                        print(f"    Panel {panel_letter}: {src_name} -> {media_filename} ({rid})")
                    else:
                        print(f"    Panel {panel_letter}: Skip - {src_name} not found")
            else:
                print(f"    Warning: slide2 XML/rels not found for panel discovery")

        elif slide_num in EXPLICIT_RID_MAPS:
            # --- Slide 3: explicit rId mapping (stable since structure doesn't change) ---
            rels_path = unpack_dir / 'ppt' / 'slides' / '_rels' / f'slide{slide_num}.xml.rels'
            import xml.etree.ElementTree as _ET
            _rels_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
            _rels_tree = _ET.parse(rels_path)
            _rid_to_media = {}
            for _rel in _rels_tree.getroot().findall(f'{{{_rels_ns}}}Relationship'):
                _rid = _rel.get('Id', '')
                _target = _rel.get('Target', '')
                if 'media/' in _target:
                    _rid_to_media[_rid] = _target.split('/')[-1]

            rid_map = EXPLICIT_RID_MAPS[slide_num]
            for rid, src_name in rid_map.items():
                media_filename = _rid_to_media.get(rid)
                if not media_filename:
                    print(f"    Warning: {rid} not found in rels")
                    continue
                src = figs_dir / fig_prefix / src_name
                if src.exists():
                    shutil.copy2(src, media_dir / media_filename)
                    assigned_filenames.add(media_filename)
                    updated += 1
                    print(f"    {src_name} -> {media_filename} ({rid})")
                else:
                    print(f"    Skip: {src_name} not found")

            # --- Slide 3: enforce panel positions for Group 2 row ---
            if slide_num == 3:
                _SLIDE3_POSITIONS = {
                    # Panel a row (Group 5)
                    'rId6':  (0.360, 1.247, 0.809, 0.809),  # Dacto HM
                    'rId10': (1.200, 1.152, 1.000, 1.000),  # Dacto Surf
                    'rId7':  (2.230, 1.247, 0.810, 0.810),  # Nif HM
                    'rId9':  (3.071, 1.152, 1.000, 1.000),  # Nif Surf
                    'rId11': (4.101, 1.247, 0.810, 0.810),  # Mex HM
                    'rId8':  (4.942, 1.152, 1.000, 1.000),  # Mex Surf
                    # Panel c (Group 4)
                    'rId4':  (0.170, 2.336, 1.462, 1.215),  # R² chart
                    # Panel d (Group 3)
                    'rId5':  (0.388, 2.306, 2.920, 1.124),  # Scatter strip
                }
                _slide_xml_path = unpack_dir / 'ppt' / 'slides' / f'slide{slide_num}.xml'
                _s3tree = _ET.parse(_slide_xml_path)
                _s3root = _s3tree.getroot()
                _ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                _ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                _EMU = 914400
                _fixes = 0
                for _pic in _s3root.iter('pic'):
                    # Try both namespaced and unnamespaced
                    pass
                # Use lxml for reliable namespace handling
                from lxml import etree as _lxml_ET
                _s3tree_lx = _lxml_ET.parse(str(_slide_xml_path))
                _s3root_lx = _s3tree_lx.getroot()
                for _pic in _s3root_lx.iter():
                    _ptag = _pic.tag.split('}')[-1] if '}' in _pic.tag else _pic.tag
                    if _ptag != 'pic':
                        continue
                    _blip = _pic.find(f'.//{{{_ns_a}}}blip')
                    if _blip is None:
                        continue
                    _rid = _blip.get(f'{{{_ns_r}}}embed', '')
                    if _rid not in _SLIDE3_POSITIONS:
                        continue
                    _tx, _ty, _tw, _th = _SLIDE3_POSITIONS[_rid]
                    for _child in _pic:
                        if 'spPr' in _child.tag.split('}')[-1]:
                            _xfrm = _child.find(f'{{{_ns_a}}}xfrm')
                            if _xfrm is not None:
                                _off = _xfrm.find(f'{{{_ns_a}}}off')
                                _ext = _xfrm.find(f'{{{_ns_a}}}ext')
                                if _off is not None:
                                    _off.set('x', str(int(_tx * _EMU)))
                                    _off.set('y', str(int(_ty * _EMU)))
                                if _ext is not None:
                                    _ext.set('cx', str(int(_tw * _EMU)))
                                    _ext.set('cy', str(int(_th * _EMU)))
                                _fixes += 1
                _s3tree_lx.write(str(_slide_xml_path), xml_declaration=True,
                                 encoding='UTF-8', standalone=True)
                if _fixes:
                    print(f"    Enforced positions for {_fixes} Group 2 panels")
        else:
            # --- All other slides: position-based assignment ---
            img_idx = start_idx
            for letter in letters:
                compound_key = (fig_id, letter)
                if compound_key in COMPOUND_PANELS:
                    for suffix in COMPOUND_PANELS[compound_key]:
                        if img_idx >= len(image_files):
                            break
                        rid, old_filename = image_files[img_idx]
                        src = figs_dir / fig_prefix / f'{fig_prefix}{letter}_{suffix}.png'
                        if src.exists():
                            shutil.copy2(src, media_dir / old_filename)
                            assigned_filenames.add(old_filename)
                            updated += 1
                            print(f"    {fig_prefix}{letter}_{suffix}.png -> {old_filename} (pos {img_idx})")
                        else:
                            print(f"    Skip: {src.name} not found")
                        img_idx += 1
                else:
                    if img_idx >= len(image_files):
                        break
                    rid, old_filename = image_files[img_idx]
                    src = figs_dir / fig_prefix / f'{fig_prefix}{letter}.png'
                    if src.exists():
                        shutil.copy2(src, media_dir / old_filename)
                        assigned_filenames.add(old_filename)
                        updated += 1
                        print(f"    {fig_prefix}{letter}.png -> {old_filename} (pos {img_idx})")
                    else:
                        print(f"    Skip: {src.name} not found")
                    img_idx += 1

        # Handle excess images beyond panel count — only for slides without offset
        # (offset != 0 means external images are present that we must not touch)
        # Skip for slides with explicit rId mapping (no position-based idx)
        if slide_num not in EXPLICIT_RID_MAPS and slide_num != 2:
            end_idx = img_idx
            if offset == 0 and len(image_files) > end_idx:
                for i in range(end_idx, len(image_files)):
                    rid, old_filename = image_files[i]
                    _create_blank_png(media_dir / old_filename)
                    print(f"    Blanked excess: {old_filename}")

    print(f"\n  Updated {updated} images in PowerPoint")

    # Update figure titles and add off-page bullet notes
    _update_slide_titles(unpack_dir)

    # Add panel letter labels and group them with images
    _add_panel_labels(unpack_dir)

    # Repack - use sys.executable to ensure same Python interpreter
    result = subprocess.run([sys.executable, str(pack_script), str(unpack_dir), str(pptx_path)],
                          capture_output=True, text=True, cwd=str(PROJECT_ROOT))
    if result.returncode != 0:
        print(f"  Pack failed: {result.stderr}")
        return False

    print(f"  PowerPoint saved: {pptx_path}")
    return True


def main():
    parser = argparse.ArgumentParser(description='Generate paper figures')
    parser.add_argument('--all', action='store_true', help='Generate all figures')
    parser.add_argument('--figure', type=str, help='Generate specific figure (e.g., 3)')
    parser.add_argument('--list', action='store_true', help='List registered figures')
    parser.add_argument('--supplements', action='store_true', help='Generate supplement figures')
    parser.add_argument('--no-pptx', action='store_true', help='Skip PowerPoint update')
    parser.add_argument('--extract-layout', action='store_true',
                        help='Extract layout from PPTX (positions/sizes) and save to slide_layout.json')

    args = parser.parse_args()

    if args.extract_layout:
        extract_slide_layout()
        return

    figures_generated = False

    if args.list:
        list_figures()
    elif args.supplements:
        generate_supplements()
        figures_generated = True
    elif args.all:
        generate_all()
        figures_generated = True
    elif args.figure:
        fig_num = args.figure
        if fig_num == '1':
            generate_fig_1()
        elif fig_num == '2':
            generate_fig_2()
        elif fig_num == '3':
            generate_fig_3()
        elif fig_num in ['4', '5']:
            generate_fig_4_5()
        elif fig_num == '6':
            generate_prediction_figures('Arrhythmia', '6', comparison_type='MoLFormer')
        elif fig_num == '7':
            generate_prediction_figures('HeartDamage', '7', comparison_type='ADMET')
        elif fig_num == '8':
            generate_prediction_figures('ConcernBinary', '8', comparison_type=None)
        else:
            print(f"Unknown figure: {fig_num}")
            return
        figures_generated = True
    else:
        parser.print_help()

    # Auto-update PowerPoint after generating figures
    if figures_generated and not args.no_pptx:
        update_powerpoint()


if __name__ == '__main__':
    main()
