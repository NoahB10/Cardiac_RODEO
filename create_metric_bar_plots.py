"""
Create bar plots for prediction model metrics (Accuracy, F1, MCC)
for Arrhythmia, Heart Damage, and Concern models.
"""
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
from pathlib import Path

# Register Helvetica fonts from local fonts folder
_font_dir = Path(__file__).parent / 'fonts'
if _font_dir.exists():
    for _font_file in _font_dir.glob('*.ttf'):
        fm.fontManager.addfont(str(_font_file))
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Helvetica', 'Arial', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference

# Paths
PROJECT_ROOT = Path(__file__).parent
METRICS_PATH = PROJECT_ROOT / 'Output' / 'Performance_Metrics' / 'model_performance_summary.csv'
OUTPUT_DIR = PROJECT_ROOT / 'Output' / 'Excel_Figures'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Load metrics
df = pd.read_csv(METRICS_PATH)
print("Loaded metrics:")
print(df)

# Define the 3 targets to plot
TARGETS = {
    'Arrhythmia': 'Arrhythmia',
    'Heart_Damage': 'heart_damage',
    'Concern': 'Concern'
}

# Colors for bars
COLORS = {
    'Arrhythmia': ['#2ecc71', '#27ae60', '#1e8449'],  # Greens
    'Heart_Damage': ['#e74c3c', '#c0392b', '#922b21'],  # Reds
    'Concern': ['#3498db', '#2980b9', '#1f618d']  # Blues
}

def create_bar_plot(target_name, target_key, df, output_dir):
    """Create a bar plot for a single target."""
    row = df[df['Target'] == target_key].iloc[0]

    metrics = ['Accuracy', 'F1', 'MCC']
    values = [row['Accuracy_Mean'], row['F1_Mean'], row['MCC_Mean']]
    errors = [row['Accuracy_Std'], row['F1_Std'], row['MCC_Std']]

    # Create figure
    fig, ax = plt.subplots(figsize=(6, 5))

    x = np.arange(len(metrics))
    bars = ax.bar(x, values, yerr=errors, capsize=5,
                  color=COLORS[target_name], edgecolor='black', linewidth=1.2)

    # Add value labels on bars
    for bar, val in zip(bars, values):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height + 0.03,
                f'{val:.3f}', ha='center', va='bottom', fontsize=11, fontweight='bold')

    # Formatting
    ax.set_xlabel('Metric', fontsize=12, fontweight='bold')
    ax.set_ylabel('Score', fontsize=12, fontweight='bold')
    ax.set_title(f'{target_name.replace("_", " ")} - Model Performance', fontsize=14, fontweight='bold')
    ax.set_xticks(x)
    ax.set_xticklabels(metrics, fontsize=11)
    ax.set_ylim(0, 1.1)
    ax.axhline(y=0.5, color='gray', linestyle='--', alpha=0.5, label='Chance level')
    ax.legend(loc='upper right')
    ax.grid(True, axis='y', alpha=0.3)

    plt.tight_layout()

    # Save figure (PNG only)
    png_path = output_dir / f'{target_name}_Metrics_Bar.png'
    fig.savefig(png_path, dpi=300, bbox_inches='tight')
    plt.close(fig)

    print(f"Saved: {png_path.name}")
    return metrics, values, errors

def create_excel_with_chart(all_data, output_path):
    """Create Excel file with data and embedded charts."""
    wb = Workbook()

    for i, (target_name, data) in enumerate(all_data.items()):
        metrics, values, errors = data

        if i == 0:
            ws = wb.active
            ws.title = target_name
        else:
            ws = wb.create_sheet(target_name)

        # Write data
        ws['A1'] = 'Metric'
        ws['B1'] = 'Score'
        ws['C1'] = 'Std'

        for row_idx, (m, v, e) in enumerate(zip(metrics, values, errors), start=2):
            ws[f'A{row_idx}'] = m
            ws[f'B{row_idx}'] = v
            ws[f'C{row_idx}'] = e

        # Create bar chart
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"{target_name.replace('_', ' ')} - Model Performance"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Metric"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1

        data_ref = Reference(ws, min_col=2, min_row=1, max_row=4, max_col=2)
        cats_ref = Reference(ws, min_col=1, min_row=2, max_row=4)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 12
        chart.height = 8

        ws.add_chart(chart, "E2")

    wb.save(output_path)
    print(f"Saved Excel: {output_path.name}")

# Main
print("="*60)
print("CREATING PREDICTION MODEL METRIC BAR PLOTS")
print("="*60)

all_data = {}

for target_name, target_key in TARGETS.items():
    print(f"\nProcessing: {target_name}")
    metrics, values, errors = create_bar_plot(target_name, target_key, df, OUTPUT_DIR)
    all_data[target_name] = (metrics, values, errors)

# Create Excel
excel_path = OUTPUT_DIR / 'Prediction_Metrics_Bars.xlsx'
create_excel_with_chart(all_data, excel_path)

print("\n" + "="*60)
print("COMPLETE!")
print("="*60)
print(f"\nOutput directory: {OUTPUT_DIR}")
