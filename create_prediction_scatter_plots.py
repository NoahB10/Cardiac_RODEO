"""
Create prediction scatter plots for Arrhythmia, Heart Damage, and Concern.
Each target gets its own PNG and Excel sheet with embedded chart.
"""
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
from pathlib import Path

# Register Helvetica fonts from local fonts folder
_font_dir = Path(__file__).parent / 'fonts'
if _font_dir.exists():
    for _font_file in _font_dir.glob('*.ttf'):
        fm.fontManager.addfont(str(_font_file))
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Helvetica', 'Arial', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.label import DataLabelList

# Paths
PROJECT_ROOT = Path(__file__).parent
DATA_DIR = PROJECT_ROOT / 'Output' / 'Prediction_Scatter_Data'
OUTPUT_DIR = PROJECT_ROOT / 'Output' / 'Excel_Figures'
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Colors
COLORS = {
    'positive': '#2ecc71',  # Green for positive/true cases
    'negative': '#e74c3c',  # Red for negative/false cases
    'threshold': '#3498db'   # Blue for threshold line
}

CONCERN_COLORS = {
    'no': '#2ecc71',      # Green
    'less': '#f39c12',    # Orange
    'most': '#e74c3c'     # Red
}

def create_scatter_plot_binary(df, pred_col, actual_col, target_name, output_dir):
    """Create scatter plot for binary classification (Arrhythmia, Heart Damage)."""

    fig, ax = plt.subplots(figsize=(12, 6))

    drugs = df['Drug'].values
    preds = df[pred_col].values
    actuals = df[actual_col].values

    # Sort by prediction value
    sort_idx = np.argsort(preds)[::-1]
    drugs = drugs[sort_idx]
    preds = preds[sort_idx]
    actuals = actuals[sort_idx]

    positions = np.arange(len(drugs))

    # Color based on actual status
    colors = [COLORS['positive'] if a else COLORS['negative'] for a in actuals]

    ax.scatter(positions, preds, c=colors, s=80, alpha=0.8, edgecolors='black', linewidth=0.5)

    # Add threshold line (dynamic: slightly above highest negative)
    neg_preds = preds[~actuals]
    if len(neg_preds) > 0:
        threshold = np.max(neg_preds) + 2
    else:
        threshold = 50
    ax.axhline(y=threshold, color=COLORS['threshold'], linestyle='--', linewidth=2,
               label=f'Threshold ({threshold:.1f}%)')

    # Formatting
    ax.set_xticks(positions)
    ax.set_xticklabels(drugs, rotation=45, ha='right', fontsize=9)
    ax.set_xlabel('Drug', fontsize=12, fontweight='bold')
    ax.set_ylabel('Predicted Probability (%)', fontsize=12, fontweight='bold')
    ax.set_title(f'{target_name} Predictions', fontsize=14, fontweight='bold')
    ax.set_ylim(0, 105)
    ax.grid(True, axis='y', alpha=0.3)

    # Legend
    from matplotlib.patches import Patch
    legend_elements = [
        Patch(facecolor=COLORS['positive'], edgecolor='black', label=f'{target_name} = True'),
        Patch(facecolor=COLORS['negative'], edgecolor='black', label=f'{target_name} = False'),
    ]
    ax.legend(handles=legend_elements, loc='upper right')

    plt.tight_layout()

    # Save
    png_path = output_dir / f'{target_name}_Prediction_Scatter.png'
    fig.savefig(png_path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"Saved: {png_path.name}")

    return drugs, preds, actuals, threshold

def create_scatter_plot_concern(df, output_dir):
    """Create scatter plots for Concern (3 classes: No, Less, Most)."""

    # Create 3 separate plots, one for each class probability
    classes = [('No', 'Pred_No_pct'), ('Less', 'Pred_Less_pct'), ('Most', 'Pred_Most_pct')]

    all_data = {}

    for class_name, pred_col in classes:
        fig, ax = plt.subplots(figsize=(12, 6))

        drugs = df['Drug'].values
        preds = df[pred_col].values
        actuals = df['Actual_Concern'].values

        # Sort by prediction value
        sort_idx = np.argsort(preds)[::-1]
        drugs = drugs[sort_idx]
        preds = preds[sort_idx]
        actuals_sorted = actuals[sort_idx]

        positions = np.arange(len(drugs))

        # Color based on whether this class matches actual
        colors = [CONCERN_COLORS[a.lower()] for a in actuals_sorted]

        ax.scatter(positions, preds, c=colors, s=80, alpha=0.8, edgecolors='black', linewidth=0.5)

        # Formatting
        ax.set_xticks(positions)
        ax.set_xticklabels(drugs, rotation=45, ha='right', fontsize=9)
        ax.set_xlabel('Drug', fontsize=12, fontweight='bold')
        ax.set_ylabel(f'Predicted {class_name} Concern (%)', fontsize=12, fontweight='bold')
        ax.set_title(f'Concern: {class_name} - Predictions', fontsize=14, fontweight='bold')
        ax.set_ylim(0, 105)
        ax.grid(True, axis='y', alpha=0.3)

        # Legend
        from matplotlib.patches import Patch
        legend_elements = [
            Patch(facecolor=CONCERN_COLORS['no'], edgecolor='black', label='Actual: No'),
            Patch(facecolor=CONCERN_COLORS['less'], edgecolor='black', label='Actual: Less'),
            Patch(facecolor=CONCERN_COLORS['most'], edgecolor='black', label='Actual: Most'),
        ]
        ax.legend(handles=legend_elements, loc='upper right')

        plt.tight_layout()

        # Save
        png_path = output_dir / f'Concern_{class_name}_Prediction_Scatter.png'
        fig.savefig(png_path, dpi=300, bbox_inches='tight')
        plt.close(fig)
        print(f"Saved: {png_path.name}")

        all_data[class_name] = (drugs, preds, actuals_sorted)

    return all_data

def add_scatter_chart_to_sheet(ws, title, n_rows):
    """Add a scatter chart to a worksheet."""
    from openpyxl.chart import ScatterChart, Reference, Series
    from openpyxl.chart.marker import Marker

    chart = ScatterChart()
    chart.title = title
    chart.x_axis.title = "Drug Index"
    chart.y_axis.title = "Predicted Probability (%)"
    chart.y_axis.scaling.min = 0
    chart.y_axis.scaling.max = 100
    chart.style = 10
    chart.width = 15
    chart.height = 10
    chart.scatterStyle = 'marker'  # Markers only, no lines

    # X values (drug index 1, 2, 3...)
    x_values = Reference(ws, min_col=4, min_row=2, max_row=n_rows + 1)
    y_values = Reference(ws, min_col=2, min_row=2, max_row=n_rows + 1)

    series = Series(y_values, x_values, title="Predictions")
    series.marker = Marker(symbol='circle', size=7)
    chart.series.append(series)

    ws.add_chart(chart, "F2")

def create_excel_with_charts(arrhythmia_data, heart_damage_data, concern_data, output_path):
    """Create Excel file with separate sheets and embedded charts."""
    wb = Workbook()

    # Sheet 1: Arrhythmia
    ws = wb.active
    ws.title = "Arrhythmia"
    drugs, preds, actuals, threshold = arrhythmia_data
    n_rows = len(drugs)

    ws['A1'] = 'Drug'
    ws['B1'] = 'Predicted_Prob'
    ws['C1'] = 'Actual'
    ws['D1'] = 'Index'
    for i, (d, p, a) in enumerate(zip(drugs, preds, actuals), start=2):
        ws[f'A{i}'] = d
        ws[f'B{i}'] = p
        ws[f'C{i}'] = str(a)
        ws[f'D{i}'] = i - 1

    add_scatter_chart_to_sheet(ws, "Arrhythmia Predictions", n_rows)

    # Sheet 2: Heart Damage
    ws2 = wb.create_sheet("Heart_Damage")
    drugs, preds, actuals, threshold = heart_damage_data
    n_rows = len(drugs)

    ws2['A1'] = 'Drug'
    ws2['B1'] = 'Predicted_Prob'
    ws2['C1'] = 'Actual'
    ws2['D1'] = 'Index'
    for i, (d, p, a) in enumerate(zip(drugs, preds, actuals), start=2):
        ws2[f'A{i}'] = d
        ws2[f'B{i}'] = p
        ws2[f'C{i}'] = str(a)
        ws2[f'D{i}'] = i - 1

    add_scatter_chart_to_sheet(ws2, "Heart Damage Predictions", n_rows)

    # Sheets 3-5: Concern classes
    for class_name, (drugs, preds, actuals) in concern_data.items():
        ws_concern = wb.create_sheet(f"Concern_{class_name}")
        n_rows = len(drugs)

        ws_concern['A1'] = 'Drug'
        ws_concern['B1'] = f'Predicted_{class_name}_Prob'
        ws_concern['C1'] = 'Actual_Concern'
        ws_concern['D1'] = 'Index'
        for i, (d, p, a) in enumerate(zip(drugs, preds, actuals), start=2):
            ws_concern[f'A{i}'] = d
            ws_concern[f'B{i}'] = p
            ws_concern[f'C{i}'] = a
            ws_concern[f'D{i}'] = i - 1

        add_scatter_chart_to_sheet(ws_concern, f"Concern {class_name} Predictions", n_rows)

    wb.save(output_path)
    print(f"Saved Excel: {output_path.name}")

# Main
print("="*60)
print("CREATING PREDICTION SCATTER PLOTS")
print("="*60)

# Load data
df_arr = pd.read_csv(DATA_DIR / 'arrhythmia_predictions.csv')
df_hd = pd.read_csv(DATA_DIR / 'heart_damage_predictions.csv')
df_concern = pd.read_csv(DATA_DIR / 'concern_predictions.csv')

# Convert actual columns to boolean
df_arr['Actual_Arrhythmia'] = df_arr['Actual_Arrhythmia'].astype(str).str.lower() == 'true'
df_hd['Actual_Heart_Damage'] = df_hd['Actual_Heart_Damage'].astype(str).str.lower() == 'true'

print("\nCreating Arrhythmia scatter plot...")
arr_data = create_scatter_plot_binary(df_arr, 'Predicted_Arrhythmia_pct', 'Actual_Arrhythmia',
                                       'Arrhythmia', OUTPUT_DIR)

print("\nCreating Heart Damage scatter plot...")
hd_data = create_scatter_plot_binary(df_hd, 'Predicted_Heart_Damage_pct', 'Actual_Heart_Damage',
                                      'Heart_Damage', OUTPUT_DIR)

print("\nCreating Concern scatter plots...")
concern_data = create_scatter_plot_concern(df_concern, OUTPUT_DIR)

# Create Excel
print("\nCreating Excel file...")
excel_path = OUTPUT_DIR / 'Prediction_Scatter_Plots.xlsx'
create_excel_with_charts(arr_data, hd_data, concern_data, excel_path)

print("\n" + "="*60)
print("COMPLETE!")
print("="*60)
print(f"\nOutput directory: {OUTPUT_DIR}")
