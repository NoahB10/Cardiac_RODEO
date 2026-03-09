"""
Update Excel files with embedded charts for MoLFormer Comparison outputs.
Creates ONE Excel file per figure type, with sheets inside.
Following same format as update_admet_excel_charts.py.

Focus on:
- MoLFormer-XL-CNN (DIQT Transfer) - pretrained on 255 DIQT drugs, tested on 25
- MoLFormer-XL-CNN (5-fold CV on 25) - fine-tuned directly on 25 drugs
- Organoid RandomForest (5-fold CV) - functional PK-PD based prediction
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, ScatterChart, Reference
from openpyxl.chart import Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill
from sklearn.metrics import roc_curve, auc
import matplotlib
matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import seaborn as sns

# Register Helvetica fonts from local fonts folder
font_dir = Path(__file__).parent / 'fonts'
for font_file in font_dir.glob('*.ttf'):
    fm.fontManager.addfont(str(font_file))

# Set Helvetica as default font
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Helvetica', 'Arial', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

# Paths
PROJECT_ROOT = Path(__file__).parent
MOLFORMER_OUTPUT = PROJECT_ROOT / 'Output' / 'MoLFormer_Comparison'
EXCEL_OUTPUT = PROJECT_ROOT / 'Output' / 'Excel_Figures' / 'MoLFormer'
EXCEL_OUTPUT.mkdir(parents=True, exist_ok=True)

# Figure dimensions (height 1.72", width scaled)
FIG_HEIGHT = 1.72

print("="*60)
print("MoLFormer Comparison Excel Charts")
print("="*60)
print(f"Output folder: {EXCEL_OUTPUT}\n")

# =============================================================================
# Load Data
# =============================================================================
print("Loading data...")

# Metrics for all 3 models
metrics_df = pd.read_csv(MOLFORMER_OUTPUT / 'comparison_metrics_all.csv')
print(f"  Loaded comparison_metrics_all.csv ({len(metrics_df)} models)")

# CNN-25 5-fold CV out-of-fold predictions
cnn25_cv = pd.read_csv(MOLFORMER_OUTPUT / 'molformer_cnn_25drugs_cv.csv')
print(f"  Loaded molformer_cnn_25drugs_cv.csv ({len(cnn25_cv)} drugs)")

# Organoid 5-fold CV out-of-fold predictions
organoid_cv = pd.read_csv(MOLFORMER_OUTPUT / 'organoid_5fold_cv_predictions.csv')
print(f"  Loaded organoid_5fold_cv_predictions.csv ({len(organoid_cv)} drugs)")

# Per-drug predictions (CNN-DIQT transfer)
per_drug_df = pd.read_csv(MOLFORMER_OUTPUT / 'per_drug_predictions_all.csv')
print(f"  Loaded per_drug_predictions_all.csv ({len(per_drug_df)} drugs)")

# Merge CNN-25 CV predictions
per_drug_df = per_drug_df.merge(
    cnn25_cv[['Drug', 'CNN_25_prob']].rename(columns={'CNN_25_prob': 'CNN_25_CV_Prob'}),
    on='Drug', how='left'
)

# Merge Organoid 5-fold CV predictions
per_drug_df = per_drug_df.merge(
    organoid_cv[['Drug', 'Organoid_5fold_Prob']],
    on='Drug', how='left'
)
print()

# =============================================================================
# 1. Overall Comparison (Accuracy + AUC Bar Chart)
# =============================================================================
print("1. Creating Overall Comparison Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Model Comparison"

# Prepare data
comparison_df = metrics_df[['Model', 'Accuracy', 'ROC_AUC', 'Sensitivity', 'Specificity']].copy()
comparison_df.columns = ['Model', 'Accuracy', 'AUC', 'Sensitivity', 'Specificity']

# Shorten model names for display
comparison_df['Model'] = comparison_df['Model'].replace({
    'CNN (DIQT Transfer)': 'CNN-DIQT',
    'CNN (5-fold on 25)': 'CNN-25',
    'Organoid (5-fold)': 'Organoid'
})

# Write data
for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)

# Add Excel bar chart
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "MoLFormer vs Organoid: Arrhythmia Prediction"
chart.y_axis.title = "Score"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1

data = Reference(ws, min_col=2, min_row=1, max_row=len(comparison_df)+1, max_col=5)
cats = Reference(ws, min_col=1, min_row=2, max_row=len(comparison_df)+1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

chart.width = 14
chart.height = 10
ws.add_chart(chart, "G2")
print("  Added 'Model Comparison' sheet with bar chart")

# Create matplotlib version - grouped bar chart
fig, ax = plt.subplots(figsize=(2.5, FIG_HEIGHT))

models = comparison_df['Model'].tolist()
x_pos = np.arange(len(models))
width = 0.35

acc = comparison_df['Accuracy'].values
auc_vals = comparison_df['AUC'].values

# Colors: blue for Accuracy, red for AUC
bars1 = ax.bar(x_pos - width/2, acc, width, label='Accuracy', color='#3498db', edgecolor='black', linewidth=0.3)
bars2 = ax.bar(x_pos + width/2, auc_vals, width, label='AUC', color='#e74c3c', edgecolor='black', linewidth=0.3)

# Add value labels
for bar, val in zip(bars1, acc):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02, f'{val:.2f}',
            ha='center', va='bottom', fontsize=4)
for bar, val in zip(bars2, auc_vals):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.02, f'{val:.2f}',
            ha='center', va='bottom', fontsize=4)

ax.set_ylabel('Score', fontsize=6)
ax.set_title('Arrhythmia: Structure vs Functional', fontsize=7, fontweight='bold')
ax.set_xticks(x_pos)
ax.set_xticklabels(models, fontsize=5)
ax.set_ylim([0, 1.1])
ax.legend(fontsize=5, loc='upper left')
ax.tick_params(axis='y', labelsize=5)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
ax.axhline(y=0.5, color='gray', linestyle='--', linewidth=0.5, label='Random')

plt.tight_layout(pad=0.2)
plt.savefig(EXCEL_OUTPUT / 'overall_comparison.png', dpi=300, bbox_inches='tight')
plt.close()

wb.save(EXCEL_OUTPUT / 'Overall_Comparison.xlsx')
print("  Saved Overall_Comparison.xlsx\n")

# =============================================================================
# 2. Per-Drug Predictions (Scatter Plot)
# =============================================================================
print("2. Creating Per-Drug Predictions Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Drug Predictions"

# Sort by true label then by organoid probability
per_drug_sorted = per_drug_df.copy()
per_drug_sorted = per_drug_sorted.sort_values(['True_Arrhythmia', 'Organoid_5fold_Prob'], ascending=[False, False])

# Columns to write (use 5-fold CV predictions for both CNN-25 and Organoid)
cols = ['Drug', 'True_Arrhythmia', 'CNN_DIQT_Prob', 'CNN_25_CV_Prob', 'Organoid_5fold_Prob']
df_write = per_drug_sorted[cols].copy()
df_write.columns = ['Drug', 'True_Label', 'CNN_DIQT', 'CNN_25', 'Organoid']

for r_idx, row in enumerate(dataframe_to_rows(df_write, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)

# Color-code true labels
for r_idx in range(2, len(df_write)+2):
    true_val = ws.cell(row=r_idx, column=2).value
    if true_val == 1:
        ws.cell(row=r_idx, column=2).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        ws.cell(row=r_idx, column=2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Add scatter chart
chart = ScatterChart()
chart.style = 10
chart.title = "Per-Drug Arrhythmia Probability (5-fold CV)"
chart.x_axis.title = "Drug"
chart.y_axis.title = "Probability"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1

x_data = Reference(ws, min_col=1, min_row=2, max_row=len(df_write)+1)
y_diqt = Reference(ws, min_col=3, min_row=2, max_row=len(df_write)+1)
y_cnn25 = Reference(ws, min_col=4, min_row=2, max_row=len(df_write)+1)
y_org = Reference(ws, min_col=5, min_row=2, max_row=len(df_write)+1)

series_diqt = Series(y_diqt, x_data, title="CNN-DIQT")
series_cnn25 = Series(y_cnn25, x_data, title="CNN-25")
series_org = Series(y_org, x_data, title="Organoid")

chart.series.append(series_diqt)
chart.series.append(series_cnn25)
chart.series.append(series_org)

chart.width = 16
chart.height = 10
ws.add_chart(chart, "G2")
print("  Added 'Drug Predictions' sheet with scatter chart")

# Create matplotlib version
fig, ax = plt.subplots(figsize=(4.5, FIG_HEIGHT))

drugs = df_write['Drug'].tolist()
x_pos = np.arange(len(drugs))

cnn_diqt = df_write['CNN_DIQT'].values
cnn_25 = df_write['CNN_25'].values
organoid = df_write['Organoid'].values
true_labels = df_write['True_Label'].values

# Colors based on true label (green=arrhythmia+, red=arrhythmia-)
colors = ['#2ecc71' if t == 1 else '#e74c3c' for t in true_labels]

# Plot all three models
ax.scatter(x_pos, cnn_diqt, c='#3498db', s=12, marker='^', label='CNN-DIQT', alpha=0.7, zorder=4)
ax.scatter(x_pos, cnn_25, c='#9b59b6', s=12, marker='s', label='CNN-25', alpha=0.7, zorder=5)
ax.scatter(x_pos, organoid, c=colors, s=18, marker='o', edgecolors='black', linewidth=0.3, label='Organoid', zorder=6)

# Threshold line
ax.axhline(y=0.5, color='red', linestyle='--', linewidth=0.8)

ax.set_xticks(x_pos)
ax.set_xticklabels(drugs, rotation=90, ha='center', fontsize=3)
ax.set_ylabel('Probability', fontsize=5)
ax.set_ylim([0, 1.05])
ax.set_title('Per-Drug Predictions (5-fold CV)\n(green=Arr+, red=Arr-, threshold=0.5)', fontsize=5, fontweight='bold')
ax.tick_params(axis='y', labelsize=4)
ax.grid(alpha=0.3, axis='y', linewidth=0.3)
ax.legend(fontsize=3, loc='upper right', ncol=3)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

plt.tight_layout(pad=0.2)
plt.savefig(EXCEL_OUTPUT / 'per_drug_predictions.png', dpi=300, bbox_inches='tight')
plt.close()

wb.save(EXCEL_OUTPUT / 'Per_Drug_Predictions.xlsx')
print("  Saved Per_Drug_Predictions.xlsx\n")

# =============================================================================
# 3. ROC Comparison (all using 5-fold CV out-of-fold predictions)
# =============================================================================
print("3. Creating ROC Comparison Excel...")

wb = Workbook()
ws = wb.active
ws.title = "ROC Curves"

# Get true labels
y_true = per_drug_df['True_Arrhythmia'].values

# CNN-DIQT: transfer predictions (these ARE test predictions)
fpr_diqt, tpr_diqt, _ = roc_curve(y_true, per_drug_df['CNN_DIQT_Prob'].values)
auc_diqt = auc(fpr_diqt, tpr_diqt)

# CNN-25: 5-fold CV out-of-fold predictions
fpr_cnn25, tpr_cnn25, _ = roc_curve(y_true, per_drug_df['CNN_25_CV_Prob'].values)
auc_cnn25 = auc(fpr_cnn25, tpr_cnn25)

# Organoid: 5-fold CV out-of-fold predictions
fpr_org, tpr_org, _ = roc_curve(y_true, per_drug_df['Organoid_5fold_Prob'].values)
auc_org = auc(fpr_org, tpr_org)

print(f"  Computed ROC curves (all from 5-fold CV out-of-fold predictions):")
print(f"    CNN-DIQT: AUC = {auc_diqt:.3f} ({len(fpr_diqt)} points)")
print(f"    CNN-25:   AUC = {auc_cnn25:.3f} ({len(fpr_cnn25)} points)")
print(f"    Organoid: AUC = {auc_org:.3f} ({len(fpr_org)} points)")

# Interpolate all curves to common grid for Excel (100 points)
common_fpr = np.linspace(0, 1, 100)
tpr_diqt_interp = np.interp(common_fpr, fpr_diqt, tpr_diqt)
tpr_cnn25_interp = np.interp(common_fpr, fpr_cnn25, tpr_cnn25)
tpr_org_interp = np.interp(common_fpr, fpr_org, tpr_org)

# Create DataFrame for Excel
roc_df = pd.DataFrame({
    'FPR': common_fpr,
    'TPR_CNN_DIQT': tpr_diqt_interp,
    'TPR_CNN_25': tpr_cnn25_interp,
    'TPR_Organoid': tpr_org_interp
})

# Write to Excel
for r_idx, row in enumerate(dataframe_to_rows(roc_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)

# Add scatter chart for ROC
chart = ScatterChart()
chart.style = 10
chart.title = "ROC: Arrhythmia (5-fold CV)"
chart.x_axis.title = "False Positive Rate"
chart.y_axis.title = "True Positive Rate"
chart.x_axis.scaling.min = 0
chart.x_axis.scaling.max = 1
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 1

x_values = Reference(ws, min_col=1, min_row=2, max_row=len(roc_df)+1)

# CNN DIQT (blue)
y_diqt = Reference(ws, min_col=2, min_row=2, max_row=len(roc_df)+1)
s1 = Series(y_diqt, x_values, title=f"CNN-DIQT ({auc_diqt:.2f})")
chart.series.append(s1)
s1.graphicalProperties.line.solidFill = "3498db"
s1.graphicalProperties.line.width = 15000

# CNN 25 (purple)
y_cnn25 = Reference(ws, min_col=3, min_row=2, max_row=len(roc_df)+1)
s2 = Series(y_cnn25, x_values, title=f"CNN-25 ({auc_cnn25:.2f})")
chart.series.append(s2)
s2.graphicalProperties.line.solidFill = "9b59b6"
s2.graphicalProperties.line.width = 15000

# Organoid (green)
y_org = Reference(ws, min_col=4, min_row=2, max_row=len(roc_df)+1)
s3 = Series(y_org, x_values, title=f"Organoid ({auc_org:.2f})")
chart.series.append(s3)
s3.graphicalProperties.line.solidFill = "2ecc71"
s3.graphicalProperties.line.width = 20000

chart.width = 12
chart.height = 12
ws.add_chart(chart, "F2")
print("  Added 'ROC Curves' sheet")

# Add AUC summary sheet
ws2 = wb.create_sheet(title="AUC Summary")
ws2.cell(row=1, column=1, value="Model").font = Font(bold=True)
ws2.cell(row=1, column=2, value="AUC").font = Font(bold=True)
ws2.cell(row=1, column=3, value="Validation").font = Font(bold=True)
ws2.cell(row=2, column=1, value="CNN (DIQT Transfer)")
ws2.cell(row=2, column=2, value=round(auc_diqt, 4))
ws2.cell(row=2, column=3, value="Transfer (test on 25)")
ws2.cell(row=3, column=1, value="CNN (5-fold on 25)")
ws2.cell(row=3, column=2, value=round(auc_cnn25, 4))
ws2.cell(row=3, column=3, value="5-fold CV out-of-fold")
ws2.cell(row=4, column=1, value="Organoid (5-fold)")
ws2.cell(row=4, column=2, value=round(auc_org, 4))
ws2.cell(row=4, column=3, value="5-fold CV out-of-fold")

# Create matplotlib ROC plot
fig, ax = plt.subplots(figsize=(1.72, FIG_HEIGHT))

# Plot all as step functions (actual predictions)
ax.step(fpr_diqt, tpr_diqt, where='post', color='#3498db', linewidth=1,
        label=f'CNN-DIQT ({auc_diqt:.2f})')

ax.step(fpr_cnn25, tpr_cnn25, where='post', color='#9b59b6', linewidth=1,
        label=f'CNN-25 ({auc_cnn25:.2f})')

ax.step(fpr_org, tpr_org, where='post', color='#2ecc71', linewidth=1.5,
        label=f'Organoid ({auc_org:.2f})')

# Random classifier
ax.plot([0, 1], [0, 1], 'k--', linewidth=0.5, label='Random')

ax.set_xlim([-0.02, 1.02])
ax.set_ylim([-0.02, 1.02])
ax.set_xlabel('FPR', fontsize=6)
ax.set_ylabel('TPR', fontsize=6)
ax.set_title('ROC Comparison (5-fold CV)', fontsize=7, fontweight='bold')
ax.legend(loc='lower right', fontsize=4, framealpha=0.9)
ax.tick_params(axis='both', labelsize=5)
ax.grid(alpha=0.3, linewidth=0.3)
ax.set_aspect('equal')

plt.tight_layout()
plt.savefig(EXCEL_OUTPUT / 'roc_comparison.png', dpi=300, bbox_inches='tight')
plt.close()

wb.save(EXCEL_OUTPUT / 'ROC_Comparison.xlsx')
print("  Saved ROC_Comparison.xlsx\n")

# =============================================================================
# 4. Confusion Matrices
# =============================================================================
print("4. Creating Confusion Matrices Excel...")

wb = Workbook()

# Get confusion matrix values from metrics
for model_row in metrics_df.itertuples():
    model_name = model_row.Model.replace(' ', '_').replace('(', '').replace(')', '').replace('-', '_')

    # Create sheet
    ws = wb.create_sheet(title=model_name[:30])  # Excel sheet name limit

    ws.cell(row=1, column=1, value=model_row.Model).font = Font(bold=True, size=12)

    # Header
    ws.cell(row=3, column=1, value="").font = Font(bold=True)
    ws.cell(row=3, column=2, value="Pred Neg").font = Font(bold=True)
    ws.cell(row=3, column=3, value="Pred Pos").font = Font(bold=True)

    ws.cell(row=4, column=1, value="True Neg").font = Font(bold=True)
    ws.cell(row=5, column=1, value="True Pos").font = Font(bold=True)

    # Values
    ws.cell(row=4, column=2, value=model_row.TN)
    ws.cell(row=4, column=3, value=model_row.FP)
    ws.cell(row=5, column=2, value=model_row.FN)
    ws.cell(row=5, column=3, value=model_row.TP)

    # Metrics
    ws.cell(row=7, column=1, value="Accuracy:")
    ws.cell(row=7, column=2, value=f"{model_row.Accuracy:.3f}")
    ws.cell(row=8, column=1, value="Sensitivity:")
    ws.cell(row=8, column=2, value=f"{model_row.Sensitivity:.3f}")
    ws.cell(row=9, column=1, value="Specificity:")
    ws.cell(row=9, column=2, value=f"{model_row.Specificity:.3f}")
    ws.cell(row=10, column=1, value="AUC:")
    ws.cell(row=10, column=2, value=f"{model_row.ROC_AUC:.3f}")

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

print("  Added confusion matrix sheets for all 3 models")

# Create matplotlib confusion matrices
fig, axes = plt.subplots(1, 3, figsize=(4.5, FIG_HEIGHT))

for ax, model_row in zip(axes, metrics_df.itertuples()):
    # Build confusion matrix
    cm = np.array([[model_row.TN, model_row.FP],
                   [model_row.FN, model_row.TP]])

    # Plot
    sns.heatmap(cm, annot=True, fmt='d', cmap='Blues', ax=ax, cbar=False,
                annot_kws={'size': 6})

    # Short name
    short_name = model_row.Model.replace('CNN (DIQT Transfer)', 'CNN-DIQT')
    short_name = short_name.replace('CNN (5-fold on 25)', 'CNN-25')
    short_name = short_name.replace('Organoid (5-fold)', 'Organoid')

    ax.set_title(f'{short_name}\nAcc={model_row.Accuracy:.2f}', fontsize=5, fontweight='bold')
    ax.set_xlabel('Pred', fontsize=4)
    ax.set_ylabel('True', fontsize=4)
    ax.set_xticklabels(['Neg', 'Pos'], fontsize=4)
    ax.set_yticklabels(['Neg', 'Pos'], fontsize=4, rotation=0)

plt.tight_layout(pad=0.3)
plt.savefig(EXCEL_OUTPUT / 'confusion_matrices.png', dpi=300, bbox_inches='tight')
plt.close()

wb.save(EXCEL_OUTPUT / 'Confusion_Matrices.xlsx')
print("  Saved Confusion_Matrices.xlsx\n")

# =============================================================================
# 5. Summary Comparison Table
# =============================================================================
print("5. Creating Summary Table Excel...")

wb = Workbook()
ws = wb.active
ws.title = "Summary"

# Full comparison table
summary_df = metrics_df[['Model', 'Accuracy', 'ROC_AUC', 'TP', 'FN', 'TN', 'FP',
                         'Sensitivity', 'Specificity']].copy()

# Add interpretation column
def interpret_model(row):
    if 'DIQT' in row['Model']:
        return 'Structure-based (transfer from 255 DIQT drugs)'
    elif '5-fold on 25' in row['Model']:
        return 'Structure-based (fine-tuned on 25 drugs)'
    else:
        return 'Functional (PK-PD coefficients from organoids)'

summary_df['Approach'] = summary_df.apply(interpret_model, axis=1)

# Reorder columns
summary_df = summary_df[['Model', 'Approach', 'Accuracy', 'ROC_AUC', 'Sensitivity',
                          'Specificity', 'TP', 'FN', 'TN', 'FP']]

for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 1):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)

# Adjust column widths
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 10

# Add key findings
ws.cell(row=len(summary_df)+3, column=1, value="Key Findings:").font = Font(bold=True)
ws.cell(row=len(summary_df)+4, column=1,
        value="1. Organoid-based functional prediction outperforms structure-based MoLFormer")
ws.cell(row=len(summary_df)+5, column=1,
        value=f"2. CNN trained on 25 drugs ({auc_cnn25:.2f} AUC) > CNN transfer from DIQT ({auc_diqt:.2f} AUC)")
ws.cell(row=len(summary_df)+6, column=1,
        value=f"3. Organoid ({auc_org:.2f} AUC) > CNN-25 ({auc_cnn25:.2f} AUC) > CNN-DIQT ({auc_diqt:.2f} AUC)")

wb.save(EXCEL_OUTPUT / 'Summary_Table.xlsx')
print("  Saved Summary_Table.xlsx\n")

# =============================================================================
# Summary
# =============================================================================
print("="*60)
print("MoLFormer Comparison Excel Charts Complete!")
print("="*60)
print(f"\nCreated {len(list(EXCEL_OUTPUT.glob('*.xlsx')))} Excel files:")
for f in sorted(EXCEL_OUTPUT.glob('*.xlsx')):
    print(f"  - {f.name}")
print(f"\nCreated {len(list(EXCEL_OUTPUT.glob('*.png')))} PNG images:")
for f in sorted(EXCEL_OUTPUT.glob('*.png')):
    print(f"  - {f.name}")
print(f"\nOutput folder: {EXCEL_OUTPUT}")
print("\nAll models use comparable validation:")
print("  - CNN-DIQT: Transfer predictions (trained on 255 DIQT, tested on 25)")
print("  - CNN-25: 5-fold CV out-of-fold predictions")
print("  - Organoid: 5-fold CV out-of-fold predictions")
