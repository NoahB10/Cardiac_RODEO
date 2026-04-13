"""
Creates a color-scaled Excel for the 3 supplementary Vandetanib heatmaps
(O2_std, O2_dom_freq, Amp_dom_freq) so user can manually delete outliers.
Also applies LOWESS w=8 preview so outliers are easier to spot.
"""
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from statsmodels.nonparametric.smoothers_lowess import lowess

PROJECT_ROOT = Path("C:/Users/NoahB/Documents/HebrewU Bioengineering/Cardiac_RODEO")
HEATMAP_DIR  = PROJECT_ROOT / 'Cleaned_Data' / 'Heatmaps' / 'Vandetanib (G11)'
OUT = Path(__file__).parent / 'S1_outlier_editor.xlsx'

LOWESS_WINDOW = 8


def apply_lowess_per_column(df, window=LOWESS_WINDOW):
    smoothed = df.copy().astype(float)
    for col in smoothed.columns:
        series = smoothed[col]
        valid = series.dropna()
        if len(valid) < 3:
            continue
        first_idx = valid.index[0]
        first_val = valid.iloc[0]
        to_smooth = valid.iloc[1:]
        if len(to_smooth) < 2:
            continue
        frac = min(1.0, max(window, 1) / len(to_smooth))
        fitted = lowess(to_smooth.values, np.arange(len(to_smooth)),
                        frac=frac, return_sorted=False)
        target = smoothed.index.get_indexer(to_smooth.index)
        smoothed.iloc[target, smoothed.columns.get_loc(col)] = fitted
        smoothed.at[first_idx, col] = first_val
    return smoothed


wb = Workbook()

sheets = [
    ('O2_std',       'O2_std.csv'),
    ('O2_dom_freq',  'O2_dom_freq.csv'),
    ('Amp_dom_freq', 'Amp_dom_freq.csv'),
]

for sheet_name, csv_name in sheets:
    df_raw = pd.read_csv(HEATMAP_DIR / csv_name, index_col=0).astype(float)
    # Apply LOWESS w=8 first so outliers stand out against smooth background
    df = apply_lowess_per_column(df_raw)

    ws = wb.create_sheet(sheet_name)
    ws.cell(1, 1, 'Time(h) / Well')
    for ci, col in enumerate(df.columns, 2):
        ws.cell(1, ci, str(col))
    for ri, (idx, row) in enumerate(df.iterrows(), 2):
        ws.cell(ri, 1, idx)
        for ci, val in enumerate(row, 2):
            ws.cell(ri, ci, None if np.isnan(val) else round(float(val), 4))

    n_rows = len(df) + 1
    n_cols = len(df.columns) + 1
    data_range = f"B2:{get_column_letter(n_cols)}{n_rows}"
    rule = ColorScaleRule(
        start_type='min',  start_color='123BFF',
        mid_type='percentile', mid_value=50, mid_color='FFFFFF',
        end_type='max',   end_color='FF2908'
    )
    ws.conditional_formatting.add(data_range, rule)
    ws.freeze_panes = 'B2'
    ws.column_dimensions['A'].width = 10
    print(f"Sheet '{sheet_name}': LOWESS w={LOWESS_WINDOW} applied, {df.shape}")

del wb['Sheet']
wb.save(OUT)
print(f"Saved: {OUT}")
