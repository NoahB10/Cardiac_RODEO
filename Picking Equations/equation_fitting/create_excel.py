"""
Create Excel Workbook from Coefficient Files

Generates all_equations_coefficients.xlsx with one sheet per equation,
matching the expected format for the prediction models.

Format matches example file:
- Sheet names: snake_case (e.g., pkpd_elimination)
- Column 2: heart_damage (not Cardiotoxicity)
- No separator column between Contractility and O2
- Drugs sorted by Concern level (most → less → no)
"""
import pandas as pd
import numpy as np
import re
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import COEFF_DIR, FINAL_EXCEL, EQUATION_NAMES, EQUATIONS, CLEANED_DATA

# =============================================================================
# STYLING
# =============================================================================

HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
GROUP_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

# =============================================================================
# CLASSIFICATION DATA
# =============================================================================

def load_classification():
    """
    Load drug classification from saved CSV or example file.

    Returns dict: {drug_name_lower: {'Arrhythmia': bool, 'heart_damage': bool, 'Concern': str}}
    """
    # Try loading from Cleaned_Data first (centralized location)
    class_file = CLEANED_DATA / 'drug_classification.csv'

    # Fall back to COEFF_DIR for backwards compatibility
    if not class_file.exists():
        class_file = COEFF_DIR / 'drug_classification.csv'

    if class_file.exists():
        df = pd.read_csv(class_file)
    else:
        # Fall back to example file
        example_path = Path(__file__).parent.parent / 'all_equations_coefficients_example.xlsx'
        if example_path.exists():
            df = pd.read_excel(example_path, sheet_name='pkpd_elimination', header=1)
            df = df[['Drug', 'Arrhythmia', 'heart_damage', 'Concern']].copy()
        else:
            return {}

    # Create lookup dict
    classification = {}
    for _, row in df.iterrows():
        drug_key = normalize_drug_name(row['Drug'])
        classification[drug_key] = {
            'Arrhythmia': row['Arrhythmia'],
            'heart_damage': row['heart_damage'],
            'Concern': row['Concern']
        }
    return classification


def normalize_drug_name(name):
    """Normalize drug name for matching."""
    name_str = str(name).lower()
    name_str = re.sub(r'\s|\(.*?\)', '', name_str)
    return name_str


def sort_drugs_by_concern(drug_list, classification):
    """
    Sort drugs by Concern level: most → less → no, then alphabetically.
    """
    concern_order = {'most': 0, 'less': 1, 'no': 2}

    def sort_key(drug):
        drug_key = normalize_drug_name(drug)
        class_info = classification.get(drug_key, {})
        concern = str(class_info.get('Concern', 'no')).lower()
        concern_rank = concern_order.get(concern, 3)
        return (concern_rank, drug.lower())

    return sorted(drug_list, key=sort_key)


# =============================================================================
# EXCEL CREATION
# =============================================================================

def create_all_equations_excel():
    """
    Create the all_equations_coefficients.xlsx workbook.

    Format matches example file:
    - Sheet names: snake_case (e.g., pkpd_elimination)
    - Column 2: heart_damage (not Cardiotoxicity)
    - No separator column between Contractility and O2
    - Drugs sorted by Concern level (most → less → no)
    """
    print("\n" + "="*80)
    print("CREATING EXCEL WORKBOOK (EXAMPLE FORMAT)")
    print("="*80 + "\n")

    # Load classification data
    classification = load_classification()
    print(f"  Loaded classification for {len(classification)} drugs")

    wb = Workbook()
    wb.remove(wb.active)

    sheets_created = 0

    for eq_name in EQUATION_NAMES:
        eq_info = EQUATIONS.get(eq_name, {})
        param_cols = eq_info.get('params', [])

        contract_file = COEFF_DIR / f"{eq_name}_coefficients_contractility.csv"
        o2_file = COEFF_DIR / f"{eq_name}_coefficients_o2.csv"

        if not contract_file.exists() and not o2_file.exists():
            print(f"  Skipping: {eq_name} (no CSV files)")
            continue

        print(f"  Adding sheet: {eq_name}")

        # Create sheet with snake_case name
        ws = wb.create_sheet(title=eq_name[:31])

        # Load data
        df_c = pd.read_csv(contract_file) if contract_file.exists() else pd.DataFrame()
        df_o = pd.read_csv(o2_file) if o2_file.exists() else pd.DataFrame()

        # Normalize drug names for matching
        if not df_c.empty:
            df_c['DrugKey'] = df_c['Drug'].apply(normalize_drug_name)
        if not df_o.empty:
            df_o['DrugKey'] = df_o['Drug'].apply(normalize_drug_name)

        # Get all unique drug names (using original case from Contractility first)
        drug_name_map = {}  # drug_key -> original name
        if not df_c.empty:
            for _, row in df_c.iterrows():
                drug_name_map[row['DrugKey']] = row['Drug']
        if not df_o.empty:
            for _, row in df_o.iterrows():
                if row['DrugKey'] not in drug_name_map:
                    drug_name_map[row['DrugKey']] = row['Drug']

        # Get list of drug names sorted by concern
        drug_names = list(drug_name_map.values())
        sorted_drugs = sort_drugs_by_concern(drug_names, classification)

        # Calculate column positions
        n_params = len(param_cols)
        # Contractility: 3 classification + n_params + 3 metrics (Cmax_used, R2, N_points)
        # O2: n_params + 3 metrics
        # NO separator column

        contract_param_start = 5  # After Drug(1) + Ar(2) + HD(3) + Concern(4)
        contract_metric_start = contract_param_start + n_params
        o2_start = contract_metric_start + 3
        o2_param_start = o2_start
        o2_metric_start = o2_param_start + n_params
        total_cols = o2_metric_start + 3 - 1

        # Row 1: Group headers
        ws.cell(row=1, column=1, value="Drug").font = HEADER_FONT
        ws.cell(row=1, column=2, value="Contractility").font = HEADER_FONT
        ws.cell(row=1, column=2).fill = GROUP_FILL

        # Merge Contractility header (columns 2 through o2_start-1)
        if o2_start > 2:
            ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=o2_start - 1)

        # O2 group header
        ws.cell(row=1, column=o2_start, value="O2").font = HEADER_FONT
        ws.cell(row=1, column=o2_start).fill = GROUP_FILL

        # Merge O2 header
        if total_cols >= o2_start:
            ws.merge_cells(start_row=1, start_column=o2_start, end_row=1, end_column=total_cols)

        # Row 2: Column names
        col_idx = 1
        ws.cell(row=2, column=col_idx, value="Drug").font = HEADER_FONT
        col_idx += 1

        # Classification columns (use heart_damage, not Cardiotoxicity)
        for name in ['Arrhythmia', 'heart_damage', 'Concern']:
            ws.cell(row=2, column=col_idx, value=name).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            col_idx += 1

        # Contractility parameter columns
        for param in param_cols:
            ws.cell(row=2, column=col_idx, value=param).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            col_idx += 1

        # Contractility metrics
        for metric in ['Cmax_used', 'R2', 'N_points']:
            ws.cell(row=2, column=col_idx, value=metric).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            col_idx += 1

        # O2 parameter columns (NO separator)
        for param in param_cols:
            ws.cell(row=2, column=col_idx, value=param).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            col_idx += 1

        # O2 metrics
        for metric in ['Cmax_used', 'R2', 'N_points']:
            ws.cell(row=2, column=col_idx, value=metric).font = HEADER_FONT
            ws.cell(row=2, column=col_idx).fill = HEADER_FILL
            col_idx += 1

        # Data rows (sorted by concern)
        row_idx = 3
        for drug_name in sorted_drugs:
            col_idx = 1
            drug_key = normalize_drug_name(drug_name)

            # Get data rows
            row_c = None
            row_o = None
            if not df_c.empty and drug_key in df_c['DrugKey'].values:
                row_c = df_c[df_c['DrugKey'] == drug_key].iloc[0]
            if not df_o.empty and drug_key in df_o['DrugKey'].values:
                row_o = df_o[df_o['DrugKey'] == drug_key].iloc[0]

            # Drug name
            ws.cell(row=row_idx, column=col_idx, value=drug_name)
            col_idx += 1

            # Classification values from loaded data
            class_info = classification.get(drug_key, {})

            # Arrhythmia
            arr_val = class_info.get('Arrhythmia', '')
            if pd.notna(arr_val):
                ws.cell(row=row_idx, column=col_idx, value=bool(arr_val) if isinstance(arr_val, (bool, np.bool_)) else arr_val)
            col_idx += 1

            # heart_damage
            hd_val = class_info.get('heart_damage', '')
            if pd.notna(hd_val):
                ws.cell(row=row_idx, column=col_idx, value=bool(hd_val) if isinstance(hd_val, (bool, np.bool_)) else hd_val)
            col_idx += 1

            # Concern
            concern_val = class_info.get('Concern', '')
            if pd.notna(concern_val):
                ws.cell(row=row_idx, column=col_idx, value=str(concern_val))
            col_idx += 1

            # Contractility parameters
            for param in param_cols:
                if row_c is not None and param in row_c.index:
                    val = row_c[param]
                    if pd.notna(val):
                        ws.cell(row=row_idx, column=col_idx, value=float(val))
                col_idx += 1

            # Contractility metrics
            for metric in ['Cmax_used', 'R2', 'N_points']:
                if row_c is not None and metric in row_c.index:
                    val = row_c[metric]
                    if pd.notna(val):
                        ws.cell(row=row_idx, column=col_idx, value=float(val))
                col_idx += 1

            # O2 parameters (NO separator)
            for param in param_cols:
                if row_o is not None and param in row_o.index:
                    val = row_o[param]
                    if pd.notna(val):
                        ws.cell(row=row_idx, column=col_idx, value=float(val))
                col_idx += 1

            # O2 metrics
            for metric in ['Cmax_used', 'R2', 'N_points']:
                if row_o is not None and metric in row_o.index:
                    val = row_o[metric]
                    if pd.notna(val):
                        ws.cell(row=row_idx, column=col_idx, value=float(val))
                col_idx += 1

            row_idx += 1

        # Adjust column widths
        for col_num in range(1, total_cols + 1):
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 12

        sheets_created += 1

    if sheets_created == 0:
        ws = wb.create_sheet(title="No Data")
        ws.cell(row=1, column=1, value="No coefficient files found")

    FINAL_EXCEL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(FINAL_EXCEL)

    print(f"\nSaved: {FINAL_EXCEL}")
    print(f"Sheets created: {sheets_created}/{len(EQUATION_NAMES)}")

    return FINAL_EXCEL


if __name__ == "__main__":
    create_all_equations_excel()
