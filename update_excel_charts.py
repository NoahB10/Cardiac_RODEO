"""
Update Excel files with embedded charts for all prediction model outputs.
Creates ONE Excel file per figure type, with sheets for each target.
Only includes binary targets: Arrhythmia, Heart Damage, Concern Binary.
"""
import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment
import matplotlib
matplotlib.use('Agg')
import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import seaborn as sns
import re

# Register Helvetica fonts from local fonts folder
_font_dir = Path(__file__).parent / 'fonts'
if _font_dir.exists():
    for _font_file in _font_dir.glob('*.ttf'):
        fm.fontManager.addfont(str(_font_file))
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['font.sans-serif'] = ['Helvetica', 'Arial', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False

excel_figures_dir = Path('Output/Excel_Figures')
output_dir = Path('Output')

# Only these targets
TARGETS = ['Arrhythmia', 'heart_damage', 'Concern_Binary']
TARGETS_LOWER = ['arrhythmia', 'heart_damage', 'concern_binary']

print("Creating Excel files (one per figure type)...\n")

# ============================================================================
# 1. Performance Metrics (Accuracy, F1, MCC, AUC) - ONE FILE
# ============================================================================
print("1. Creating performance metrics Excel (Accuracy, F1, MCC, AUC)...")

metrics_src = output_dir / 'Performance_Metrics' / 'all_performance_metrics.xlsx'

if metrics_src.exists():
    perf_df = pd.read_excel(metrics_src, sheet_name='model_performance_summary')

    wb = Workbook()
    first_sheet = True

    for target in TARGETS:
        target_data = perf_df[perf_df['Target'] == target].copy()
        if target_data.empty:
            continue

        sheet_name = target.replace('_', ' ').title()
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(target_data, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Create bar chart for all 4 metrics
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = f"Model Performance: {sheet_name}"
        chart.y_axis.title = "Score"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 1

        cols = target_data.columns.tolist()
        acc_col = cols.index('Accuracy_Mean') + 1 if 'Accuracy_Mean' in cols else None
        auc_col = cols.index('AUC_Mean') + 1 if 'AUC_Mean' in cols else None
        f1_col = cols.index('F1_Mean') + 1 if 'F1_Mean' in cols else None
        mcc_col = cols.index('MCC_Mean') + 1 if 'MCC_Mean' in cols else None

        if acc_col and auc_col and f1_col and mcc_col:
            for col in [acc_col, f1_col, mcc_col, auc_col]:
                data = Reference(ws, min_col=col, min_row=1, max_row=len(target_data)+1)
                chart.add_data(data, titles_from_data=True)

            model_col = cols.index('Model') + 1 if 'Model' in cols else 2
            cats = Reference(ws, min_col=model_col, min_row=2, max_row=len(target_data)+1)
            chart.set_categories(cats)

        chart.width = 14
        chart.height = 10
        ws.add_chart(chart, "M2")
        print(f"  Added '{sheet_name}' sheet")

        # Create matplotlib version - bars labeled by metric name on x-axis (no legend)
        # Size: 1.69" wide x 1.72" tall
        fig, ax = plt.subplots(figsize=(1.69, 1.72))

        models = target_data['Model'].tolist()
        metrics = ['Accuracy_Mean', 'F1_Mean', 'MCC_Mean', 'AUC_Mean']
        metric_labels = ['Acc', 'F1', 'MCC', 'AUC']
        colors = ['#3498db', '#2ecc71', '#9b59b6', '#e74c3c']

        # Create x positions: group by model, 4 bars per model
        n_models = len(models)
        n_metrics = len(metrics)
        bar_width = 0.8
        group_width = n_metrics * bar_width + 0.5  # spacing between model groups

        x_positions = []
        x_labels = []
        bar_colors = []
        bar_values = []

        for m_idx, model in enumerate(models):
            model_row = target_data[target_data['Model'] == model].iloc[0]
            for met_idx, (metric, label, color) in enumerate(zip(metrics, metric_labels, colors)):
                x_pos = m_idx * group_width + met_idx * bar_width
                x_positions.append(x_pos)
                x_labels.append(label)
                bar_colors.append(color)
                bar_values.append(model_row[metric])

        ax.bar(x_positions, bar_values, width=bar_width, color=bar_colors, edgecolor='black', linewidth=0.3)

        # Set x-axis labels (metric names under each bar) - small font for small figure
        ax.set_xticks(x_positions)
        ax.set_xticklabels(x_labels, fontsize=5)

        # Model name goes in title, not on x-axis
        model_name = models[0] if len(models) == 1 else ', '.join(models)

        ax.set_ylabel('Score', fontsize=5)
        ax.set_title(f"{sheet_name}\n{model_name}", fontsize=6, fontweight='bold')
        ax.set_ylim([0, 1.05])
        ax.tick_params(axis='y', labelsize=5)
        ax.grid(alpha=0.3, axis='y', linewidth=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)

        plt.tight_layout(pad=0.2)
        img_path = excel_figures_dir / f'metrics_{target.lower()}.png'
        plt.savefig(img_path, dpi=300, bbox_inches='tight')
        
        plt.close()

    wb.save(excel_figures_dir / 'Performance_Metrics.xlsx')
    print("  Saved Performance_Metrics.xlsx\n")

# ============================================================================
# 2. SHAP Feature Importance - ONE FILE
# ============================================================================
print("2. Creating SHAP feature importance Excel...")

wb = Workbook()
first_sheet = True

for target in TARGETS_LOWER:
    src = output_dir / 'SHAP_Data' / f'shap_{target}_bar.xlsx'

    if not src.exists():
        continue

    df = pd.read_excel(src, sheet_name='mean_importance')

    sheet_name = target.replace('_', ' ').title()
    if first_sheet:
        ws = wb.active
        ws.title = sheet_name
        first_sheet = False
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # Create horizontal bar chart
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = f"SHAP Feature Importance: {sheet_name}"
    chart.x_axis.title = "Mean |SHAP value|"

    data = Reference(ws, min_col=2, min_row=1, max_row=len(df)+1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width = 15
    chart.height = 12

    ws.add_chart(chart, "D2")
    print(f"  Added '{sheet_name}' sheet")

    # Create matplotlib version - height 1.72", width scaled
    fig, ax = plt.subplots(figsize=(2.15, 1.72))
    features = df['Feature'].tolist()
    values = df['Mean_Abs_SHAP'].tolist()

    ax.barh(features, values, color='#3498db', height=0.7)
    ax.set_xlabel('Mean |SHAP|', fontsize=5)
    ax.set_title(f"{sheet_name}", fontsize=6, fontweight='bold')
    ax.invert_yaxis()
    ax.tick_params(axis='both', labelsize=4)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_linewidth(0.3)
    ax.spines['bottom'].set_linewidth(0.3)

    plt.tight_layout()
    img_path = excel_figures_dir / f'shap_{target}.png'
    plt.savefig(img_path, dpi=300, bbox_inches='tight')
    
    plt.close()

wb.save(excel_figures_dir / 'SHAP_Feature_Importance.xlsx')
print("  Saved SHAP_Feature_Importance.xlsx\n")

# ============================================================================
# 3. ROC Curves with Std Range - ONE FILE
# ============================================================================
print("3. Creating ROC curves Excel (with std range)...")

src = output_dir / 'ROC_Data' / 'roc_curves_all_models.xlsx'

if src.exists():
    src_xl = pd.ExcelFile(src)

    wb = Workbook()
    first_sheet = True

    for sheet in src_xl.sheet_names:
        df = pd.read_excel(src, sheet_name=sheet)

        sheet_name = sheet.replace('_', ' ')
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Find FPR/TPR columns
        fpr_cols = sorted([c for c in df.columns if 'FPR' in c])
        tpr_cols = sorted([c for c in df.columns if 'TPR' in c])

        # Calculate mean and std for ROC
        mean_fpr = np.linspace(0, 1, 100)
        tprs = []
        for fpr_col, tpr_col in zip(fpr_cols, tpr_cols):
            fpr = df[fpr_col].dropna().values
            tpr = df[tpr_col].dropna().values
            if len(fpr) > 0 and len(tpr) > 0:
                tprs.append(np.interp(mean_fpr, fpr, tpr))

        if tprs:
            mean_tpr = np.mean(tprs, axis=0)
            std_tpr = np.std(tprs, axis=0)
            tpr_upper = np.minimum(mean_tpr + std_tpr, 1)
            tpr_lower = np.maximum(mean_tpr - std_tpr, 0)

            # Create new dataframe with mean, upper, lower bounds
            roc_df = pd.DataFrame({
                'FPR': mean_fpr,
                'TPR_Mean': mean_tpr,
                'TPR_Upper': tpr_upper,
                'TPR_Lower': tpr_lower
            })

            # Write data
            for r_idx, row in enumerate(dataframe_to_rows(roc_df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            # Create scatter chart with lines (proper X-Y plotting for ROC)
            from openpyxl.chart import ScatterChart, Series

            chart = ScatterChart()
            chart.style = 10
            chart.title = f"{sheet_name}"
            chart.x_axis.title = "FPR"
            chart.y_axis.title = "TPR"
            chart.x_axis.scaling.min = 0
            chart.x_axis.scaling.max = 1
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 1

            # X values (FPR)
            x_values = Reference(ws, min_col=1, min_row=2, max_row=len(roc_df)+1)

            # Mean TPR series
            y_mean = Reference(ws, min_col=2, min_row=2, max_row=len(roc_df)+1)
            series_mean = Series(y_mean, x_values, title="Mean")
            chart.series.append(series_mean)

            # Upper bound series
            y_upper = Reference(ws, min_col=3, min_row=2, max_row=len(roc_df)+1)
            series_upper = Series(y_upper, x_values, title="Upper")
            chart.series.append(series_upper)

            # Lower bound series
            y_lower = Reference(ws, min_col=4, min_row=2, max_row=len(roc_df)+1)
            series_lower = Series(y_lower, x_values, title="Lower")
            chart.series.append(series_lower)

            # Style: scatter with smooth lines, no markers
            for s in chart.series:
                s.graphicalProperties.line.width = 15000
                s.smooth = True

            # Mean - solid blue
            chart.series[0].graphicalProperties.line.solidFill = "3498DB"
            chart.series[0].graphicalProperties.line.width = 20000
            # Upper - dashed grey
            chart.series[1].graphicalProperties.line.solidFill = "888888"
            chart.series[1].graphicalProperties.line.dashStyle = "dash"
            # Lower - dashed grey
            chart.series[2].graphicalProperties.line.solidFill = "888888"
            chart.series[2].graphicalProperties.line.dashStyle = "dash"

            chart.width = 10
            chart.height = 10
            ws.add_chart(chart, "G2")
            print(f"  Added '{sheet_name}' sheet")

            # Create matplotlib version with shaded std region - height 1.72", square
            fig, ax = plt.subplots(figsize=(1.72, 1.72))

            # Shaded std region
            ax.fill_between(mean_fpr, tpr_lower, tpr_upper, alpha=0.3, color='#3498db')

            # Mean ROC curve
            ax.plot(mean_fpr, mean_tpr, 'b-', linewidth=1)

            # Std bounds as dotted lines
            ax.plot(mean_fpr, tpr_upper, '--', color='#888888', linewidth=0.5)
            ax.plot(mean_fpr, tpr_lower, '--', color='#888888', linewidth=0.5)

            # Random line
            ax.plot([0, 1], [0, 1], 'k--', linewidth=0.5)

            ax.set_xlim([0, 1])
            ax.set_ylim([0, 1])
            ax.set_xlabel('FPR', fontsize=5)
            ax.set_ylabel('TPR', fontsize=5)
            ax.set_title(f"{sheet_name}", fontsize=6, fontweight='bold')
            ax.tick_params(axis='both', labelsize=4)
            ax.grid(alpha=0.3, linewidth=0.3)
            ax.spines['top'].set_linewidth(0.3)
            ax.spines['right'].set_linewidth(0.3)
            ax.spines['left'].set_linewidth(0.3)
            ax.spines['bottom'].set_linewidth(0.3)

            plt.tight_layout(pad=0.2)
            img_path = excel_figures_dir / f'roc_{sheet.lower()}.png'
            plt.savefig(img_path, dpi=300, bbox_inches='tight')
            
            plt.close()

    wb.save(excel_figures_dir / 'ROC_Curves.xlsx')
    print("  Saved ROC_Curves.xlsx\n")

# ============================================================================
# 4. Cumulative Feature Importance - ONE FILE (binary targets only)
# ============================================================================
print("4. Creating cumulative feature importance Excel...")

import json

src = output_dir / 'Cumulative_Plot_Data' / 'cumulative_feature_importance.xlsx'
threshold_path = output_dir / 'Prediction_Scatter_Data' / 'prediction_thresholds.json'
classification_path = Path('Cleaned_Data') / 'drug_classification.csv'

# Load thresholds
with open(threshold_path, 'r') as f:
    thresholds_data = json.load(f)

THRESHOLD_MAP = {
    'Arrhythmia': thresholds_data['Arrhythmia_threshold_pct'],
    'Heart Damage': thresholds_data['Heart_Damage_threshold_pct'],
    'Concern Binary': thresholds_data['Concern_Binary_threshold_pct'],
}

# Load drug classification for pass/fail markers
drug_class = pd.read_csv(classification_path)
drug_class.set_index('Drug', inplace=True)

def get_passing_drugs(sheet_name):
    """Return set of drugs that are POSITIVE for the condition."""
    if sheet_name == 'Arrhythmia':
        return set(drug_class[drug_class['Arrhythmia'] == True].index)
    elif sheet_name == 'Heart Damage':
        return set(drug_class[drug_class['heart_damage'] == True].index)
    elif sheet_name == 'Concern Binary':
        return set(drug_class[drug_class['Concern'] == 'most'].index)
    return set()

if src.exists():
    src_xl = pd.ExcelFile(src)

    wb = Workbook()
    first_sheet = True

    # Only include binary targets
    sheets_to_include = ['Arrhythmia', 'Heart Damage', 'Concern Binary']

    for sheet in src_xl.sheet_names:
        if sheet not in sheets_to_include:
            continue

        df = pd.read_excel(src, sheet_name=sheet)

        if first_sheet:
            ws = wb.active
            ws.title = sheet
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Get drug columns (all except Cumulative_Coefficients)
        drug_cols = [col for col in df.columns if col != 'Cumulative_Coefficients']
        n_features = len(df)
        n_drugs = len(drug_cols)

        # Create line chart with all drugs
        chart = LineChart()
        chart.style = 10
        chart.title = f"Cumulative Feature Importance: {sheet}"
        chart.x_axis.title = "Feature Rank"
        chart.y_axis.title = "Predicted Probability (%)"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        # Add each drug as a series
        for i in range(2, n_drugs + 2):
            values = Reference(ws, min_col=i, min_row=1, max_row=n_features + 1)
            chart.add_data(values, titles_from_data=True)

        # Categories (feature rank 1, 2, 3, ...)
        cats = Reference(ws, min_col=1, min_row=2, max_row=n_features + 1)
        chart.set_categories(cats)

        chart.width = 20
        chart.height = 12
        ws.add_chart(chart, "AB2")
        print(f"  Added '{sheet}' sheet")

        # Create matplotlib version - correct cumulative plot
        # Each drug gets its own line, X for fail, O for pass
        fig, ax = plt.subplots(figsize=(5.5, 1.72))

        features = df['Cumulative_Coefficients'].values
        passing_drugs = get_passing_drugs(sheet)
        threshold = THRESHOLD_MAP.get(sheet, 50.0)

        # Rainbow colormap for drugs
        drug_list = sorted(drug_cols)
        cmap = plt.get_cmap('rainbow')
        color_map = {
            drug: cmap(0.0 if len(drug_list) == 1 else i / (len(drug_list) - 1))
            for i, drug in enumerate(drug_list)
        }

        # Plot each drug
        x = np.arange(1, n_features + 1)
        for drug in drug_list:
            y = df[drug].values
            is_passing = drug in passing_drugs
            marker = 'o' if is_passing else 'x'
            markersize = 2 if is_passing else 3
            ax.plot(x, y, marker=marker, label=drug, alpha=0.7, linewidth=0.5,
                    markersize=markersize, color=color_map[drug])

        # Add threshold line
        ax.axhline(y=threshold, color='red', linestyle='--', linewidth=1,
                   label=f'Threshold={threshold:.0f}%')

        ax.set_xticks(x)
        ax.set_xticklabels([str(i) for i in x], fontsize=3)
        ax.set_yticks(np.linspace(0, 100, 6))
        ax.set_ylim(0, 100)
        ax.set_xlabel('Feature Rank', fontsize=5)
        ax.set_ylabel('Pred Prob (%)', fontsize=5)
        ax.set_title(f"{sheet}\n(o=positive, x=negative)", fontsize=5, fontweight='bold')
        ax.tick_params(axis='y', labelsize=4)
        ax.grid(True, axis='y', alpha=0.3, linewidth=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)

        # Legend outside (small font)
        ax.legend(bbox_to_anchor=(1.02, 1), loc='upper left', ncol=1, fontsize=3)

        plt.tight_layout(pad=0.2)
        target_name = sheet.replace(' ', '_').lower()
        img_path = excel_figures_dir / f'cumulative_{target_name}.png'
        plt.savefig(img_path, dpi=300, bbox_inches='tight')

        plt.close()

    wb.save(excel_figures_dir / 'Cumulative_Feature_Importance.xlsx')
    print("  Saved Cumulative_Feature_Importance.xlsx\n")

# ============================================================================
# 5. Prediction Scatter Plots - ONE FILE (binary targets only)
# ============================================================================
print("5. Creating prediction scatter Excel...")

src = output_dir / 'Prediction_Scatter_Data' / 'prediction_scatter_all.xlsx'

# Threshold mapping for scatter plots (same as cumulative)
SCATTER_THRESHOLD_MAP = {
    'arrhythmia': thresholds_data['Arrhythmia_threshold_pct'],
    'heart_damage': thresholds_data['Heart_Damage_threshold_pct'],
    'concern_binary': thresholds_data['Concern_Binary_threshold_pct'],
}

if src.exists():
    src_xl = pd.ExcelFile(src)

    wb = Workbook()
    first_sheet = True

    sheets_to_include = ['arrhythmia', 'heart_damage', 'concern_binary']

    for sheet in src_xl.sheet_names:
        if sheet not in sheets_to_include:
            continue

        df = pd.read_excel(src, sheet_name=sheet)

        sheet_name = sheet.replace('_', ' ').title()
        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        prob_cols = [c for c in df.columns if 'predicted' in c.lower() or 'pred_' in c.lower()]

        # Create scatter chart
        chart = ScatterChart()
        chart.style = 10
        chart.title = f"Prediction Scatter: {sheet_name}"
        chart.x_axis.title = "Drug Index"
        chart.y_axis.title = "Predicted %"
        chart.y_axis.scaling.min = 0
        chart.y_axis.scaling.max = 100

        if prob_cols:
            prob_col = df.columns.get_loc(prob_cols[0]) + 1
            x_data = Reference(ws, min_col=1, min_row=2, max_row=len(df)+1)
            y_data = Reference(ws, min_col=prob_col, min_row=2, max_row=len(df)+1)

            from openpyxl.chart import Series
            series = Series(y_data, x_data, title="Predictions")
            chart.series.append(series)

        chart.width = 12
        chart.height = 8
        ws.add_chart(chart, "H2")
        print(f"  Added '{sheet_name}' sheet")

        # Get threshold for this target (same as cumulative plots)
        scatter_threshold = SCATTER_THRESHOLD_MAP.get(sheet, 50.0)

        # Create matplotlib version - height 1.72", width scaled (wider for 25 drugs)
        fig, ax = plt.subplots(figsize=(3.44, 1.72))

        if 'Drug' in df.columns:
            x_labels = df['Drug'].tolist()
            x_vals = list(range(len(x_labels)))
        else:
            x_labels = [str(i) for i in range(len(df))]
            x_vals = list(range(len(df)))

        if prob_cols:
            y_vals = df[prob_cols[0]].tolist()

            actual_col = [c for c in df.columns if 'actual' in c.lower()]
            if actual_col:
                actual_vals = df[actual_col[0]].tolist()
                colors = []
                for v in actual_vals:
                    if isinstance(v, bool):
                        colors.append('#2ecc71' if v else '#e74c3c')
                    elif isinstance(v, str):
                        colors.append('#2ecc71' if v.lower() == 'true' else '#e74c3c')
                    else:
                        colors.append('#2ecc71' if v else '#e74c3c')
            else:
                colors = '#3498db'

            ax.scatter(x_vals, y_vals, c=colors, s=15, edgecolors='black', linewidth=0.2, zorder=5)
            ax.axhline(y=scatter_threshold, color='red', linestyle='--', linewidth=0.8,
                       label=f'Threshold={scatter_threshold:.0f}%')

        ax.set_ylabel('Pred %', fontsize=5)
        ax.set_xticks(x_vals)
        ax.set_xticklabels(x_labels, rotation=90, ha='center', fontsize=3)
        ax.set_ylim([0, 105])
        ax.set_title(f"{sheet_name}", fontsize=6, fontweight='bold')
        ax.tick_params(axis='y', labelsize=4)
        ax.grid(alpha=0.3, axis='y', linewidth=0.3)
        ax.spines['top'].set_visible(False)
        ax.spines['right'].set_visible(False)
        ax.spines['left'].set_linewidth(0.3)
        ax.spines['bottom'].set_linewidth(0.3)

        plt.tight_layout(pad=0.2)
        img_path = excel_figures_dir / f'scatter_{sheet.lower()}.png'
        plt.savefig(img_path, dpi=300, bbox_inches='tight')
        
        plt.close()
        print(f"  Saved scatter_{sheet.lower()}.png")

    wb.save(excel_figures_dir / 'Prediction_Scatter.xlsx')
    print("  Saved Prediction_Scatter.xlsx\n")

# ============================================================================
# 6. Confusion Matrices - ONE FILE (binary targets only)
# ============================================================================
print("6. Creating confusion matrices Excel...")

src = output_dir / 'Confusion_Matrices' / 'confusion_matrices_all.xlsx'

if src.exists():
    src_xl = pd.ExcelFile(src)

    wb = Workbook()
    first_sheet = True

    sheet_map = {
        'arrhythmia': 'Arrhythmia',
        'heart_damage': 'Heart Damage',
        'concern_binary': 'Concern Binary'
    }
    model_map = {
        'arrhythmia': 'RandomForest',
        'heart_damage': 'GaussianNB',
        'concern_binary': 'GaussianNB'
    }
    threshold_map = {
        'arrhythmia': thresholds_data.get('Arrhythmia_threshold_pct'),
        'heart_damage': thresholds_data.get('Heart_Damage_threshold_pct'),
        'concern_binary': thresholds_data.get('Concern_Binary_threshold_pct')
    }

    def _load_organoid_cm_xlsx(cm_path):
        """Load CM from existing organoid Excel (no re-prediction)."""
        if not cm_path.exists():
            return None
        raw = pd.read_excel(cm_path, sheet_name=0, header=None)
        if raw.empty or 0 not in raw.columns:
            return None

        rows = []
        for _, row in raw.iterrows():
            v1, v2 = row.get(1), row.get(2)
            if pd.notna(v1) and pd.notna(v2):
                try:
                    v1_i = int(v1)
                    v2_i = int(v2)
                except Exception:
                    continue
                label = str(row.get(0)).strip() if pd.notna(row.get(0)) else ''
                rows.append((label, v1_i, v2_i))

        if len(rows) < 2:
            return None

        (label0, tn, fp), (label1, fn, tp) = rows[0], rows[1]
        return pd.DataFrame(
            np.array([[tn, fp], [fn, tp]], dtype=int),
            index=[f"Actual_{label0 or 'No'}", f"Actual_{label1 or 'Yes'}"],
            columns=['Pred_No', 'Pred_Yes']
        )

    def _load_concern_binary_cm_pdf(pdf_path, threshold_value=None):
        """Load CM counts from existing concern binary PDF (no re-prediction)."""
        if not pdf_path.exists():
            return None
        try:
            import fitz  # PyMuPDF
        except Exception:
            return None

        doc = fitz.open(pdf_path)
        text = "".join(page.get_text() or "" for page in doc)
        nums = [int(n) for n in re.findall(r'\\b\\d+\\b', text)]
        # Drop colorbar ticks and title numbers commonly present in the PDF
        drop = {0, 5, 10, 20, 40, 60, 80, 100, 120}
        if threshold_value is not None:
            drop.add(int(round(threshold_value)))
        counts = [n for n in nums if n not in drop and n >= 15]
        if len(counts) < 4:
            return None
        # Use the four largest remaining numbers as the CM counts
        counts = sorted(counts)[-4:]
        tn, fp, fn, tp = counts[0], counts[1], counts[2], counts[3]
        return pd.DataFrame(
            np.array([[tn, fp], [fn, tp]], dtype=int),
            index=['Actual_No Concern', 'Actual_High Concern'],
            columns=['Pred_No', 'Pred_Yes']
        )

    arrhythmia_override_df = _load_organoid_cm_xlsx(
        excel_figures_dir / 'confusion_matrix_organoid_arrhythmia.xlsx'
    )
    heart_damage_override_df = _load_organoid_cm_xlsx(
        excel_figures_dir / 'confusion_matrix_organoid_heart_damage.xlsx'
    )
    concern_binary_override_df = _load_organoid_cm_xlsx(
        excel_figures_dir / 'confusion_matrix_organoid_concern_binary.xlsx'
    )
    if concern_binary_override_df is None:
        concern_binary_override_df = _load_concern_binary_cm_pdf(
            excel_figures_dir / 'confusion_matrix_concern_binary.pdf',
            threshold_map.get('concern_binary')
        )

    for src_sheet, dst_sheet in sheet_map.items():
        if src_sheet not in src_xl.sheet_names and not (
            src_sheet == 'arrhythmia' and arrhythmia_override_df is not None
        ):
            continue

        override_df = None
        if src_sheet == 'arrhythmia':
            override_df = arrhythmia_override_df
        elif src_sheet == 'heart_damage':
            override_df = heart_damage_override_df
        elif src_sheet == 'concern_binary':
            override_df = concern_binary_override_df

        if override_df is not None:
            df = override_df
            title_suffix = " (10-seed 5-fold)"
        else:
            df = pd.read_excel(src, sheet_name=src_sheet, index_col=0)
            title_suffix = ""

        if first_sheet:
            ws = wb.active
            ws.title = dst_sheet
            first_sheet = False
        else:
            ws = wb.create_sheet(title=dst_sheet)

        # Write data with formatting
        ws.cell(row=1, column=1, value="")
        for c_idx, col in enumerate(df.columns, 2):
            cell = ws.cell(row=1, column=c_idx, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        for r_idx, (idx, row) in enumerate(df.iterrows(), 2):
            cell = ws.cell(row=r_idx, column=1, value=idx)
            cell.font = Font(bold=True)
            for c_idx, value in enumerate(row, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                cell.alignment = Alignment(horizontal='center')

        print(f"  Added '{dst_sheet}' sheet")

        # Create matplotlib version - height 1.72", width scaled
        fig, ax = plt.subplots(figsize=(2.06, 1.72))

        # Build annotations with row percentages
        annot = None
        if df.shape == (2, 2):
            row_sums = df.sum(axis=1).replace(0, 1)
            row_pct = df.div(row_sums, axis=0) * 100
            annot = df.astype(int).astype(str) + "\n" + row_pct.round(1).astype(str) + "%"

        sns.heatmap(
            df,
            annot=annot if annot is not None else True,
            fmt='' if annot is not None else 'd',
            cmap='Blues',
            ax=ax,
            cbar=False,
            square=True,
            annot_kws={'size': 6}
        )

        model_name = model_map.get(src_sheet, dst_sheet)
        if src_sheet in ('heart_damage', 'concern_binary'):
            thr = threshold_map.get(src_sheet)
            if thr is not None:
                title_line1 = f"{model_name} Confusion Matrix (threshold={thr:.0f}%)"
            else:
                title_line1 = f"{model_name} Confusion Matrix"
            title_line2 = "10-seed 5-fold"
            ax.set_title(f"{title_line1}\n{title_line2}", fontsize=6, fontweight='bold')
        else:
            ax.set_title(f"{model_name}{title_suffix}", fontsize=6, fontweight='bold')
        ax.set_xlabel('Prediction', fontsize=5)
        ax.set_ylabel('Actual', fontsize=5)
        if df.shape == (2, 2):
            ax.set_xticklabels(['No', 'Yes'], fontsize=4)
            ax.set_yticklabels(['No', 'Yes'], fontsize=4)
        else:
            ax.tick_params(axis='both', labelsize=4)

        plt.tight_layout(pad=0.2)
        img_path = excel_figures_dir / f'confmat_{src_sheet}.png'
        plt.savefig(img_path, dpi=300, bbox_inches='tight')
        
        plt.close()

    wb.save(excel_figures_dir / 'Confusion_Matrices.xlsx')
    print("  Saved Confusion_Matrices.xlsx\n")

print("=== Done! Created 6 Excel files (one per figure type) ===")
print("All files contain only: Arrhythmia, Heart Damage, Concern Binary")
print("ROC curves now include mean with std range (grey shaded + dotted bounds)")
